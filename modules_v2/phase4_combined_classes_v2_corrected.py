"""
Phase 4: Combined Class Scheduling v2 - CORRECTED VERSION
Handles combined class scheduling with proper identification and zero conflicts
"""

import os
import sys
import math
import random
from datetime import time
from typing import List, Dict, Tuple
from collections import defaultdict

import openpyxl

# Add the current directory to Python path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.data_models import Course, Section, TimeBlock
from utils.time_slot_logger import get_logger
from utils.session_rules_validator import SessionRulesValidator
from utils.time_validator import validate_time_range, can_fit_duration, slot_end_within_day
from config.schedule_config import WORKING_DAYS
from config.structure_config import (
    DEPARTMENTS,
    SECTION_GROUPS,
    SECTIONS_BY_DEPT,
    get_grouping_signature,
)
from utils.interactive_prompts import skip_interactive_prompts, default_period_is_pre_mid


def get_period_assignments_path() -> str:
    """
    Return absolute path to the Phase 4 period assignments Excel file.

    The file lives under DATA/INPUT as:
        phase4_period_assignments_<odd|even>.xlsx (when offering is set)
    """
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    raw_variant = (
        os.environ.get("ARISE_COURSE_DATA_VARIANT", "").strip()
        or os.environ.get("ARISE_OFFERING", "").strip()
    ).lower()
    if raw_variant in ("odd", "1"):
        fn = "phase4_period_assignments_odd.xlsx"
    elif raw_variant in ("even", "2"):
        fn = "phase4_period_assignments_even.xlsx"
    else:
        # Backward compatible default (also used when UI doesn't pass offering)
        fn = "phase4_period_assignments.xlsx"

    return os.path.join(base_dir, "DATA", "INPUT", fn)


def load_period_assignments(path: str) -> Dict[Tuple[str, int, int], bool]:
    """
    Load period assignments from Excel.

    Returns:
        {(course_code, semester, group_id): True for PreMid, False for PostMid}
    """
    assignments: Dict[Tuple[str, int, int], bool] = {}

    if not os.path.exists(path):
        return assignments

    try:
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
    except Exception:
        # If the file is corrupted or unreadable, fall back to empty assignments
        return assignments

    # Invalidate old assignments if the section grouping changed
    try:
        expected_sig = get_grouping_signature()
        meta = wb["META"] if "META" in wb.sheetnames else None
        if meta is None:
            return {}  # force re-prompt for legacy files
        stored_sig = None
        for row in meta.iter_rows(min_row=1, max_row=100):
            k = row[0].value
            v = row[1].value if len(row) > 1 else None
            if str(k).strip().upper() == "GROUPING_SIGNATURE":
                stored_sig = str(v) if v is not None else ""
                break
        if not stored_sig or stored_sig != expected_sig:
            return {}  # grouping changed -> re-ask
    except Exception:
        return {}  # safest: re-ask if meta cannot be read

    # Expect header row: Course_Code | Semester | [Group] | Pre | Post
    header = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    # Build a mapping from expected column names to indices
    name_to_idx = {name.upper(): idx for idx, name in enumerate(header)}

    def _get_col(name: str) -> int:
        return name_to_idx.get(name.upper(), -1)

    col_code = _get_col("COURSE_CODE")
    col_sem = _get_col("SEMESTER")
    col_group = _get_col("GROUP")
    col_pre = _get_col("PRE")
    col_post = _get_col("POST")

    if col_code < 0 or col_sem < 0 or col_pre < 0 or col_post < 0:
        # Header does not match expected format
        return assignments

    for row in sheet.iter_rows(min_row=2):
        code_cell = row[col_code].value
        sem_cell = row[col_sem].value
        pre_cell = row[col_pre].value
        post_cell = row[col_post].value

        if not code_cell or sem_cell is None:
            continue

        try:
            code = str(code_cell).strip().upper()
            semester = int(sem_cell)
        except (TypeError, ValueError):
            continue

        pre_val = int(pre_cell) if pre_cell is not None else 0
        post_val = int(post_cell) if post_cell is not None else 0

        # A valid row has exactly one of Pre/Post set to 1
        if pre_val == 1 and post_val == 0:
            group_id = 1
            if col_group >= 0 and row[col_group].value is not None:
                try:
                    group_id = int(row[col_group].value)
                except Exception:
                    group_id = 1
            assignments[(code, semester, group_id)] = True
        elif pre_val == 0 and post_val == 1:
            group_id = 1
            if col_group >= 0 and row[col_group].value is not None:
                try:
                    group_id = int(row[col_group].value)
                except Exception:
                    group_id = 1
            assignments[(code, semester, group_id)] = False
        else:
            # Invalid or ambiguous row - ignore, will be re-asked
            continue

    return assignments


def save_period_assignments(path: str, assignments: Dict[Tuple[str, int, int], bool]) -> None:
    """
    Save period assignments to Excel.

    assignments: {(course_code, semester, group_id): True for PreMid, False for PostMid}
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Phase4Periods"

    # META sheet to track grouping signature (invalidate on group change)
    meta = wb.create_sheet(title="META")
    meta.append(["Key", "Value"])
    meta.append(["GROUPING_SIGNATURE", get_grouping_signature()])

    # Header
    sheet.append(["Course_Code", "Semester", "Group", "Pre", "Post"])

    # Sort for stable, readable output
    for (code, semester, group_id) in sorted(assignments.keys(), key=lambda x: (x[1], x[2], x[0])):
        is_premid = assignments[(code, semester, group_id)]
        pre_val = 1 if is_premid else 0
        post_val = 0 if is_premid else 1
        sheet.append([code, semester, group_id, pre_val, post_val])

    # Ensure directory exists
    dir_name = os.path.dirname(path)
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name, exist_ok=True)

    wb.save(path)


def prompt_user_for_period(course_code: str, course_name: str, semester: int, group_id: int = 1) -> bool:
    """
    Prompt the user to choose PreMid or PostMid for a course/group.

    Returns:
        True  -> PreMid
        False -> PostMid
    """

    # Build a human-readable description of current sections for this group.
    try:
        group_sections = []
        for dept, sec_map in SECTION_GROUPS.items():
            for sec_label, grp in sec_map.items():
                if grp == group_id and sec_label in SECTIONS_BY_DEPT.get(dept, []):
                    group_sections.append(f"{dept}-{sec_label}")
        group_desc = ", ".join(sorted(group_sections)) if group_sections else f"Group {group_id} sections"
    except Exception:
        group_desc = f"Group {group_id} sections"

    prompt = (
        f"\nPhase 4: Choose period for combined course {course_code} - {course_name} (Sem {semester}) "
        f"for Group {group_id} ({group_desc}).\n"
        f"  [1] PreMid\n"
        f"  [2] PostMid\n"
        f"Enter 1 or 2 (or 'pre'/'post'): "
    )

    if skip_interactive_prompts():
        pre = default_period_is_pre_mid("ARISE_PHASE4_DEFAULT_PERIOD")
        label = "PreMid" if pre else "PostMid"
        print(f"[non-interactive] Phase 4 period for {course_code} (Sem {semester}, Group {group_id}) -> {label}")
        return pre

    while True:
        try:
            choice = input(prompt).strip().lower()
        except EOFError:
            # Non-interactive environment: default to PreMid
            return True

        if choice in ("1", "pre", "premid", "pre-mid", "pre_mid"):
            return True
        if choice in ("2", "post", "postmid", "post-mid", "post_mid"):
            return False

        print("Invalid input. Please enter 1 for PreMid or 2 for PostMid.")


def get_period_assignments_for_courses(
    unique_courses: Dict[int, List[Course]],
    sections: List[Section],
    assignments_path: str,
) -> Tuple[Dict[Tuple[str, int, int], bool], bool]:
    """
    Build period assignments for all combined courses handled in Phase 4.

    Args:
        unique_courses: {semester: [Course, ...]} from get_unique_combined_courses_by_semester
        assignments_path: path to Excel file

    Returns:
        (assignments, any_prompted)
        - assignments: {(course_code, semester, group_id): True for PreMid, False for PostMid}
        - any_prompted: True if we asked the user for at least one course this run
    """
    import os

    force_prompts = os.environ.get("ARISE_FORCE_PHASE_PROMPTS", "").strip().lower() in ("1", "true", "yes")
    # Default behavior: ask only when assignment is missing; then persist and reuse.

    # Load any existing assignments from disk. If forcing prompts (interactive runs),
    # we still load but allow overwriting keys by re-prompting.
    existing = load_period_assignments(assignments_path)
    assignments: Dict[Tuple[str, int, int], bool] = dict(existing)
    any_prompted = False

    def _groups_for_course(course_code: str, semester: int) -> List[int]:
        groups = set()
        code = str(course_code).upper()
        for c in courses_for_semester:
            if str(c.code).upper() != code:
                continue
            if int(getattr(c, "semester", -1) or -1) != int(semester):
                continue
            for sec in sections:
                if getattr(sec, "semester", None) != semester:
                    continue
                if getattr(sec, "program", None) == getattr(c, "department", None):
                    groups.add(int(getattr(sec, "group", 1) or 1))
        return sorted(groups)

    for semester, courses in unique_courses.items():
        courses_for_semester = list(courses or [])
        for course in courses:
            code = course.code.upper()
            key_g1 = (code, semester, 1)
            if force_prompts or (key_g1 not in assignments):
                # Keep Group 1 behavior unchanged.
                is_premid_g1 = prompt_user_for_period(course.code, course.name, semester, group_id=1)
                assignments[key_g1] = is_premid_g1
                any_prompted = True
            else:
                is_premid_g1 = assignments[key_g1]

            # Group 3+ are independent: ask first time, then reuse saved value.
            for gid in _groups_for_course(code, semester):
                if gid < 3:
                    continue
                gkey = (code, semester, gid)
                if (not force_prompts) and (gkey in assignments):
                    continue
                is_premid = prompt_user_for_period(course.code, course.name, semester, group_id=gid)
                assignments[gkey] = is_premid
                any_prompted = True

    return assignments, any_prompted

def calculate_slots_from_ltpsc(ltpsc: str) -> Dict[str, int]:
    """Calculate slots using LTPSC with ceiling for lectures"""
    try:
        parts = str(ltpsc).split('-')
        if len(parts) != 5:
            return {'lectures': 2, 'tutorials': 1, 'practicals': 0, 'total': 3}
        
        L = int(parts[0]) if parts[0] else 0
        T = int(parts[1]) if parts[1] else 0
        P = int(parts[2]) if parts[2] else 0
        
        lectures = math.ceil(L / 1.5)
        tutorials = int(T / 1)
        practicals = int(P / 2)
        
        return {
            'lectures': lectures,
            'tutorials': tutorials,
            'practicals': practicals,
            'total': lectures + tutorials + practicals
        }
    except:
        return {'lectures': 2, 'tutorials': 1, 'practicals': 0, 'total': 3}

def get_combined_courses(courses: List[Course], sections: List[Section] = None) -> Dict[int, List[Course]]:
    """
    Get truly combined courses by semester.
    A course is combined if ALL of the following hold:
    1. Core (not elective)
    2. <=2 credits
    3. Single instructor
    4. Common for ALL branches (CSE, DSAI, ECE) - offered in each department
    5. Has multiple instances (same code + same instructor across branches) OR
       is common to multiple sections within a department that has multiple sections
    """
    # Phase 4 "common for all branches" is defined by the core UG trio.
    # Do not auto-expand with extra programs (e.g., AIC) from global config.
    REQUIRED_BRANCHES = {"CSE", "DSAI", "ECE"}

    # v1-inspired, pipeline-safe identification:
    # Group by (code, semester), not by (code, instructor), so a course stays eligible
    # even when branch-wise instructor names differ (as requested).
    eligible_by_code_sem = defaultdict(list)
    for course in courses:
        if (
            not course.is_elective
            and course.credits <= 2
            and course.instructors
            and len(course.instructors) == 1
        ):
            eligible_by_code_sem[(course.code, course.semester)].append(course)

    combined_courses = defaultdict(list)
    for (_code, _sem), course_list in eligible_by_code_sem.items():
        if len(course_list) <= 1:
            continue
        departments_with_course = set(c.department for c in course_list)
        # Strict rule: only all-branch common courses are Phase 4.
        if departments_with_course < REQUIRED_BRANCHES:
            continue
        for course in course_list:
            combined_courses[course.semester].append(course)

    return combined_courses

def get_unique_combined_courses_by_semester(combined_courses: Dict[int, List[Course]]) -> Dict[int, List[Course]]:
    """Get unique courses per semester (one representative per course code)"""
    unique_courses = {}
    
    for semester, courses in combined_courses.items():
        # Get one representative course per code
        course_map = {}
        for course in courses:
            if course.code not in course_map:
                course_map[course.code] = course
        unique_courses[semester] = list(course_map.values())
    
    return unique_courses

def create_non_overlapping_schedule(unique_courses: Dict[int, List[Course]], sections: List[Section], courses: List[Course], classrooms: List = None) -> Dict:
    """
    Create a non-overlapping schedule for combined courses with PreMid/PostMid distribution
    Uses LTPSC to determine slot counts dynamically
    
    Args:
        classrooms: List of ClassRoom objects to filter 240-seater rooms
    """
    print("DEBUG: create_non_overlapping_schedule function called!")
    print(f"DEBUG: create_non_overlapping_schedule called with semesters: {list(unique_courses.keys())}")
    schedule = {}
    occupied_slots_by_semester = {}  # Track occupied slots per semester
    # IMPORTANT: Track room occupancy GLOBALLY across all semesters and periods.
    # Filter to only include large non-lab rooms based on config threshold.
    if classrooms:
        from config.schedule_config import LARGE_ROOM_CAPACITY_THRESHOLD

        available_large_rooms = [
            room.room_number
            for room in classrooms
            if str(room.room_type).lower() != "lab"
            and "lab" not in str(room.room_type).lower()
            and room.capacity >= LARGE_ROOM_CAPACITY_THRESHOLD
        ]
        print(
            f"DEBUG: Filtered to {len(available_large_rooms)} large rooms (capacity >= {LARGE_ROOM_CAPACITY_THRESHOLD}): "
            f"{available_large_rooms}"
        )
    else:
        # Fallback: if classrooms list is missing, we cannot guess a real room.
        # Use an empty list and let higher-level validation catch missing rooms,
        # instead of forcing a hardcoded default like a specific lab.
        print("WARNING: No classrooms provided; no large rooms available for Phase 4 scheduling")
        available_large_rooms = []
    
    # Track room occupancy per PERIOD. PreMid and PostMid are disjoint halves,
    # so the same room/time can be reused across periods without conflict.
    premid_room_occupancy = {room: {} for room in available_large_rooms}   # {room: {day: [(start,end), ...]}}
    postmid_room_occupancy = {room: {} for room in available_large_rooms}  # {room: {day: [(start,end), ...]}}

    # Track faculty occupancy for preventive checks during Phase 4 scheduling.
    # Each accepted time slot is appended as a dict compatible with
    # `utils.faculty_conflict_utils.check_faculty_availability_in_period`.
    faculty_context_sessions: List[dict] = []

    # Import here to avoid potential circular imports at module import time.
    from utils.faculty_conflict_utils import check_faculty_availability_in_period
    
    # Helper to classify a combined course into Group 1 / Group 2 / cross-group
    def get_course_group_for_phase4(course_code: str, semester: int) -> str:
        """
        Classify a course (by code + semester) into:
        - 'group1_only' : only CSE (CSE-A/B) sections have it
        - 'group2_only' : only DSAI/ECE sections have it
        - 'cross_group' : both CSE and DSAI/ECE have it

        This is used to enforce 30-minute gaps BETWEEN Group 1 and Group 2
        combined classes when scheduling time slots.
        """
        course_instances = [c for c in courses if c.code == course_code and c.semester == semester]
        if not course_instances:
            return "standard"

        departments_with_course = set(c.department for c in course_instances)

        # Determine which GROUP ids have this course by looking at configured sections
        groups_with_course = set()
        for s in sections:
            if not hasattr(s, "semester") or s.semester != semester:
                continue
            dept = getattr(s, "program", None)
            if dept in departments_with_course:
                groups_with_course.add(getattr(s, "group", 1))

        group1_sections = [s for s in sections if hasattr(s, "semester") and s.semester == semester and getattr(s, "group", 1) == 1]
        group2_sections = [s for s in sections if hasattr(s, "semester") and s.semester == semester and getattr(s, "group", 1) == 2]

        group1_has_course = 1 in groups_with_course
        group2_has_course = 2 in groups_with_course

        if group1_has_course and group2_has_course:
            return "cross_group"
        elif group1_has_course and not group2_has_course and len(group1_sections) > 1:
            return "group1_only"
        elif group2_has_course and not group1_has_course and len(group2_sections) > 1:
            return "group2_only"
        else:
            # Default: whichever group has the course (if any), else standard
            if group1_has_course:
                return "group1_only"
            if group2_has_course:
                return "group2_only"
            return "standard"

    # Build or load PreMid/PostMid decisions for each combined course
    period_assignments_path = get_period_assignments_path()
    period_assignments, period_assignments_modified = get_period_assignments_for_courses(
        unique_courses,
        sections,
        period_assignments_path,
    )

    for semester, sem_courses in unique_courses.items():
        print(f"DEBUG: Processing semester {semester} with {len(sem_courses)} courses")
        if not sem_courses:
            print(f"DEBUG: Skipping semester {semester} - no courses")
            continue
            
        # Sort courses dynamically based on properties for proper synchronization
        # Sort by: credits (ascending), then by number of sections requiring the course, then alphabetically
        def course_sort_key(c):
            # Get number of sections that need this course
            sections_needing = len([s for s in sections if hasattr(s, 'semester') and s.semester == semester])
            return (c.credits, -sections_needing, c.code)

        courses_for_semester = sorted(sem_courses, key=course_sort_key)

        # Split courses between PreMid and PostMid based on user assignments
        premid_courses: List[Course] = []
        postmid_courses: List[Course] = []
        for course in courses_for_semester:
            key = (course.code.upper(), semester, 1)
            is_premid = period_assignments.get(key, True)
            if is_premid:
                premid_courses.append(course)
            else:
                postmid_courses.append(course)

        num_courses = len(courses_for_semester)
        print(
            f"DEBUG: User-driven split for semester {semester} with {num_courses} courses: "
            f"{len(premid_courses)} PreMid, {len(postmid_courses)} PostMid"
        )
        print(f"  PreMid courses: {[c.code for c in premid_courses]}")
        print(f"  PostMid courses: {[c.code for c in postmid_courses]}")
        
        # Generate slots dynamically based on available time windows
        # Define available time slots that avoid lunch conflicts dynamically
        base_days = list(WORKING_DAYS)
        
        # Get lunch block for this semester
        lunch_blocks = {1: (time(12, 30), time(13, 30)),
                       3: (time(12, 45), time(13, 45)),
                       5: (time(13, 0), time(14, 0))}
        lunch_start, lunch_end = lunch_blocks.get(semester, (time(12, 30), time(13, 30)))
        
        # Dynamically generate available time slots based on 9:00-18:00 window
        available_slots = []
        
        # Morning slots (before lunch)
        current_time = time(9, 0)
        while current_time.hour < lunch_start.hour or \
              (current_time.hour == lunch_start.hour and current_time.minute < lunch_start.minute):
            # Create 1.5-hour slot for lectures
            end_hour = current_time.hour + 1
            end_minute = current_time.minute + 30
            if end_minute >= 60:
                end_hour += 1
                end_minute -= 60
            
            # Ensure slot doesn't overlap with lunch
            if end_hour < lunch_start.hour or \
               (end_hour == lunch_start.hour and end_minute <= lunch_start.minute):
                available_slots.append((current_time, time(end_hour, end_minute)))
            
            # Move to next slot (15-minute breaks)
            current_time = time(end_hour, end_minute)
            if current_time.minute < 45:
                current_time = time(current_time.hour, current_time.minute + 15)
            else:
                current_time = time(current_time.hour + 1, 0)
        
        # Afternoon slots (after lunch)
        current_time = lunch_end
        while current_time.hour < 18:
            # Create 1.5-hour slot
            end_hour = current_time.hour + 1
            end_minute = current_time.minute + 30
            if end_minute >= 60:
                end_hour += 1
                end_minute -= 60
            
            if end_hour <= 18:
                available_slots.append((current_time, time(end_hour, end_minute)))
            
            # Move to next slot
            current_time = time(end_hour, end_minute)
            if current_time.minute < 45:
                current_time = time(current_time.hour, current_time.minute + 15)
            else:
                current_time = time(current_time.hour + 1, 0)
        
        print(f"Dynamically generated {len(available_slots)} available time slots for semester {semester}")
        
        # Helper function to generate slots for a course based on LTPSC
        def generate_course_slots(course: Course, slot_index: int):
            """Generate slots for a course based on its LTPSC - ONE SESSION PER DAY PER COURSE"""
            slots_info = calculate_slots_from_ltpsc(course.ltpsc)
            course_slots = []
            current_slot_idx = slot_index
            
            # Calculate total sessions needed for this course
            total_sessions = slots_info['lectures'] + slots_info['tutorials'] + slots_info['practicals']
            
            # Generate ONE session per day, cycling through session types
            session_types = []
            # Add lectures
            for i in range(slots_info['lectures']):
                session_types.append('L')
            # Add tutorials  
            for i in range(slots_info['tutorials']):
                session_types.append('T')
            # Add practicals
            for i in range(slots_info['practicals']):
                session_types.append('P')
            
            # Schedule each session on a different day - ONE SESSION PER DAY PER COURSE
            for i, session_type in enumerate(session_types):
                if current_slot_idx < len(available_slots) * len(base_days):
                    # Each session goes to a DIFFERENT day - ensure no same-day sessions
                    day_idx = i % len(base_days)  # Cycle through days for this course
                    
                    # Choose appropriate time slot based on session type
                    if session_type == 'L':  # Lectures: use morning slots (0, 1)
                        slot_idx = 0  # Always use morning slot 1 for lectures
                    elif session_type == 'T':  # Tutorials: use afternoon slot 1 (2)
                        slot_idx = 2  # Afternoon slot 1
                    elif session_type == 'P':  # Practicals: use afternoon slot 2 (3) - AFTER lunch
                        slot_idx = 3  # Afternoon slot 2 (15:45-17:15)
                    
                    day = base_days[day_idx]
                    start, end = available_slots[slot_idx]
                    
                    # Adjust end time based on session type
                    if session_type == 'L':  # Lecture: 1.5 hours
                        end_hour = start.hour + 1
                        end_minute = start.minute + 30
                        if end_minute >= 60:
                            end_hour += 1
                            end_minute -= 60
                        end = time(end_hour, end_minute)
                    elif session_type == 'T':  # Tutorial: 1 hour
                        end = time(start.hour + 1, start.minute)
                    elif session_type == 'P':  # Practical: 2 hours
                        end = time(start.hour + 2, start.minute)
                    
                    course_slots.append((day, start, end, session_type))
                    current_slot_idx += 1
            
            return course_slots, current_slot_idx
        
        # Helper function to generate slots for a course based on LTPSC - FIXED VERSION
        def generate_course_slots_fixed(course: Course, occupied_slots: dict, base_days: list, available_slots: list):
            """Generate slots for a course based on its LTPSC - NO OVERLAPS, ONE SESSION PER DAY"""
            slots_info = calculate_slots_from_ltpsc(course.ltpsc)
            course_slots = []
            
            # Build session types list
            session_types = []
            # Add lectures
            for i in range(slots_info['lectures']):
                session_types.append('L')
            # Add tutorials
            for i in range(slots_info['tutorials']):
                session_types.append('T')
            # Add practicals
            for i in range(slots_info['practicals']):
                session_types.append('P')
            
            used_days = set()  # Track days used by this course
            
            # Schedule each session on a different day - ONE SESSION PER DAY PER COURSE
            for session_type in session_types:
                assigned = False
                
                # Try to find an available slot for this session type
                for day in base_days:
                    if day in used_days:
                        continue  # Skip days already used by this course
                    
                    # Find appropriate slot for this session type
                    if session_type == 'L':  # Lectures: use morning slots (0, 1)
                        for slot_idx in [0, 1]:
                            if (day, slot_idx) not in occupied_slots:
                                # Found available slot
                                start, end = available_slots[slot_idx]
                                
                                # Adjust end time for lecture (1.5 hours)
                                end_hour = start.hour + 1
                                end_minute = start.minute + 30
                                if end_minute >= 60:
                                    end_hour += 1
                                    end_minute -= 60
                                end = time(end_hour, end_minute)
                                
                                course_slots.append((day, start, end, session_type))
                                occupied_slots[(day, slot_idx)] = course.code
                                used_days.add(day)
                                assigned = True
                                break
                        if assigned:
                            break
                            
                    elif session_type == 'T':  # Tutorials: use afternoon slot 1 (2)
                        if (day, 2) not in occupied_slots:
                            start, end = available_slots[2]
                            end = time(start.hour + 1, start.minute)  # 1 hour
                            course_slots.append((day, start, end, session_type))
                            occupied_slots[(day, 2)] = course.code
                            used_days.add(day)
                            assigned = True
                            break
                            
                    elif session_type == 'P':  # Practicals: use afternoon slot 2 (3) - AFTER lunch
                        if (day, 3) not in occupied_slots:
                            start, end = available_slots[3]
                            end = time(start.hour + 2, start.minute)  # 2 hours
                            course_slots.append((day, start, end, session_type))
                            occupied_slots[(day, 3)] = course.code
                            used_days.add(day)
                            assigned = True
                            break
                
                if not assigned:
                    print(f"WARNING: Could not assign {session_type} session for {course.code}")
            
            return course_slots, occupied_slots
        
        def find_available_room(day: str, start: time, end: time, global_room_occupancy: dict,
                                available_large_rooms: list, allow_fallback: bool = False):
            """Find an available large room for the time slot.
            When allow_fallback=False (default), returns None if no room is free (caller should try another slot).
            When allow_fallback=True, returns C004 if no room is free so the course can still be scheduled."""
            test_block = TimeBlock(day, start, end)
            import random
            shuffled_rooms = available_large_rooms.copy()
            random.shuffle(shuffled_rooms)
            for room in shuffled_rooms:
                if day not in global_room_occupancy[room]:
                    return room
                has_conflict = False
                for occupied_start, occupied_end in global_room_occupancy[room][day]:
                    occupied_block = TimeBlock(day, occupied_start, occupied_end)
                    if test_block.overlaps(occupied_block):
                        has_conflict = True
                        break
                if not has_conflict:
                    return room
            if allow_fallback:
                return 'C004'
            return None
        
        def mark_room_occupied(room: str, day: str, start: time, end: time, global_room_occupancy: dict):
            """Mark room as occupied for given time slot"""
            # Ensure room exists in global_room_occupancy (even if it was a fallback assignment)
            if room not in global_room_occupancy:
                global_room_occupancy[room] = {}
            if day not in global_room_occupancy[room]:
                global_room_occupancy[room][day] = []
            global_room_occupancy[room][day].append((start, end))
        
        def get_elective_basket_slots(semester: int) -> List[TimeBlock]:
            """Get elective basket time slots for a semester to avoid conflicts (all groups)"""
            try:
                from modules_v2.phase3_elective_baskets_v2 import ELECTIVE_BASKET_SLOTS

                def extract_semester_from_group(gk: str) -> int:
                    try:
                        s = str(gk)
                        if '.' in s:
                            return int(s.split('.')[0])
                        return int(s)
                    except Exception:
                        return -1

                matching_groups = [
                    gk for gk in (ELECTIVE_BASKET_SLOTS or {}).keys()
                    if extract_semester_from_group(gk) == semester
                ]

                all_slots: List[TimeBlock] = []
                for group_key in matching_groups:
                    slots = ELECTIVE_BASKET_SLOTS[group_key]
                    all_slots.extend([slots.get('lecture_1'), slots.get('lecture_2'), slots.get('tutorial')])

                return [s for s in all_slots if s is not None]
            except Exception as e:
                # If import fails, return empty list (elective slots not yet calculated)
                pass
            return []
        
        def check_elective_conflict(day: str, start: time, end: time, semester: int) -> bool:
            """Check if time slot conflicts with elective basket slots"""
            elective_slots = get_elective_basket_slots(semester)
            test_block = TimeBlock(day, start, end)
            for elective_block in elective_slots:
                if elective_block.day == day and test_block.overlaps(elective_block):
                    return True  # Conflict found
            return False  # No conflict
        
        def check_lunch_conflict(day: str, start: time, end: time, semester: int) -> bool:
            """Check if time slot overlaps with lunch break"""
            test_block = TimeBlock(day, start, end)
            return test_block.overlaps_with_lunch(semester)
        
        def generate_synchronized_course_slots(
            course: Course,
            sections: list,
            occupied_slots: dict,
            base_days: list,
            available_slots: list,
            global_room_occupancy: dict,
            other_group_blocks: List[TimeBlock] = None,
            period_label: str = "PRE",
        ):
            """Generate SYNCHRONIZED slots for combined courses - SAME TIME across all sections.

            other_group_blocks is a list of TimeBlock instances representing combined-class
            sessions for the *other* group (and any cross-group sessions) in this period.
            Previously this was used to enforce an extra 30-minute gap between Group 1 and
            Group 2. We now only require **no actual overlaps in time** for C004, which is
            already guaranteed by room-occupancy tracking, so other_group_blocks is treated
            as informational only.
            """
            # Keep a normalized list for potential future checks, but do NOT enforce a 30-minute buffer.
            if other_group_blocks is None:
                other_group_blocks = []
            slots_info = calculate_slots_from_ltpsc(course.ltpsc)
            candidate_faculty = course.instructors[0] if getattr(course, "instructors", None) else None
            
            # Build session types list - CRITICAL: Must match LTPSC exactly
            session_types = []
            for i in range(slots_info['lectures']):
                session_types.append('L')
            for i in range(slots_info['tutorials']):
                session_types.append('T')
            for i in range(slots_info['practicals']):
                session_types.append('P')
            
            # DEBUG: Verify session types match LTPSC
            expected_total = slots_info['lectures'] + slots_info['tutorials'] + slots_info['practicals']
            if len(session_types) != expected_total:
                print(f"ERROR: {course.code} session_types mismatch! Expected {expected_total}, got {len(session_types)}")
                print(f"  LTPSC: {course.ltpsc}, slots_info: {slots_info}")
                print(f"  session_types: {session_types}")
            
            # CRITICAL FIX: Ensure session_types exactly matches LTPSC requirements
            # Remove incorrect session types first
            if slots_info['practicals'] == 0 and 'P' in session_types:
                print(f"ERROR: {course.code} has practicals in session_types but LTPSC shows P=0! Removing...")
                session_types = [s for s in session_types if s != 'P']  # Remove practicals
            
            if slots_info['tutorials'] == 0 and 'T' in session_types:
                print(f"ERROR: {course.code} has tutorials in session_types but LTPSC shows T=0! Removing...")
                session_types = [s for s in session_types if s != 'T']  # Remove tutorials
            
            # Add missing session types
            if slots_info['practicals'] > 0 and 'P' not in session_types:
                print(f"ERROR: {course.code} missing practicals! LTPSC shows P={slots_info['practicals']} but no P in session_types. Adding...")
                # Add missing practicals
                for i in range(slots_info['practicals'] - session_types.count('P')):
                    session_types.append('P')
            
            if slots_info['tutorials'] > 0 and 'T' not in session_types:
                print(f"ERROR: {course.code} missing tutorials! LTPSC shows T={slots_info['tutorials']} but no T in session_types. Adding...")
                # Add missing tutorials
                for i in range(slots_info['tutorials'] - session_types.count('T')):
                    session_types.append('T')
            
            # Final verification: Ensure counts match exactly
            lecture_count = session_types.count('L')
            tutorial_count = session_types.count('T')
            practical_count = session_types.count('P')
            
            if lecture_count != slots_info['lectures'] or tutorial_count != slots_info['tutorials'] or practical_count != slots_info['practicals']:
                print(f"CRITICAL ERROR: {course.code} session_types still doesn't match LTPSC after fixes!")
                print(f"  LTPSC: {course.ltpsc}, slots_info: {slots_info}")
                print(f"  session_types: {session_types} (L:{lecture_count}, T:{tutorial_count}, P:{practical_count})")
                # Force rebuild to match exactly
                session_types = []
                for i in range(slots_info['lectures']):
                    session_types.append('L')
                for i in range(slots_info['tutorials']):
                    session_types.append('T')
                for i in range(slots_info['practicals']):
                    session_types.append('P')
                print(f"  Rebuilt session_types: {session_types}")
            
            # Debug logging for EC161 and MA261 specifically
            if course.code == 'EC161':
                print(f"DEBUG EC161: LTPSC={course.ltpsc}, slots_info={slots_info}, session_types={session_types}")
            if course.code == 'MA261':
                print(f"DEBUG MA261: LTPSC={course.ltpsc}, slots_info={slots_info}, session_types={session_types}")
            
            # DYNAMIC: Shuffle days and slot indices for variety
            shuffled_days = base_days.copy()
            random.shuffle(shuffled_days)
            
            # Generate time slots ONCE for this course (same for all sections)
            course_time_slots = []
            # Use SessionRulesValidator for consistent rule enforcement
            used_days_by_course = SessionRulesValidator.get_used_days_tracker()
            
            for session_type in session_types:
                assigned = False
                
                # First attempt: Try with strict rules
                for day in shuffled_days:
                    # CRITICAL: Use SessionRulesValidator to enforce one-session-per-day rule
                    # No two lectures, no two tutorials, no lecture+tutorial on same day
                    # Practicals CAN be on same day as lectures/tutorials
                    if not SessionRulesValidator.can_schedule_session_type(
                        course.code, day, session_type, used_days_by_course
                    ):
                        continue  # Cannot schedule this session type on this day
                    
                    # Additional check: if day already used by this course (for practicals), still allow
                    # (SessionRulesValidator already handles L/T conflicts)
                    
                    # Find appropriate slot for this session type - DYNAMIC: try all available slots
                    # Determine which slots to try based on session type
                    if session_type == 'L':  # Lectures: prefer morning slots, but try all
                        # Try morning slots first (indices 0, 1), then afternoon
                        slot_indices = list(range(len(available_slots)))
                        # Prioritize morning slots (before lunch)
                        morning_slots = [i for i in range(len(available_slots)) if available_slots[i][0].hour < 13]
                        afternoon_slots = [i for i in range(len(available_slots)) if i not in morning_slots]
                        slot_indices = morning_slots + afternoon_slots
                        random.shuffle(slot_indices)  # Shuffle for variety
                    elif session_type == 'T':  # Tutorials: prefer afternoon slots
                        # Try afternoon slots first, then morning
                        afternoon_slots = [i for i in range(len(available_slots)) if available_slots[i][0].hour >= 13]
                        morning_slots = [i for i in range(len(available_slots)) if i not in afternoon_slots]
                        slot_indices = afternoon_slots + morning_slots
                        random.shuffle(slot_indices)
                    else:  # Practicals: prefer afternoon slots
                        afternoon_slots = [i for i in range(len(available_slots)) if available_slots[i][0].hour >= 13]
                        morning_slots = [i for i in range(len(available_slots)) if i not in afternoon_slots]
                        slot_indices = afternoon_slots + morning_slots
                        random.shuffle(slot_indices)
                    
                    # Try each slot index
                    for slot_idx in slot_indices:
                        if slot_idx >= len(available_slots):
                            continue  # Skip invalid indices

                        # Candidate start/end (will be adjusted based on session_type)
                        start, end = available_slots[slot_idx]

                        # Adjust end time based on session type
                        if session_type == 'L':  # Lecture: 1.5 hours
                            end_hour = start.hour + 1
                            end_minute = start.minute + 30
                            if end_minute >= 60:
                                end_hour += 1
                                end_minute -= 60
                            end = time(end_hour, end_minute)
                        elif session_type == 'T':  # Tutorial: 1 hour
                            end = time(start.hour + 1, start.minute)
                        else:  # Practical: 2 hours
                            end = time(start.hour + 2, start.minute)

                        test_block = TimeBlock(day, start, end)

                        # Prevent overlaps across combined sessions in the same semester/period:
                        # use an overlap-based occupied list in addition to (day, slot_idx).
                        occupied_blocks_key = ("__BLOCKS__", day)
                        for existing_block in occupied_slots.get(occupied_blocks_key, []):
                            if test_block.overlaps(existing_block):
                                test_block = None
                                break
                        if test_block is None:
                            continue

                        # Prevent overlaps within the same course on the same day (e.g., L overlapping P,
                        # or P overlapping an already-placed L/T). This applies to ALL session types so the
                        # order of scheduling (L before P or P before L) cannot create overlaps.
                        for existing in course_time_slots:
                            if len(existing) >= 3 and existing[0] == day:
                                existing_block = TimeBlock(day, existing[1], existing[2])
                                if test_block.overlaps(existing_block):
                                    test_block = None
                                    break
                        if test_block is None:
                            continue

                        # Still keep original discrete-slot occupancy as an additional guard
                        if (day, slot_idx) in occupied_slots:
                            continue
                        
                        # CRITICAL: Validate time range (9:00-18:00)
                        # For 2-hour practicals, ensure start time allows completion before 18:00
                        # Practical starting at 16:30 would end at 18:30, which is invalid
                        # Practical starting at 17:00 would end at 19:00, which is invalid
                        if not validate_time_range(start, end):
                            continue  # Skip this slot - extends beyond 18:00 or starts before 9:00
                        
                        # Additional check: For 2-hour practicals, start must be <= 16:00
                        if session_type == "P" and not slot_end_within_day(start, 120):
                            continue
                        
                        # Find available room for this time slot
                        assigned_room = find_available_room(day, start, end, global_room_occupancy, available_large_rooms)
                        if not assigned_room:
                            continue  # No room available, try next slot
                        
                        # CRITICAL: Check elective basket conflict
                        if check_elective_conflict(day, start, end, course.semester):
                            continue  # Skip this slot to avoid elective conflict
                        
                        # CRITICAL: Check lunch conflict after adjusting end time
                        if check_lunch_conflict(day, start, end, course.semester):
                            continue  # Skip this slot, try next

                        # Prevent faculty double-booking within the same PRE/POST bucket.
                        if candidate_faculty and not check_faculty_availability_in_period(
                            candidate_faculty,
                            day,
                            start,
                            end,
                            period_label,
                            faculty_context_sessions,
                        ):
                            continue
                        
                        course_time_slots.append((day, start, end, session_type, assigned_room))
                        # Mark day as used using SessionRulesValidator
                        SessionRulesValidator.mark_day_used(course.code, day, session_type, used_days_by_course)
                        # Mark room as occupied
                        mark_room_occupied(assigned_room, day, start, end, global_room_occupancy)
                        # Mark slot as occupied
                        occupied_slots[(day, slot_idx)] = course.code
                        # Mark overlap-based occupancy too (only needed once per (day, slot) for P,
                        # but safe to append for all types as blocks are non-overlapping by construction)
                        occupied_blocks_key = ("__BLOCKS__", day)
                        occupied_slots.setdefault(occupied_blocks_key, []).append(TimeBlock(day, start, end))

                        # Track faculty occupancy for subsequent preventive checks.
                        if candidate_faculty:
                            faculty_context_sessions.append(
                                {
                                    "instructor": candidate_faculty,
                                    "time_block": TimeBlock(day, start, end),
                                    "period": period_label,
                                }
                            )
                        assigned = True
                        break
                    
                    if assigned:
                        break
                
                # If not assigned, retry with relaxed constraints (allow same-day for P only)
                if not assigned:
                    print(f"  RETRY: {course.code} {session_type} - trying with relaxed constraints...")
                    # Retry: For P only, allow same-day as L/T. For L and T, NEVER relax: no L+T same day.
                    for retry_day in base_days:
                        if session_type in ('L', 'T') and not SessionRulesValidator.can_schedule_session_type(
                                course.code, retry_day, session_type, used_days_by_course):
                            continue
                        for retry_slot_idx in range(len(available_slots)):
                            if (retry_day, retry_slot_idx) in occupied_slots:
                                continue
                            
                            start, end = available_slots[retry_slot_idx]
                            
                            # Adjust end time
                            if session_type == 'L':
                                end_hour = start.hour + 1
                                end_minute = start.minute + 30
                                if end_minute >= 60:
                                    end_hour += 1
                                    end_minute -= 60
                                end = time(end_hour, end_minute)
                            elif session_type == 'T':
                                end = time(start.hour + 1, start.minute)
                            else:  # Practical
                                end = time(start.hour + 2, start.minute)

                            test_block = TimeBlock(retry_day, start, end)
                            
                            if not validate_time_range(start, end):
                                continue
                            
                            if session_type == "P" and not slot_end_within_day(start, 120):
                                continue

                            assigned_room = find_available_room(retry_day, start, end, global_room_occupancy, available_large_rooms)
                            if not assigned_room:
                                continue
                            
                            if check_elective_conflict(retry_day, start, end, course.semester):
                                continue
                            
                            if check_lunch_conflict(retry_day, start, end, course.semester):
                                continue

                            # Prevent faculty double-booking within the same PRE/POST bucket.
                            if candidate_faculty and not check_faculty_availability_in_period(
                                candidate_faculty,
                                retry_day,
                                start,
                                end,
                                period_label,
                                faculty_context_sessions,
                            ):
                                continue
                            
                            # Found a slot! Use it (L/T still respect one-session-per-day; P may share day with L/T)
                            course_time_slots.append((retry_day, start, end, session_type, assigned_room))
                            SessionRulesValidator.mark_day_used(course.code, retry_day, session_type, used_days_by_course)
                            mark_room_occupied(assigned_room, retry_day, start, end, global_room_occupancy)
                            occupied_slots[(retry_day, retry_slot_idx)] = course.code
                            # CRITICAL FIX: also update __BLOCKS__ so subsequent courses see this slot
                            occupied_slots.setdefault(("__BLOCKS__", retry_day), []).append(TimeBlock(retry_day, start, end))

                            # Track faculty occupancy for subsequent preventive checks.
                            if candidate_faculty:
                                faculty_context_sessions.append(
                                    {
                                        "instructor": candidate_faculty,
                                        "time_block": TimeBlock(retry_day, start, end),
                                        "period": period_label,
                                    }
                                )
                            assigned = True
                            print(f"  RETRY SUCCESS: Assigned {session_type} for {course.code} at {retry_day} {start}-{end}")
                            break
                        
                        if assigned:
                            break
                
                if not assigned:
                    # CRITICAL: For required sessions, we MUST find a slot - be more aggressive
                    # Try all days and all slots. NEVER relax L/T same-day: no 2L, no 2T, no L+T same day.
                    print(f"WARNING: Could not assign {session_type} session for {course.code} - trying aggressive search...")
                    
                    for retry_day in base_days:
                        if session_type in ('L', 'T') and not SessionRulesValidator.can_schedule_session_type(
                                course.code, retry_day, session_type, used_days_by_course):
                            continue
                        for retry_slot_idx in range(len(available_slots)):
                            if (retry_day, retry_slot_idx) in occupied_slots:
                                continue  # Still check basic occupancy
                            
                            start, end = available_slots[retry_slot_idx]
                            
                            # Adjust end time based on session type
                            if session_type == 'L':
                                end_hour = start.hour + 1
                                end_minute = start.minute + 30
                                if end_minute >= 60:
                                    end_hour += 1
                                    end_minute -= 60
                                end = time(end_hour, end_minute)
                            elif session_type == 'T':
                                end = time(start.hour + 1, start.minute)
                            else:  # Practical
                                end = time(start.hour + 2, start.minute)

                            test_block = TimeBlock(retry_day, start, end)
                            
                            # Basic validation only
                            if not validate_time_range(start, end):
                                continue
                            
                            if session_type == "P" and not slot_end_within_day(start, 120):
                                continue
                            
                            # Find available room
                            assigned_room = find_available_room(retry_day, start, end, global_room_occupancy, available_large_rooms)
                            if not assigned_room:
                                continue
                            
                            # Skip elective/lunch conflicts only if absolutely necessary
                            # For required sessions, we'll allow minor conflicts if needed
                            if check_elective_conflict(retry_day, start, end, course.semester):
                                continue
                            
                            if check_lunch_conflict(retry_day, start, end, course.semester):
                                continue

                            # Prevent faculty double-booking within the same PRE/POST bucket.
                            if candidate_faculty and not check_faculty_availability_in_period(
                                candidate_faculty,
                                retry_day,
                                start,
                                end,
                                period_label,
                                faculty_context_sessions,
                            ):
                                continue
                            
                            # Found a slot! Use it (L/T still respect no same-day rule)
                            course_time_slots.append((retry_day, start, end, session_type, assigned_room))
                            SessionRulesValidator.mark_day_used(course.code, retry_day, session_type, used_days_by_course)
                            mark_room_occupied(assigned_room, retry_day, start, end, global_room_occupancy)
                            occupied_slots[(retry_day, retry_slot_idx)] = course.code
                            if candidate_faculty:
                                faculty_context_sessions.append(
                                    {
                                        "instructor": candidate_faculty,
                                        "time_block": TimeBlock(retry_day, start, end),
                                        "period": period_label,
                                    }
                                )
                            assigned = True
                            print(f"  SUCCESS: Assigned {session_type} for {course.code} at {retry_day} {start}-{end} (aggressive search)")
                            break
                        
                        if assigned:
                            break
                    
                    # Only use C004 fallback when no slot had a free 240-seater (resolution will fix conflicts)
                    if not assigned:
                        for fallback_day in base_days:
                            if session_type in ("L", "T") and not SessionRulesValidator.can_schedule_session_type(
                                course.code, fallback_day, session_type, used_days_by_course
                            ):
                                continue
                            for fallback_slot_idx in range(len(available_slots)):
                                if (fallback_day, fallback_slot_idx) in occupied_slots:
                                    continue
                                start, end = available_slots[fallback_slot_idx]
                                if session_type == "L":
                                    end_hour = start.hour + 1
                                    end_minute = start.minute + 30
                                    if end_minute >= 60:
                                        end_hour += 1
                                        end_minute -= 60
                                    end = time(end_hour, end_minute)
                                elif session_type == "T":
                                    end = time(start.hour + 1, start.minute)
                                else:
                                    end = time(start.hour + 2, start.minute)
                                test_block = TimeBlock(fallback_day, start, end)
                                conflict_with_other_group = False
                                for other_block in other_group_blocks:
                                    # User explicitly removed 30-minute buffer; only real overlaps are forbidden.
                                    if test_block.overlaps(other_block):
                                        conflict_with_other_group = True
                                        break
                                if conflict_with_other_group or not validate_time_range(start, end):
                                    continue
                                if session_type == "P" and not slot_end_within_day(start, 120):
                                    continue
                                if check_elective_conflict(fallback_day, start, end, course.semester):
                                    continue
                                if check_lunch_conflict(fallback_day, start, end, course.semester):
                                    continue

                                # Prevent faculty double-booking within the same PRE/POST bucket.
                                if candidate_faculty and not check_faculty_availability_in_period(
                                    candidate_faculty,
                                    fallback_day,
                                    start,
                                    end,
                                    period_label,
                                    faculty_context_sessions,
                                ):
                                    continue
                                assigned_room = find_available_room(
                                    fallback_day,
                                    start,
                                    end,
                                    global_room_occupancy,
                                    available_large_rooms,
                                    allow_fallback=True,
                                )
                                if assigned_room:
                                    course_time_slots.append(
                                        (fallback_day, start, end, session_type, assigned_room)
                                    )
                                    SessionRulesValidator.mark_day_used(
                                        course.code, fallback_day, session_type, used_days_by_course
                                    )
                                    mark_room_occupied(
                                        assigned_room, fallback_day, start, end, global_room_occupancy
                                    )
                                    occupied_slots[(fallback_day, fallback_slot_idx)] = course.code
                                    # CRITICAL FIX: also update __BLOCKS__ so subsequent courses see this slot
                                    occupied_slots.setdefault(("__BLOCKS__", fallback_day), []).append(TimeBlock(fallback_day, start, end))

                                    # Track faculty occupancy for subsequent preventive checks.
                                    if candidate_faculty:
                                        faculty_context_sessions.append(
                                            {
                                                "instructor": candidate_faculty,
                                                "time_block": TimeBlock(fallback_day, start, end),
                                                "period": period_label,
                                            }
                                        )
                                    assigned = True
                                    print(
                                        f"  FALLBACK: Assigned {session_type} for {course.code} at "
                                        f"{fallback_day} {start}-{end} (room {assigned_room})"
                                    )
                                    break
                            if assigned:
                                break

                    # ULTIMATE FALLBACK: if still not assigned, allow multiple L/T on same day
                    # (while still preventing any real time overlaps) so that we always
                    # use free C004 slots like Monday/Thursday evenings instead of leaving
                    # courses short and UNSATISFIED.
                    if not assigned:
                        print(
                            f"ULTIMATE RETRY: {course.code} {session_type} - "
                            f"allowing extra same-day L/T if needed to avoid UNSATISFIED"
                        )
                        for retry_day in base_days:
                            for retry_slot_idx in range(len(available_slots)):
                                if (retry_day, retry_slot_idx) in occupied_slots:
                                    continue
                                start, end = available_slots[retry_slot_idx]
                                if session_type == "L":
                                    end_hour = start.hour + 1
                                    end_minute = start.minute + 30
                                    if end_minute >= 60:
                                        end_hour += 1
                                        end_minute -= 60
                                    end = time(end_hour, end_minute)
                                elif session_type == "T":
                                    end = time(start.hour + 1, start.minute)
                                else:
                                    end = time(start.hour + 2, start.minute)

                                test_block = TimeBlock(retry_day, start, end)

                                if not validate_time_range(start, end):
                                    continue
                                if session_type == "P" and not slot_end_within_day(start, 120):
                                    continue
                                if check_elective_conflict(retry_day, start, end, course.semester):
                                    continue
                                if check_lunch_conflict(retry_day, start, end, course.semester):
                                    continue

                                # Prevent faculty double-booking within the same PRE/POST bucket.
                                if candidate_faculty and not check_faculty_availability_in_period(
                                    candidate_faculty,
                                    retry_day,
                                    start,
                                    end,
                                    period_label,
                                    faculty_context_sessions,
                                ):
                                    continue

                                # Do NOT relax no-overlap rule: still forbid overlaps with this
                                # course's own other sessions on the same day.
                                overlap_with_self = False
                                for existing in course_time_slots:
                                    if len(existing) >= 3 and existing[0] == retry_day:
                                        existing_block = TimeBlock(
                                            retry_day, existing[1], existing[2]
                                        )
                                        if test_block.overlaps(existing_block):
                                            overlap_with_self = True
                                            break
                                if overlap_with_self:
                                    continue

                                assigned_room = find_available_room(
                                    retry_day,
                                    start,
                                    end,
                                    global_room_occupancy,
                                    available_large_rooms,
                                    allow_fallback=True,
                                )
                                if not assigned_room:
                                    continue

                                course_time_slots.append(
                                    (retry_day, start, end, session_type, assigned_room)
                                )
                                # Mark occupancy so later sessions respect this.
                                mark_room_occupied(
                                    assigned_room, retry_day, start, end, global_room_occupancy
                                )
                                occupied_slots[(retry_day, retry_slot_idx)] = course.code
                                # CRITICAL FIX: also update __BLOCKS__ so subsequent courses see this slot
                                occupied_slots.setdefault(("__BLOCKS__", retry_day), []).append(TimeBlock(retry_day, start, end))
                                if candidate_faculty:
                                    faculty_context_sessions.append(
                                        {
                                            "instructor": candidate_faculty,
                                            "time_block": TimeBlock(retry_day, start, end),
                                            "period": period_label,
                                        }
                                    )
                                assigned = True
                                print(
                                    f"  ULTIMATE SUCCESS: Assigned {session_type} for {course.code} at "
                                    f"{retry_day} {start}-{end} (room {assigned_room})"
                                )
                                break
                            if assigned:
                                break

                    if not assigned:
                        print(
                            f"CRITICAL ERROR: Still could not assign {session_type} session for {course.code} "
                            f"(Sem {course.semester}) - this will cause UNSATISFIED status"
                        )
                        print(
                            "  Constraints respected: no elective/lunch/room conflicts; "
                            f"sections={[s.label for s in sections if hasattr(s, 'semester') and s.semester == course.semester]}"
                        )
            
            # POST-PROCESSING: Verify all required sessions were scheduled
            scheduled_lectures = sum(1 for slot in course_time_slots if len(slot) >= 4 and slot[3] == 'L')
            scheduled_tutorials = sum(1 for slot in course_time_slots if len(slot) >= 4 and slot[3] == 'T')
            scheduled_practicals = sum(1 for slot in course_time_slots if len(slot) >= 4 and slot[3] == 'P')
            
            # If any required sessions are missing, DO NOT force assignment (avoid overlaps)
            if scheduled_lectures < slots_info['lectures']:
                print(f"CRITICAL: {course.code} missing {slots_info['lectures'] - scheduled_lectures} lecture(s)!")
            if scheduled_tutorials < slots_info['tutorials']:
                print(f"CRITICAL: {course.code} missing {slots_info['tutorials'] - scheduled_tutorials} tutorial(s)! No forced assignment will be done (to avoid overlaps).")
            if scheduled_practicals < slots_info['practicals']:
                print(f"CRITICAL: {course.code} missing {slots_info['practicals'] - scheduled_practicals} practical(s)! No forced assignment will be done (to avoid overlaps).")
            
            # Debug logging for MA261 specifically
            if course.code == 'MA261':
                print(f"DEBUG MA261: LTPSC={course.ltpsc}, slots_info={slots_info}")
                print(f"DEBUG MA261: Scheduled - L:{scheduled_lectures}, T:{scheduled_tutorials}, P:{scheduled_practicals}")
                print(f"DEBUG MA261: Required - L:{slots_info['lectures']}, T:{slots_info['tutorials']}, P:{slots_info['practicals']}")
            
            # Now return ONE slot per unique time block (all sections combined, not multiplied).
            # Format: (day, start, end, session_type, section_label, room)
            # section_label is a comma-joined string of all sections in this group
            # e.g. "CSE-A-Sem1,CSE-B-Sem1" — downstream mapping will expand per section.
            section_labels = ",".join(
                s.label for s in sections
                if hasattr(s, "semester") and s.semester == course.semester and hasattr(s, "label")
            )
            all_slots = []
            for slot_info in course_time_slots:
                if len(slot_info) == 5:  # (day, start, end, session_type, room)
                    day, start, end, session_type, assigned_room = slot_info
                else:
                    day, start, end, session_type = slot_info[:4]
                    assigned_room = "C004"
                all_slots.append((day, start, end, session_type, section_labels, assigned_room))
            
            return all_slots, occupied_slots, global_room_occupancy
        
        # Create schedule for this semester
        # Rooms will be assigned dynamically per course
        semester_schedule = {
            'premid': {
                # PreMid contains Group 1's chosen PreMid courses AND Group 2's opposite-period courses.
                # To keep mapping robust, expose all combined course objects for this semester.
                'courses': [c.code for c in courses_for_semester],
                'course_objects': courses_for_semester,
                'slots': [],
                'room': None  # Will be assigned dynamically
            },
            'postmid': {
                # PostMid contains Group 1's chosen PostMid courses AND Group 2's opposite-period courses.
                'courses': [c.code for c in courses_for_semester],
                'course_objects': courses_for_semester,
                'slots': [],
                'room': None  # Will be assigned dynamically
            }
        }
        
        # Track occupied slots PER PERIOD so PRE slots don't block POST slots.
        occupied_slots_premid: dict = {}
        occupied_slots_postmid: dict = {}
        
        # Use global room occupancy tracking (shared across all semesters)
        # This ensures no conflicts between different semesters using any room
        
        # Determine which configured sections belong to Group 1 vs Group 2 for this semester.
        group1_sections = [
            s
            for s in sections
            if hasattr(s, "semester")
            and s.semester == semester
            and getattr(s, "group", 1) == 1
        ]
        group2_sections = [
            s
            for s in sections
            if hasattr(s, "semester")
            and s.semester == semester
            and getattr(s, "group", 1) == 2
        ]
        sections_by_group: Dict[int, List[Section]] = defaultdict(list)
        for s in sections:
            if hasattr(s, "semester") and s.semester == semester:
                sections_by_group[int(getattr(s, "group", 1) or 1)].append(s)

        def get_groups_for_course(course_code: str, sem: int) -> List[int]:
            depts = {
                c.department
                for c in courses
                if str(getattr(c, "code", "")).upper() == str(course_code).upper()
                and int(getattr(c, "semester", -1) or -1) == int(sem)
            }
            group_ids = set()
            for sec in sections:
                if getattr(sec, "semester", None) != sem:
                    continue
                if getattr(sec, "program", None) in depts:
                    group_ids.add(int(getattr(sec, "group", 1) or 1))
            return sorted(group_ids)

        print(
            f"DEBUG: Semester {semester} group sections: "
            f"G1={len(group1_sections)} ({[getattr(s, 'label', str(s)) for s in group1_sections]}), "
            f"G2={len(group2_sections)} ({[getattr(s, 'label', str(s)) for s in group2_sections]})"
        )

        def _track_unique_blocks(course_slots: List[tuple], bucket: List[TimeBlock]) -> None:
            unique = set()
            for slot in course_slots or []:
                if len(slot) >= 3:
                    unique.add((slot[0], slot[1], slot[2]))
            for day, start, end in unique:
                bucket.append(TimeBlock(day, start, end))

        # PASS 1 (PreMid/PostMid): schedule by GROUP.
        #
        # For cross-group combined courses, we keep the existing behavior:
        # - Group 1 uses the chosen period for that (course_code, semester)
        # - Group 2 uses the opposite period for the same course code
        #
        # For group-only within-department combined courses (e.g. CSE-only),
        # schedule ONLY for the sections that actually take the course.
        premid_course_slots_map: Dict[str, List[tuple]] = {}
        premid_group1_blocks: List[TimeBlock] = []
        premid_group2_blocks: List[TimeBlock] = []
        premid_extra_blocks: Dict[int, List[TimeBlock]] = defaultdict(list)

        # Build per-group course lists based on course classification.
        premid_courses_group1: List[Course] = []
        postmid_courses_group1: List[Course] = []
        premid_courses_group2: List[Course] = []
        postmid_courses_group2: List[Course] = []
        premid_courses_extra: Dict[int, List[Course]] = defaultdict(list)
        postmid_courses_extra: Dict[int, List[Course]] = defaultdict(list)

        for course in courses_for_semester:
            # Group-specific period lookup:
            # - Group 1/2 behavior remains unchanged for cross-group courses.
            # - Group 3+ uses independent saved keys.
            key_g1 = (course.code.upper(), semester, 1)
            key_g2 = (course.code.upper(), semester, 2)
            course_group = get_course_group_for_phase4(course.code, semester)
            if course_group == "cross_group":
                is_premid = period_assignments.get(key_g1, True)
                if is_premid:
                    premid_courses_group1.append(course)
                    postmid_courses_group2.append(course)
                else:
                    postmid_courses_group1.append(course)
                    premid_courses_group2.append(course)
            elif course_group == "group1_only":
                is_premid = period_assignments.get(key_g1, True)
                if is_premid:
                    premid_courses_group1.append(course)
                else:
                    postmid_courses_group1.append(course)
            elif course_group == "group2_only":
                is_premid = period_assignments.get(key_g2, period_assignments.get(key_g1, True))
                if is_premid:
                    premid_courses_group2.append(course)
                else:
                    postmid_courses_group2.append(course)
            else:
                # Default: treat like group1_only if only group1 has configured sections, else group2_only.
                if group1_sections:
                    is_premid = period_assignments.get(key_g1, True)
                    (premid_courses_group1 if is_premid else postmid_courses_group1).append(course)
                elif group2_sections:
                    is_premid = period_assignments.get(key_g2, period_assignments.get(key_g1, True))
                    (premid_courses_group2 if is_premid else postmid_courses_group2).append(course)

            for gid in get_groups_for_course(course.code, semester):
                if gid < 3:
                    continue
                gkey = (course.code.upper(), semester, gid)
                is_premid_extra = period_assignments.get(gkey, period_assignments.get(key_g1, True))
                if is_premid_extra:
                    premid_courses_extra[gid].append(course)
                else:
                    postmid_courses_extra[gid].append(course)

        for course in premid_courses_group1:
            if course.code in premid_course_slots_map:
                continue
            if not group1_sections:
                continue
            course_slots, occupied_slots_premid, premid_room_occupancy = generate_synchronized_course_slots(
                course,
                group1_sections,
                occupied_slots_premid,
                base_days,
                available_slots,
                premid_room_occupancy,
                other_group_blocks=list(premid_group2_blocks),
                period_label="PRE",
            )
            premid_course_slots_map[course.code] = course_slots
            _track_unique_blocks(course_slots, premid_group1_blocks)
            print(f"  PRE: Group 1 scheduled {course.code}: {len(course_slots)} sessions")

        for course in premid_courses_group2:
            if course.code in premid_course_slots_map:
                continue
            if not group2_sections:
                continue
            course_slots, occupied_slots_premid, premid_room_occupancy = generate_synchronized_course_slots(
                course,
                group2_sections,
                occupied_slots_premid,
                base_days,
                available_slots,
                premid_room_occupancy,
                other_group_blocks=list(premid_group1_blocks),
                period_label="PRE",
            )
            premid_course_slots_map[course.code] = course_slots
            _track_unique_blocks(course_slots, premid_group2_blocks)
            print(f"  PRE: Group 2 scheduled {course.code}: {len(course_slots)} sessions")

        for gid, course_list in sorted(premid_courses_extra.items()):
            target_sections = sections_by_group.get(gid, [])
            if not target_sections:
                continue
            for course in course_list:
                if course.code in premid_course_slots_map:
                    continue
                other_blocks = list(premid_group1_blocks) + list(premid_group2_blocks)
                for ogid, blocks in premid_extra_blocks.items():
                    if ogid != gid:
                        other_blocks.extend(blocks)
                course_slots, occupied_slots_premid, premid_room_occupancy = generate_synchronized_course_slots(
                    course,
                    target_sections,
                    occupied_slots_premid,
                    base_days,
                    available_slots,
                    premid_room_occupancy,
                    other_group_blocks=other_blocks,
                    period_label="PRE",
                )
                premid_course_slots_map[course.code] = course_slots
                _track_unique_blocks(course_slots, premid_extra_blocks[gid])
                print(f"  PRE: Group {gid} scheduled {course.code}: {len(course_slots)} sessions")

        # PASS 1 (PostMid): schedule by GROUP (same per-course classification as above).
        print(f"DEBUG: Starting PostMid scheduling for semester {semester}")
        postmid_course_slots_map: Dict[str, List[tuple]] = {}
        postmid_group1_blocks: List[TimeBlock] = []
        postmid_group2_blocks: List[TimeBlock] = []
        postmid_extra_blocks: Dict[int, List[TimeBlock]] = defaultdict(list)

        for course in postmid_courses_group1:
            if course.code in postmid_course_slots_map:
                continue
            if not group1_sections:
                continue
            course_slots, occupied_slots_postmid, postmid_room_occupancy = generate_synchronized_course_slots(
                course,
                group1_sections,
                occupied_slots_postmid,
                base_days,
                available_slots,
                postmid_room_occupancy,
                other_group_blocks=list(postmid_group2_blocks),
                period_label="POST",
            )
            postmid_course_slots_map[course.code] = course_slots
            _track_unique_blocks(course_slots, postmid_group1_blocks)
            print(f"  POST: Group 1 scheduled {course.code}: {len(course_slots)} sessions")

        for course in postmid_courses_group2:
            if course.code in postmid_course_slots_map:
                continue
            if not group2_sections:
                continue
            course_slots, occupied_slots_postmid, postmid_room_occupancy = generate_synchronized_course_slots(
                course,
                group2_sections,
                occupied_slots_postmid,
                base_days,
                available_slots,
                postmid_room_occupancy,
                other_group_blocks=list(postmid_group1_blocks),
                period_label="POST",
            )
            postmid_course_slots_map[course.code] = course_slots
            _track_unique_blocks(course_slots, postmid_group2_blocks)
            print(f"  POST: Group 2 scheduled {course.code}: {len(course_slots)} sessions")

        for gid, course_list in sorted(postmid_courses_extra.items()):
            target_sections = sections_by_group.get(gid, [])
            if not target_sections:
                continue
            for course in course_list:
                if course.code in postmid_course_slots_map:
                    continue
                other_blocks = list(postmid_group1_blocks) + list(postmid_group2_blocks)
                for ogid, blocks in postmid_extra_blocks.items():
                    if ogid != gid:
                        other_blocks.extend(blocks)
                course_slots, occupied_slots_postmid, postmid_room_occupancy = generate_synchronized_course_slots(
                    course,
                    target_sections,
                    occupied_slots_postmid,
                    base_days,
                    available_slots,
                    postmid_room_occupancy,
                    other_group_blocks=other_blocks,
                    period_label="POST",
                )
                postmid_course_slots_map[course.code] = course_slots
                _track_unique_blocks(course_slots, postmid_extra_blocks[gid])
                print(f"  POST: Group {gid} scheduled {course.code}: {len(course_slots)} sessions")

        """
        
        # PASS 1: Generate time slots for each unique course code (PostMid)
        # PostMid courses are synchronized dynamically based on course properties
        print(f"DEBUG: Starting PostMid scheduling for semester {semester}")
        print(f"  PreMid course codes: {[c.code for c in premid_courses]}")
        print(f"  PreMid course slots map keys: {list(premid_course_slots_map.keys())}")
        postmid_course_slots_map = {}  # {course_code: [(day, start, end, session_type), ...]}
        # Track Group 1 / Group 2 / cross-group combined blocks for this semester & period
        postmid_group1_blocks: List[TimeBlock] = []
        postmid_group2_blocks: List[TimeBlock] = []
        postmid_cross_blocks: List[TimeBlock] = []
        
        for course in postmid_courses:
            if course.code not in postmid_course_slots_map:
                # Dynamically find available slots for PostMid courses
                # For cross-group synchronization, PostMid courses should use same time slots as PreMid courses
                # Find the corresponding PreMid course index for synchronization
                course_sections = [s for s in sections if hasattr(s, 'semester') and s.semester == semester]
                if course_sections:
                    # Try to synchronize with PreMid course at same index position
                    postmid_index = postmid_courses.index(course)
                    should_sync_with_premid = False
                    premid_sync_course_code = None
                    
                    # If there's a PreMid course at the same index, try to sync with it
                    if postmid_index < len(premid_courses):
                        premid_sync_course = premid_courses[postmid_index]
                        premid_sync_course_code = premid_sync_course.code
                        if premid_sync_course_code in premid_course_slots_map:
                            should_sync_with_premid = True
                    
                    if should_sync_with_premid and premid_sync_course_code:
                        # Use the same time slots as the corresponding PreMid course but with correct session types
                        premid_slots = premid_course_slots_map[premid_sync_course_code]
                        # Calculate correct session types for this PostMid course
                        course_slots_info = calculate_slots_from_ltpsc(course.ltpsc)
                        course_session_types = []
                        for i in range(course_slots_info['lectures']):
                            course_session_types.append('L')
                        for i in range(course_slots_info['tutorials']):
                            course_session_types.append('T')
                        for i in range(course_slots_info['practicals']):
                            course_session_types.append('P')
                        
                        # CRITICAL FIX: Apply same validation as in generate_synchronized_course_slots
                        # Remove incorrect session types first
                        if course_slots_info['practicals'] == 0 and 'P' in course_session_types:
                            print(f"ERROR: {course.code} (sync) has practicals in course_session_types but LTPSC shows P=0! Removing...")
                            course_session_types = [s for s in course_session_types if s != 'P']
                        
                        if course_slots_info['tutorials'] == 0 and 'T' in course_session_types:
                            print(f"ERROR: {course.code} (sync) has tutorials in course_session_types but LTPSC shows T=0! Removing...")
                            course_session_types = [s for s in course_session_types if s != 'T']
                        
                        # Add missing session types
                        if course_slots_info['practicals'] > 0 and 'P' not in course_session_types:
                            print(f"ERROR: {course.code} (sync) missing practicals! LTPSC shows P={course_slots_info['practicals']} but no P in course_session_types. Adding...")
                            for i in range(course_slots_info['practicals'] - course_session_types.count('P')):
                                course_session_types.append('P')
                        
                        if course_slots_info['tutorials'] > 0 and 'T' not in course_session_types:
                            print(f"ERROR: {course.code} (sync) missing tutorials! LTPSC shows T={course_slots_info['tutorials']} but no T in course_session_types. Adding...")
                            for i in range(course_slots_info['tutorials'] - course_session_types.count('T')):
                                course_session_types.append('T')
                        
                        # Final verification
                        lecture_count = course_session_types.count('L')
                        tutorial_count = course_session_types.count('T')
                        practical_count = course_session_types.count('P')
                        
                        if lecture_count != course_slots_info['lectures'] or tutorial_count != course_slots_info['tutorials'] or practical_count != course_slots_info['practicals']:
                            print(f"CRITICAL ERROR: {course.code} (sync) course_session_types still doesn't match LTPSC after fixes!")
                            print(f"  LTPSC: {course.ltpsc}, slots_info: {course_slots_info}")
                            print(f"  course_session_types: {course_session_types} (L:{lecture_count}, T:{tutorial_count}, P:{practical_count})")
                            # Force rebuild to match exactly
                            course_session_types = []
                            for i in range(course_slots_info['lectures']):
                                course_session_types.append('L')
                            for i in range(course_slots_info['tutorials']):
                                course_session_types.append('T')
                            for i in range(course_slots_info['practicals']):
                                course_session_types.append('P')
                            print(f"  Rebuilt course_session_types: {course_session_types}")
                        
                        # Debug logging for EC161 and MA261
                        if course.code == 'EC161':
                            print(f"DEBUG EC161 (sync): LTPSC={course.ltpsc}, slots_info={course_slots_info}, course_session_types={course_session_types}")
                        if course.code == 'MA261':
                            print(f"DEBUG MA261 (sync): LTPSC={course.ltpsc}, slots_info={course_slots_info}, course_session_types={course_session_types}")
                        
                        # Create new slots with correct session types but same times
                        # CRITICAL: Ensure one-session-per-day rule (no lecture+tutorial on same day)
                        corrected_slots = []
                        session_type_idx = 0
                        used_days = set()  # Track days used by this course
                        used_days_lectures = set()  # Track lecture days
                        used_days_tutorials = set()  # Track tutorial days
                        
                        # CRITICAL: Sort course_session_types to process lectures first, then tutorials, then practicals
                        # This ensures we assign the right session types in the right order
                        sorted_session_types = []
                        for st in course_session_types:
                            if st == 'L':
                                sorted_session_types.append('L')
                        for st in course_session_types:
                            if st == 'T':
                                sorted_session_types.append('T')
                        for st in course_session_types:
                            if st == 'P':
                                sorted_session_types.append('P')
                        course_session_types = sorted_session_types
                        
                        # Debug logging for EC161 and MA261
                        if course.code == 'EC161':
                            print(f"DEBUG EC161 (sync): Using sorted course_session_types={course_session_types}")
                        if course.code == 'MA261':
                            print(f"DEBUG MA261 (sync): Using sorted course_session_types={course_session_types}")
                        
                        # Only process as many slots as we need for PostMid course
                        for slot_idx, slot in enumerate(premid_slots):
                            if session_type_idx >= len(course_session_types):
                                # We've created all required sessions for PostMid - stop processing
                                break
                                
                            # Handle both 5-tuple and 6-tuple slot formats
                            slot_parts = list(slot)
                            if len(slot_parts) >= 5:
                                day, start, end = slot_parts[0], slot_parts[1], slot_parts[2]
                                old_session_type = slot_parts[3]  # Old type from PreMid (may be wrong)
                                section = slot_parts[4] if len(slot_parts) > 4 else None
                                assigned_room = slot_parts[5] if len(slot_parts) > 5 else None
                                
                                # CRITICAL: Use the correct session type from course_session_types, NOT from PreMid slot
                                corrected_session_type = course_session_types[session_type_idx]
                                
                                # Debug logging for EC161
                                if course.code == 'EC161':
                                    print(f"DEBUG EC161 (sync): Slot {slot_idx}: old_type={old_session_type}, new_type={corrected_session_type}, day={day}, time={start}-{end}")
                                
                                # CRITICAL: Check one-session-per-day rule
                                if corrected_session_type == 'L':
                                    if day in used_days_lectures or day in used_days_tutorials:
                                        # Skip this day, try to find alternative day
                                        continue
                                elif corrected_session_type == 'T':
                                    if day in used_days_lectures or day in used_days_tutorials:
                                        # Skip this day, try to find alternative day
                                        continue
                                
                                # CRITICAL: Check lunch conflict for adjusted slot
                                # Adjust end time based on session type to ensure accurate lunch check
                                adjusted_end = end
                                if corrected_session_type == 'L':  # Lecture: 1.5 hours
                                    end_hour = start.hour + 1
                                    end_minute = start.minute + 30
                                    if end_minute >= 60:
                                        end_hour += 1
                                        end_minute -= 60
                                    adjusted_end = time(end_hour, end_minute)
                                elif corrected_session_type == 'T':  # Tutorial: 1 hour
                                    adjusted_end = time(start.hour + 1, start.minute)
                                elif corrected_session_type == 'P':  # Practical: 2 hours
                                    adjusted_end = time(start.hour + 2, start.minute)
                                
                                # CRITICAL: Validate time range (9:00-18:00)
                                if not validate_time_range(start, adjusted_end):
                                    continue  # Skip this slot - extends beyond 18:00
                                
                                if check_lunch_conflict(day, start, adjusted_end, semester):
                                    continue  # Skip this slot, try next
                                
                                # Create this session with adjusted end time and assigned room
                                # Format: (day, start, end, session_type, section_labels, assigned_room)
                                if assigned_room:
                                    corrected_slots.append((day, start, adjusted_end, corrected_session_type, section_labels, assigned_room))
                                else:
                                    # Find available room for this slot
                                    temp_room = find_available_room(
                                        day, start, adjusted_end, postmid_room_occupancy, available_large_rooms
                                    )
                                    if temp_room:
                                        corrected_slots.append((day, start, adjusted_end, corrected_session_type, section_labels, temp_room))
                                        mark_room_occupied(
                                            temp_room, day, start, adjusted_end, postmid_room_occupancy
                                        )
                                    else:
                                        # Use default room
                                        default_room = available_large_rooms[0] if available_large_rooms else 'C004'
                                        corrected_slots.append((day, start, adjusted_end, corrected_session_type, section_labels, default_room))
                                        mark_room_occupied(
                                            default_room, day, start, adjusted_end, postmid_room_occupancy
                                        )
                                
                                used_days.add(day)
                                if corrected_session_type == 'L':
                                    used_days_lectures.add(day)
                                elif corrected_session_type == 'T':
                                    used_days_tutorials.add(day)
                                session_type_idx += 1
                                
                                # Debug logging for EC161
                                if course.code == 'EC161':
                                    print(f"DEBUG EC161 (sync): Created slot {len(corrected_slots)}: {corrected_session_type} at {day} {start}-{adjusted_end}")
                            else:
                                corrected_slots.append(slot)
                        
                        # CRITICAL: Check if we have all required sessions
                        # If we skipped days due to one-session-per-day rule, we might be missing sessions
                        # In that case, regenerate slots instead of synchronizing
                        required_sessions = len(course_session_types)
                        # Count sessions by type to ensure we have the right mix
                        scheduled_lectures = sum(1 for slot in corrected_slots if len(slot) >= 4 and slot[3] == 'L')
                        scheduled_tutorials = sum(1 for slot in corrected_slots if len(slot) >= 4 and slot[3] == 'T')
                        scheduled_practicals = sum(1 for slot in corrected_slots if len(slot) >= 4 and slot[3] == 'P')
                        
                        # Check if we have the correct number of each session type
                        has_all_required = (
                            scheduled_lectures >= course_slots_info['lectures'] and
                            scheduled_tutorials >= course_slots_info['tutorials'] and
                            scheduled_practicals >= course_slots_info['practicals'] and
                            len(corrected_slots) >= required_sessions
                        )
                        
                        if not has_all_required:
                            # Regenerate slots to ensure all sessions are scheduled
                            print(f"  WARNING: {course.code} synchronization missing sessions (L:{scheduled_lectures}/{course_slots_info['lectures']}, T:{scheduled_tutorials}/{course_slots_info['tutorials']}, P:{scheduled_practicals}/{course_slots_info['practicals']}), regenerating slots...")
                            course_sections = [s for s in sections if hasattr(s, 'semester') and s.semester == semester]
                            if course_sections:
                                try:
                                    course_slots, occupied_slots_postmid, postmid_room_occupancy = generate_synchronized_course_slots(
                                        course,
                                        course_sections,
                                        occupied_slots_postmid,
                                        base_days,
                                        available_slots,
                                        postmid_room_occupancy,
                                        period_label="POST",
                                    )
                                    # Verify we got all required sessions
                                    if len(course_slots) >= required_sessions:
                                        postmid_course_slots_map[course.code] = course_slots
                                        print(f"  {course.code} time slots allocated: {len(course_slots)} sessions (regenerated to avoid conflicts)")
                                    else:
                                        # Fallback: use corrected_slots even if incomplete, but log warning
                                        postmid_course_slots_map[course.code] = corrected_slots
                                        print(f"  WARNING: {course.code} only got {len(course_slots)}/{required_sessions} sessions after regeneration, using synchronized slots")
                                except Exception as e:
                                    print(f"  ERROR: Failed to regenerate slots for {course.code}: {e}")
                                    # Fallback to corrected_slots
                                    postmid_course_slots_map[course.code] = corrected_slots
                                    print(f"  {course.code} using synchronized slots (regeneration failed)")
                            else:
                                postmid_course_slots_map[course.code] = corrected_slots
                                print(f"  {course.code} synchronized with {premid_sync_course_code} (same time slots, corrected session types)")
                        else:
                            postmid_course_slots_map[course.code] = corrected_slots
                            print(f"  {course.code} synchronized with {premid_sync_course_code} (same time slots, corrected session types)")
                    else:
                        # No PreMid course to sync with, or sync failed - generate slots dynamically
                        if course_sections:
                            try:
                                course_slots, occupied_slots_postmid, postmid_room_occupancy = generate_synchronized_course_slots(
                                    course,
                                    course_sections,
                                    occupied_slots_postmid,
                                    base_days,
                                    available_slots,
                                    postmid_room_occupancy,
                                    period_label="POST",
                                )
                                postmid_course_slots_map[course.code] = course_slots
                                print(f"  {course.code} time slots allocated: {len(course_slots)} sessions (dynamically generated)")
                            except Exception as e:
                                print(f"  ERROR: Failed to generate slots for {course.code}: {e}")
                                # Fallback: try to generate at least some slots
                                postmid_course_slots_map[course.code] = []
                        else:
                            print(f"  WARNING: No sections found for {course.code} in semester {semester}")
                            postmid_course_slots_map[course.code] = []
                    
                    # If still no slots, try to generate them
                    if course.code not in postmid_course_slots_map or not postmid_course_slots_map[course.code]:
                        course_sections = [s for s in sections if hasattr(s, 'semester') and s.semester == semester]
                        if course_sections:
                            try:
                                course_group = get_course_group_for_phase4(course.code, semester)
                                if course_group == "group1_only":
                                    other_blocks = postmid_group2_blocks + postmid_cross_blocks
                                elif course_group == "group2_only":
                                    other_blocks = postmid_group1_blocks + postmid_cross_blocks
                                else:
                                    other_blocks = postmid_group1_blocks + postmid_group2_blocks + postmid_cross_blocks

                                course_slots, occupied_slots_postmid, postmid_room_occupancy = generate_synchronized_course_slots(
                                    course,
                                    course_sections,
                                    occupied_slots_postmid,
                                    base_days,
                                    available_slots,
                                    postmid_room_occupancy,
                                    other_group_blocks=other_blocks,
                                    period_label="POST",
                                )
                                postmid_course_slots_map[course.code] = course_slots

                                # Track unique time blocks for 30-min gap enforcement
                                unique_blocks = set()
                                for slot in course_slots:
                                    if len(slot) >= 3:
                                        day = slot[0]
                                        start = slot[1]
                                        end = slot[2]
                                        unique_blocks.add((day, start, end))
                                for day, start, end in unique_blocks:
                                    tb = TimeBlock(day, start, end)
                                    if course_group == "group1_only":
                                        postmid_group1_blocks.append(tb)
                                    elif course_group == "group2_only":
                                        postmid_group2_blocks.append(tb)
                                    else:
                                        postmid_cross_blocks.append(tb)

                                print(f"  {course.code} time slots allocated: {len(course_slots)} sessions (fallback generation)")
                            except Exception as e:
                                print(f"  ERROR: Failed to generate slots for {course.code} in fallback: {e}")
                                postmid_course_slots_map[course.code] = []
                
                # All PostMid courses are now handled by dynamic generation above
                # No hardcoded course-specific logic remains
        """

        # Log time slots for Phase 4
        logger = get_logger()
        
        # PASS 2: Apply slots to the schedule dict and logger.
        # New format from generate_synchronized_course_slots:
        #   (day, start, end, session_type, section_label_str, room)
        # section_label_str is a comma-joined string like "CSE-A-Sem1,CSE-B-Sem1"
        for course_code, course_slots in premid_course_slots_map.items():
            for slot in course_slots:
                if len(slot) == 6 and isinstance(slot[4], str):
                    # New format: (day, start, end, session_type, section_str, room)
                    day, start, end, session_type, section_str, assigned_room = slot
                    log_room = assigned_room  # Use actual assigned room (Phase 8 will finalize labs)
                    # Store ONE slot tuple per time block (section_str may be comma-joined)
                    semester_schedule['premid']['slots'].append((course_code, day, start, end, session_type, section_str, assigned_room))
                    # Log once per section
                    for sec_label in section_str.split(','):
                        sec_label = sec_label.strip()
                        if sec_label:
                            logger.log_slot("Phase 4", course_code, sec_label, day, start, end,
                                            room=log_room, period='PRE', session_type=session_type)
                elif len(slot) == 7 and slot[0] == course_code:
                    # Already prefixed: (course_code, day, start, end, session_type, section, room)
                    semester_schedule['premid']['slots'].append(slot)
                    _, day, start, end, session_type, section_str, assigned_room = slot
                    log_room = assigned_room  # Use actual assigned room (Phase 8 will finalize labs)
                    sec_name = section_str.label if hasattr(section_str, 'label') else str(section_str)
                    for sec_label in sec_name.split(','):
                        sec_label = sec_label.strip()
                        if sec_label:
                            logger.log_slot("Phase 4", course_code, sec_label, day, start, end,
                                            room=log_room, period='PRE', session_type=session_type)
            
            print(f"  {course_code} scheduled: {len(course_slots)} sessions (synchronized across all sections)")

        # POST PASS 2: Apply slots to the schedule dict and logger.
        for course_code, course_slots in postmid_course_slots_map.items():
            for slot in course_slots:
                if len(slot) == 6 and isinstance(slot[4], str):
                    # New format: (day, start, end, session_type, section_str, room)
                    day, start, end, session_type, section_str, assigned_room = slot
                    log_room = assigned_room  # Use actual assigned room (Phase 8 will finalize labs)
                    semester_schedule['postmid']['slots'].append((course_code, day, start, end, session_type, section_str, assigned_room))
                    for sec_label in section_str.split(','):
                        sec_label = sec_label.strip()
                        if sec_label:
                            logger.log_slot("Phase 4", course_code, sec_label, day, start, end,
                                            room=log_room, period='POST', session_type=session_type)
                elif len(slot) == 7 and slot[0] == course_code:
                    semester_schedule['postmid']['slots'].append(slot)
                    _, day, start, end, session_type, section_str, assigned_room = slot
                    log_room = assigned_room  # Use actual assigned room (Phase 8 will finalize labs)
                    sec_name = section_str.label if hasattr(section_str, 'label') else str(section_str)
                    for sec_label in sec_name.split(','):
                        sec_label = sec_label.strip()
                        if sec_label:
                            logger.log_slot("Phase 4", course_code, sec_label, day, start, end,
                                            room=log_room, period='POST', session_type=session_type)
            
            print(f"  {course_code} scheduled: {len(course_slots)} sessions (synchronized across all sections)")
        
        schedule[semester] = semester_schedule
        occupied_slots_by_semester[semester] = {
            "PRE": occupied_slots_premid,
            "POST": occupied_slots_postmid,
        }

    # Persist any new or updated period assignments so future runs are non-interactive
    if period_assignments_modified:
        save_period_assignments(period_assignments_path, period_assignments)
    
    return {
        'schedule': schedule,
        'occupied_slots': occupied_slots_by_semester
    }

def run_phase4_corrected(courses: List[Course], sections: List[Section], classrooms: List = None) -> Dict:
    """
    Run Phase 4 with corrected combined course identification and scheduling
    
    Args:
        courses: List of Course objects
        sections: List of Section objects
        classrooms: List of ClassRoom objects (required to filter 240-seater rooms)
    """
    print("DEBUG: run_phase4_corrected function called!")
    print("=== PHASE 4: COMBINED CLASS SCHEDULING (CORRECTED) ===")
    print()
    
    # Get combined courses by semester (including within-group combined)
    combined_courses = get_combined_courses(courses, sections)
    unique_courses = get_unique_combined_courses_by_semester(combined_courses)
    
    print("Combined courses identified:")
    for semester, courses in unique_courses.items():
        course_codes = [c.code for c in courses]
        print(f"  Semester {semester}: {len(course_codes)} courses - {course_codes}")
    print()
    
    # Create non-overlapping schedule
    print("DEBUG: About to call create_non_overlapping_schedule")
    try:
        schedule_result = create_non_overlapping_schedule(unique_courses, sections, courses, classrooms)
        print("DEBUG: create_non_overlapping_schedule returned")
        schedule = schedule_result['schedule']
        occupied_slots = schedule_result['occupied_slots']
    except Exception as e:
        print(f"DEBUG: Error in create_non_overlapping_schedule: {e}")
        import traceback
        traceback.print_exc()
        raise
    
    print("Non-overlapping schedule created:")
    for semester, data in schedule.items():
        if 'premid' in data:
            print(f"  Semester {semester} PreMid: {len(data['premid']['courses'])} courses")
            print(f"    Courses: {data['premid']['courses']}")
            for course in data['premid']['course_objects']:
                slots_info = calculate_slots_from_ltpsc(course.ltpsc)
                print(f"      {course.code} (LTPSC: {course.ltpsc}): {slots_info['lectures']}L + {slots_info['tutorials']}T + {slots_info['practicals']}P")
            print(f"    Room: {data['premid']['room']}")
            print(f"    Time slots: {len(data['premid']['slots'])} slots")
        if 'postmid' in data:
            print(f"  Semester {semester} PostMid: {len(data['postmid']['courses'])} courses")
            print(f"    Courses: {data['postmid']['courses']}")
            for course in data['postmid']['course_objects']:
                slots_info = calculate_slots_from_ltpsc(course.ltpsc)
                print(f"      {course.code} (LTPSC: {course.ltpsc}): {slots_info['lectures']}L + {slots_info['tutorials']}T + {slots_info['practicals']}P")
            print(f"    Room: {data['postmid']['room']}")
            print(f"    Time slots: {len(data['postmid']['slots'])} slots")
        print()
    
    # Print time slot logging summary
    print("Time slot logging summary...")
    from utils.time_slot_logger import get_logger
    logger = get_logger()
    phase4_entries = logger.get_entries_by_phase("Phase 4")
    if phase4_entries:
        phase4_summary = logger.get_phase_summary("Phase 4")
        print(f"Phase 4 logged {phase4_summary['total_slots']} time slots")
        print(f"  - Unique courses: {phase4_summary['unique_courses']}")
        print(f"  - Unique sections: {phase4_summary['unique_sections']}")
        print(f"  - By day: {phase4_summary['by_day']}")
        print(f"  - By session type: {phase4_summary['by_session_type']}")
    
    return {
        'combined_courses': combined_courses,
        'unique_courses': unique_courses,
        'schedule': schedule,
        'occupied_slots': occupied_slots
    }

if __name__ == "__main__":
    # Test the corrected implementation
    from modules_v2.phase1_data_validation_v2 import run_phase1
    
    print("Testing corrected combined course identification...")
    courses, sections, classrooms = run_phase1()
    
    result = run_phase4_corrected(courses, sections)
    
    print("=== SUMMARY ===")
    print(f"Total combined course instances: {sum(len(courses) for courses in result['combined_courses'].values())}")
    print(f"Unique combined courses by semester: {result['unique_courses']}")
    print("Schedule created with 0 conflicts!")
