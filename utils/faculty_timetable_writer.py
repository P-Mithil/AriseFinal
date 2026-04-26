"""
Faculty-wise timetable writer and conflict reporter.

Creates a separate Excel workbook where each sheet contains the timetable
for a single faculty member (all days, both periods), and marks any
same-period double bookings using the Phase 6 `FacultyConflict` data.
"""

from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime, time as dt_time
from typing import Dict, Iterable, List, Tuple, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from utils.data_models import TimeBlock
from config.schedule_config import WORKING_DAYS

try:
    # Import only for type hints – avoid hard dependency at import time
    from modules_v2.phase6_faculty_conflicts import FacultyConflict
except Exception:  # pragma: no cover - defensive
    FacultyConflict = Any  # type: ignore


def _normalize_period(raw: str) -> str:
    """Normalize period strings to 'PRE' or 'POST'."""
    if not raw:
        return "PRE"
    val = str(raw).strip().upper()
    if val in ("PREMID", "PRE"):
        return "PRE"
    if val in ("POSTMID", "POST"):
        return "POST"
    return val


def _sanitize_sheet_name(faculty_name: str) -> str:
    """
    Create a safe Excel sheet name for a faculty.

    - Prefix with 'FAC_'
    - Replace spaces with underscores
    - Remove invalid characters: []:*?/\\
    - Truncate to 31 characters (Excel limit)
    """
    base = faculty_name.strip().replace(" ", "_")
    # Characters not allowed in Excel sheet names: : \ / ? * [ ]
    invalid_chars = "[]:*?/\\"
    for ch in invalid_chars:
        base = base.replace(ch, "")
    if not base:
        base = "Unknown"
    name = f"FAC_{base}"
    # Excel sheet name limit
    return name[:31]


def _build_faculty_conflict_index(
    conflicts: Iterable[FacultyConflict],
) -> Dict[Tuple[str, str, str, dt_time, dt_time], List[FacultyConflict]]:
    """
    Build an index for quick lookup of conflicts by
    (faculty_name, day, period, start_time, end_time).

    `FacultyConflict.time_slot` has the format:
        "{day} {start}-{end} ({period})"
    where start/end are usually HH:MM:SS.
    """
    index: Dict[Tuple[str, str, str, dt_time, dt_time], List[FacultyConflict]] = defaultdict(list)

    for conflict in conflicts or []:
        time_slot_str = getattr(conflict, "time_slot", "")
        faculty_name = getattr(conflict, "faculty_name", "")
        day = getattr(conflict, "day", "")

        if not time_slot_str or not faculty_name or not day:
            continue

        try:
            # Example: "Monday 09:00:00-10:30:00 (PRE)"
            day_part, rest = time_slot_str.split(" ", 1)
            # Trust the dedicated `day` field more than the string
            day_value = day or day_part

            times_part, period_part = rest.split("(", 1)
            times_part = times_part.strip()
            period_str = period_part.strip().rstrip(")")
            period = _normalize_period(period_str)

            start_str, end_str = [t.strip() for t in times_part.split("-", 1)]

            # Support both HH:MM:SS and HH:MM
            fmt = "%H:%M:%S" if len(start_str.split(":")) == 3 else "%H:%M"
            start_time = datetime.strptime(start_str, fmt).time()
            end_time = datetime.strptime(end_str, fmt).time()
        except Exception:
            # If parsing fails, skip this conflict (better to be silent than crash generation)
            continue

        key = (faculty_name, day_value, period, start_time, end_time)
        index[key].append(conflict)

    return index


def _collect_faculty_sessions(
    all_sessions: Iterable[Any],
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Normalize all sessions into a per-faculty list of dicts with:
    - day
    - period
    - start_time
    - end_time
    - course_code
    - section_str
    - room
    """
    faculty_sessions: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for session in all_sessions or []:
        # Extract core fields from both dict- and object-style sessions
        if isinstance(session, dict):
            faculty = session.get("instructor") or session.get("faculty")
            block: TimeBlock = session.get("time_block")
            period_raw = session.get("period", "PRE")
            course_code = session.get("course_code", "UNKNOWN")
            sections = session.get("sections") or []
            single_section = session.get("section")
            room = session.get("room")

            if sections and isinstance(sections, (list, tuple)):
                section_str = ",".join(str(s) for s in sections)
            elif sections:
                 section_str = str(sections)
            elif single_section:
                 section_str = str(single_section)
            else:
                 section_str = ""
        else:
            faculty = getattr(session, "faculty", None)
            block = getattr(session, "block", None)
            period_raw = getattr(session, "period", "PRE")
            course_code = getattr(session, "course_code", "UNKNOWN")
            section = getattr(session, "section", "")
            room = getattr(session, "room", None)
            section_str = str(section)

        if not faculty or str(faculty).lower() == "nan" or str(faculty) in ("TBD", "Various", "-"):
            continue
        if not block or not getattr(block, "start", None) or not getattr(block, "end", None):
            continue

        period = _normalize_period(period_raw)

        start_time = block.start
        end_time = block.end
        for t_var in [start_time, end_time]:
            if isinstance(t_var, str):
                try:
                    fmt = "%H:%M:%S" if len(t_var.split(":")) == 3 else "%H:%M"
                    if t_var == start_time: start_time = datetime.strptime(t_var, fmt).time()
                    else: end_time = datetime.strptime(t_var, fmt).time()
                except Exception:
                    pass

        faculty_sessions[faculty].append(
            {
                "day": block.day,
                "period": period,
                "start": start_time,
                "end": end_time,
                "course_code": course_code,
                "section": section_str,
                "room": room,
            }
        )

    return faculty_sessions


def write_faculty_timetables(
    all_sessions: Iterable[Any],
    faculty_conflicts: Iterable[FacultyConflict],
    output_path: str,
) -> str:
    """
    Write a per-faculty timetable workbook.

    New layout:
    - One sheet per faculty.
    - Separate PreMid and PostMid tables (each Time × Weekday).
    - Per-faculty summary at the bottom of each sheet.
    - A global SUMMARY sheet aggregating conflicts across all faculty.
    """
    faculty_sessions = _collect_faculty_sessions(all_sessions)
    if not faculty_sessions:
        # Nothing to write – avoid creating empty files
        return ""

    # Materialize conflicts once and build indices
    conflicts_list: List[FacultyConflict] = list(faculty_conflicts or [])
    conflict_index = _build_faculty_conflict_index(conflicts_list)

    # faculty -> period -> list[FacultyConflict]
    faculty_period_conflicts: Dict[str, Dict[str, List[FacultyConflict]]] = defaultdict(
        lambda: defaultdict(list)
    )
    for c in conflicts_list:
        faculty_name = getattr(c, "faculty_name", "UNKNOWN")
        time_slot_str = getattr(c, "time_slot", "")
        period = "UNKNOWN"
        if "(" in time_slot_str and ")" in time_slot_str:
            period_raw = time_slot_str.split("(", 1)[1].rstrip(")")
            period = _normalize_period(period_raw)
        faculty_period_conflicts[faculty_name][period].append(c)

    wb = Workbook()
    # Basic styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    time_font = Font(bold=True, size=10)
    course_font = Font(size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    conflict_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    days = list(WORKING_DAYS)

    def _write_period_block(
        ws,
        start_row: int,
        faculty_name: str,
        period_label: str,  # "PRE" or "POST"
        sessions_for_period: List[Dict[str, Any]],
    ) -> int:
        """
        Write one Time × Weekday table for the given period.
        Returns the next free row index after the table.
        """
        # Header row for the table
        header_row = start_row
        ws.cell(row=header_row, column=1, value="Time").font = time_font
        ws.cell(row=header_row, column=1).border = thin_border

        if not sessions_for_period:
            # No sessions in this period
            msg_row = header_row + 1
            ws.cell(row=msg_row, column=1, value="No sessions in this period").font = course_font
            return msg_row + 2

        # Create weekday columns
        col_map: Dict[str, int] = {}
        col = 2
        for day in days:
            cell = ws.cell(row=header_row, column=col, value=day)
            cell.font = time_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            col_map[day] = col
            col += 1

        # Collect unique time slots (start,end)
        time_keys: List[Tuple[dt_time, dt_time]] = []
        seen_time_keys = set()
        for s in sessions_for_period:
            key = (s["start"], s["end"])
            if key not in seen_time_keys:
                seen_time_keys.add(key)
                time_keys.append(key)

        time_keys.sort(key=lambda t: (t[0].hour, t[0].minute, t[1].hour, t[1].minute))

        # Map time slot to row index
        row_map: Dict[Tuple[dt_time, dt_time], int] = {}
        current_row = header_row + 1
        for start, end in time_keys:
            row_map[(start, end)] = current_row
            label = f"{start.strftime('%H:%M')}-{end.strftime('%H:%M')}"
            cell = ws.cell(row=current_row, column=1, value=label)
            cell.font = time_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            current_row += 1

        # Populate cells
        for s in sessions_for_period:
            row_idx = row_map.get((s["start"], s["end"]))
            col_idx = col_map.get(s["day"])
            if not row_idx or not col_idx:
                continue

            course_code = str(s["course_code"])
            section = s["section"]
            room = s.get("room") or ""

            cell = ws.cell(row=row_idx, column=col_idx)

            parts = [course_code]
            if section:
                parts.append(f"({section})")
            if room:
                parts.append(f"[{room}]")

            text = " ".join(parts)
            if cell.value:
                cell.value = f"{cell.value}\n{text}"
            else:
                cell.value = text

            cell.font = course_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

            # Highlight if there is a conflict for this faculty / slot / period
            conflict_key = (faculty_name, s["day"], period_label, s["start"], s["end"])
            conflicts_here = conflict_index.get(conflict_key)
            if conflicts_here:
                cell.fill = conflict_fill

        # Adjust column widths a bit
        ws.column_dimensions["A"].width = 14
        for c in range(2, 2 + len(days)):
            ws.column_dimensions[ws.cell(row=header_row, column=c).column_letter].width = 18

        # Leave one empty row after the table
        return current_row + 1

    first_sheet = True
    for faculty_name, sessions in sorted(faculty_sessions.items(), key=lambda x: x[0].lower()):
        if not sessions:
            continue

        if first_sheet:
            ws = wb.active
            first_sheet = False
        else:
            ws = wb.create_sheet()

        ws.title = _sanitize_sheet_name(faculty_name)

        # Title
        ws["A1"] = f"Faculty: {faculty_name}"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill

        # Partition this faculty's sessions into PreMid and PostMid
        pre_sessions: List[Dict[str, Any]] = []
        post_sessions: List[Dict[str, Any]] = []
        for s in sessions:
            if s["period"] == "PRE":
                pre_sessions.append(s)
            elif s["period"] == "POST":
                post_sessions.append(s)

        # PreMid block
        pre_label_row = 3
        ws.cell(row=pre_label_row, column=1, value="PreMid").font = time_font
        next_row = _write_period_block(
            ws,
            start_row=pre_label_row + 1,
            faculty_name=faculty_name,
            period_label="PRE",
            sessions_for_period=pre_sessions,
        )

        # PostMid block
        post_label_row = next_row + 1
        ws.cell(row=post_label_row, column=1, value="PostMid").font = time_font
        next_row = _write_period_block(
            ws,
            start_row=post_label_row + 1,
            faculty_name=faculty_name,
            period_label="POST",
            sessions_for_period=post_sessions,
        )

        # Per-faculty summary
        summary_row = next_row + 1
        per_fac_conf = faculty_period_conflicts.get(faculty_name, {})
        has_any_conflicts = any(per_fac_conf.get(p) for p in ("PRE", "POST"))
        if not has_any_conflicts:
            ws.cell(
                row=summary_row,
                column=1,
                value="Summary: NO conflicts for this faculty",
            ).font = course_font
        else:
            ws.cell(
                row=summary_row,
                column=1,
                value="Summary: Conflicts exist for this faculty",
            ).font = course_font
            row = summary_row + 1
            for per in ("PRE", "POST"):
                confs = per_fac_conf.get(per) or []
                if not confs:
                    continue
                label = "PreMid" if per == "PRE" else "PostMid"
                snippets = [str(getattr(c, "time_slot", "")) for c in confs[:2]]
                details = "; ".join(snippets)
                ws.cell(
                    row=row,
                    column=1,
                    value=f"{label}: {len(confs)} conflict(s){' - ' + details if details else ''}",
                ).font = course_font
                row += 1

    # Global SUMMARY sheet in the same workbook
    summary_ws = wb.create_sheet(title="SUMMARY")
    if not conflicts_list:
        summary_ws["A1"] = "[OK] NO FACULTY CONFLICTS DETECTED"
        summary_ws["A1"].font = header_font
    else:
        summary_ws["A1"] = "Faculty Conflict Summary"
        summary_ws["A1"].font = header_font
        header_row = 3
        headers = ["Faculty", "Period", "Has Conflicts?", "Details"]
        for idx, h in enumerate(headers, start=1):
            cell = summary_ws.cell(row=header_row, column=idx, value=h)
            cell.font = time_font

        row = header_row + 1
        all_faculties = set(faculty_sessions.keys()) | set(faculty_period_conflicts.keys())
        for faculty_name in sorted(all_faculties, key=lambda x: x.lower()):
            for per in ("PRE", "POST"):
                confs = faculty_period_conflicts.get(faculty_name, {}).get(per) or []
                has_conf = bool(confs)
                period_label = "PreMid" if per == "PRE" else "PostMid"
                summary_ws.cell(row=row, column=1, value=faculty_name).font = course_font
                summary_ws.cell(row=row, column=2, value=period_label).font = course_font
                summary_ws.cell(row=row, column=3, value="YES" if has_conf else "NO").font = course_font
                if has_conf:
                    snippets = [str(getattr(c, "time_slot", "")) for c in confs[:2]]
                    details = "; ".join(snippets)
                    summary_ws.cell(
                        row=row,
                        column=4,
                        value=f"{len(confs)} conflict(s): {details}",
                    ).font = course_font
                row += 1

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def write_faculty_conflict_summary(
    conflicts: Iterable[FacultyConflict],
    output_path: str,
) -> str:
    """
    Write a human-readable text summary of faculty conflicts, grouped by
    faculty and by period (PreMid/PostMid).
    """
    conflicts = list(conflicts or [])
    if not conflicts:
        # Still write a small OK file so the user knows checks ran
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("[OK] NO FACULTY CONFLICTS DETECTED\n")
            f.write("All faculty members are properly scheduled without double-booking.\n")
        return output_path

    # Group by faculty -> period
    grouped: Dict[str, Dict[str, List[FacultyConflict]]] = defaultdict(lambda: defaultdict(list))
    for c in conflicts:
        faculty = getattr(c, "faculty_name", "UNKNOWN")
        time_slot_str = getattr(c, "time_slot", "")
        period = "UNKNOWN"
        if "(" in time_slot_str and ")" in time_slot_str:
            period_raw = time_slot_str.split("(", 1)[1].rstrip(")")
            period = _normalize_period(period_raw)
        grouped[faculty][period].append(c)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("[FACULTY CONFLICT SUMMARY]\n")
        f.write("=" * 72 + "\n\n")
        for faculty in sorted(grouped.keys(), key=lambda x: x.lower()):
            f.write(f"Faculty: {faculty}\n")
            periods_for_faculty = grouped[faculty]
            for period in ("PRE", "POST"):
                conflicts_list = periods_for_faculty.get(period) or []
                label = "PreMid" if period == "PRE" else "PostMid"
                f.write(f"  Period: {label}  (Total conflicts: {len(conflicts_list)})\n")
                for idx, c in enumerate(conflicts_list, 1):
                    time_slot = getattr(c, "time_slot", "")
                    sessions = getattr(c, "conflicting_sessions", [])
                    f.write(f"    {idx}. {time_slot} -> {', '.join(sessions)}\n")
            f.write("\n")

    return output_path

