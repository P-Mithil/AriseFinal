#!/usr/bin/env python3
"""
Minor Timetable Generator

Reads minor course data from DATA/INPUT/Minor.xlsx and creates a separate
minor timetable workbook with two sheets:
- "minor t.t sem3"
- "minor t.t sem5"

Each sheet:
- Rows: Monday–Friday
- Columns: 7:30–9:00 and 18:00–19:30
- Each course gets 2 slots:
  - On different days
  - In the same time band (both morning or both evening)
If all slots are already occupied, courses are allowed to share a slot.
"""

import os
from collections import defaultdict
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


DAYS: List[str] = ["Mon", "Tue", "Wed", "Thu", "Fri"]
TIME_SLOTS: List[str] = ["7:30-9:00", "18:00-19:30"]


def get_project_root() -> str:
    """Return the project root directory (two levels above this file)."""
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.dirname(here)


def load_minor_courses() -> Dict[int, List[str]]:
    """
    Load minor courses from DATA/INPUT/Minor.xlsx.

    Expected logical columns:
    - MINOR COURSE: course name/code
    - SEMESTER: semester number (we care about 3 and 5)

    Returns:
        Dict[semester, List[course_name]]
    """
    project_root = get_project_root()
    input_path = os.path.join(project_root, "iiitdwd_timetable_v2", "DATA", "INPUT", "Minor.xlsx")

    if not os.path.exists(input_path):
        # Fallback to relative DATA/INPUT if script is moved
        input_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "DATA", "INPUT", "Minor.xlsx")

    df = pd.read_excel(input_path)

    # Make column access robust to minor variations in header spacing/case
    normalized_cols = {str(c).strip().upper(): c for c in df.columns}

    course_col_key = "MINOR COURSE"
    sem_col_key = "SEMESTER"

    if course_col_key not in normalized_cols or sem_col_key not in normalized_cols:
        raise ValueError(
            f"Minor.xlsx is missing required columns. "
            f"Found columns: {list(df.columns)}; "
            f"expected at least '{course_col_key}' and '{sem_col_key}'."
        )

    course_col = normalized_cols[course_col_key]
    sem_col = normalized_cols[sem_col_key]

    courses_by_sem: Dict[int, List[str]] = defaultdict(list)

    for _, row in df.iterrows():
        sem_val = row[sem_col]
        try:
            sem = int(sem_val)
        except (TypeError, ValueError):
            continue

        if sem not in (3, 5):
            continue

        name = str(row[course_col]).strip()
        if not name:
            continue

        courses_by_sem[sem].append(name)

    return courses_by_sem


def schedule_courses_for_semester(courses: List[str]) -> Dict[Tuple[str, str], List[str]]:
    """
    Schedule courses for a single semester.

    Rules:
    - Each course gets exactly 2 slots.
    - Slots must be on different days.
    - Slots must be in the same time band (both morning or both evening).
    - Slots may be shared if needed.

    Returns:
        Dict[(day, time_slot)] -> List[course_names]
    """
    grid: Dict[Tuple[str, str], List[str]] = defaultdict(list)

    if not courses:
        return grid

    # Round-robin pointers per time band
    morning_pointer = 0  # for TIME_SLOTS[0]
    evening_pointer = 0  # for TIME_SLOTS[1]

    for idx, course in enumerate(courses):
        # Alternate bands for some balance: odd -> morning, even -> evening
        use_morning = (idx % 2 == 0)
        if use_morning:
            band_slot = TIME_SLOTS[0]
            base_pointer = morning_pointer
        else:
            band_slot = TIME_SLOTS[1]
            base_pointer = evening_pointer

        # First day
        day1_idx = base_pointer % len(DAYS)
        day1 = DAYS[day1_idx]

        # Second day: shift by 2 days for better spread, but ensure different day
        day2_idx = (base_pointer + 2) % len(DAYS)
        if day2_idx == day1_idx:
            day2_idx = (day1_idx + 1) % len(DAYS)
        day2 = DAYS[day2_idx]

        grid[(day1, band_slot)].append(course)
        grid[(day2, band_slot)].append(course)

        # Advance pointer for this band
        if use_morning:
            morning_pointer += 1
        else:
            evening_pointer += 1

    return grid


def write_minor_timetable(courses_by_sem: Dict[int, List[str]]) -> str:
    """
    Write minor timetable workbook to DATA/OUTPUT/Minor_timetable.xlsx.

    Returns:
        Full path of the generated workbook.
    """
    project_root = get_project_root()
    output_dir = os.path.join(project_root, "iiitdwd_timetable_v2", "DATA", "OUTPUT")

    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    output_path = os.path.join(output_dir, "Minor_timetable.xlsx")

    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for sem in (3, 5):
        sheet_name = f"minor t.t sem{sem}"
        ws = wb.create_sheet(title=sheet_name)

        # Header row
        ws["A1"] = "Day"
        ws["B1"] = TIME_SLOTS[0]
        ws["C1"] = TIME_SLOTS[1]

        for cell in ("A1", "B1", "C1"):
            ws[cell].font = header_font
            ws[cell].alignment = center_align

        # Schedule grid for this semester
        sem_courses = courses_by_sem.get(sem, [])
        grid = schedule_courses_for_semester(sem_courses)

        # Fill days and slots
        for row_idx, day in enumerate(DAYS, start=2):
            ws.cell(row=row_idx, column=1, value=day)
            ws.cell(row=row_idx, column=1).alignment = center_align

            for col_idx, slot in enumerate(TIME_SLOTS, start=2):
                cell = ws.cell(row=row_idx, column=col_idx)
                courses_here = grid.get((day, slot), [])
                if courses_here:
                    cell.value = "\n".join(courses_here)
                else:
                    cell.value = ""
                cell.alignment = center_align

        # Adjust column widths a bit
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25

    wb.save(output_path)
    return output_path


def generate_minor_timetable() -> str:
    """
    High-level entry point to generate minor timetable.

    Returns:
        Path to the generated Excel file.
    """
    courses_by_sem = load_minor_courses()
    output_path = write_minor_timetable(courses_by_sem)
    return output_path


if __name__ == "__main__":
    path = generate_minor_timetable()
    print(f"Minor timetable generated at: {path}")

