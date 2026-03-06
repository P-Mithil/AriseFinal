"""
Classroom-wise timetable writer and clash reporter.

Creates a separate Excel workbook where each sheet contains the timetable
for a single classroom/lab (all days, both periods), and marks any
same-period double bookings in that room. A SUMMARY sheet aggregates
clash information per room.
"""

from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime, time as dt_time
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from utils.data_models import TimeBlock
from config.schedule_config import WORKING_DAYS


def _normalize_period(raw: str) -> str:
    """Normalize period strings to 'PRE' or 'POST'."""
    if not raw:
        return "PRE"
    val = str(raw).strip().upper()
    if val in ("PREMID", "PRE"):
        return "PRE"
    if val in ("POSTMID", "POST"):
        return "POST"
    return val


def _sanitize_room_code(room: Any) -> str:
    """Normalize room code to a stable uppercase string."""
    if room is None:
        return ""
    return str(room).strip().upper()


def _sanitize_room_sheet_name(room_code: str) -> str:
    """
    Create a safe Excel sheet name for a room.

    - Prefix with 'ROOM_'
    - Replace spaces with underscores
    - Remove invalid characters: []:*?/\\
    - Truncate to 31 characters (Excel limit)
    """
    base = room_code.strip().replace(" ", "_") if room_code else "UNKNOWN"
    invalid_chars = "[]:*?/\\"
    for ch in invalid_chars:
        base = base.replace(ch, "")
    if not base:
        base = "UNKNOWN"
    name = f"ROOM_{base}"
    return name[:31]


def _collect_room_sessions(
    all_sessions: Iterable[Any],
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Normalize all sessions into a per-room list of dicts with:
    - day
    - period
    - start_time
    - end_time
    - course_code
    - section_str
    - faculty_name
    """
    room_sessions: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for session in all_sessions or []:
        # Extract core fields from both dict- and object-style sessions
        if isinstance(session, dict):
            block: TimeBlock = session.get("time_block")
            period_raw = session.get("period", "PRE")
            course_code = session.get("course_code", "UNKNOWN")
            sections = session.get("sections") or []
            room = session.get("room")
            faculty = session.get("instructor") or session.get("faculty") or ""

            if sections and isinstance(sections, (list, tuple)):
                section_str = ",".join(str(s) for s in sections)
            else:
                section_str = str(sections) if sections else ""
        else:
            block = getattr(session, "block", None)
            period_raw = getattr(session, "period", "PRE")
            course_code = getattr(session, "course_code", "UNKNOWN")
            section = getattr(session, "section", "")
            room = getattr(session, "room", None)
            faculty = getattr(session, "faculty", "") or ""
            section_str = str(section)

        # Elective basket placeholders (ELECTIVE_BASKET_*) are virtual slots,
        # not real room occupancy; concrete elective courses use their own rooms.
        if isinstance(course_code, str) and course_code.startswith("ELECTIVE_BASKET_"):
            continue

        if not block or not getattr(block, "start", None) or not getattr(block, "end", None):
            continue

        room_code_raw = _sanitize_room_code(room)
        if not room_code_raw:
            # Skip sessions without a concrete room – no way to build a room timetable
            continue

        period = _normalize_period(period_raw)

        # Multi-lab assignments may appear as 'L105, L206'. Treat each
        # physical room separately so per-room sheets and clash counts
        # are computed per lab, not for the combined label.
        room_codes = [p.strip() for p in room_code_raw.split(",") if p.strip()] or [room_code_raw]

        for rc in room_codes:
            room_sessions[rc].append(
                {
                    "day": block.day,
                    "period": period,
                    "start": block.start,
                    "end": block.end,
                    "course_code": course_code,
                    "section": section_str,
                    "faculty": faculty,
                }
            )

    return room_sessions


def _build_room_conflicts(
    room_sessions: Dict[str, List[Dict[str, Any]]],
) -> Tuple[
    Dict[str, Dict[str, Dict[Tuple[str, dt_time, dt_time], List[Dict[str, Any]]]]],
    Dict[str, int],
]:
    """
    Detect intra-room clashes (same room, same day, same period, overlapping time).

    Returns:
        room_conflicts: room -> period -> (day,start,end) -> list[sessions]
        room_clash_counts: room -> int (number of conflicting time slots)
    """
    room_conflicts: Dict[str, Dict[str, Dict[Tuple[str, dt_time, dt_time], List[Dict[str, Any]]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(list))
    )
    room_clash_counts: Dict[str, int] = defaultdict(int)

    for room, sessions in room_sessions.items():
        # Treat lab rooms (e.g. L105, L206) as non-classroom for clash purposes.
        # Lab sharing/parallel use is allowed; we don't want them flagged red.
        if isinstance(room, str) and room.upper().startswith("L"):
            room_clash_counts[room] += 0
            continue

        # Group by period and day for overlap checks
        per_period_day: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
        for s in sessions:
            key = (s["period"], str(s["day"]))
            per_period_day[key].append(s)

        for (period, day), day_sessions in per_period_day.items():
            # Sort by start time
            ordered = sorted(day_sessions, key=lambda x: (x["start"], x["end"]))
            n = len(ordered)
            for i in range(n):
                s1 = ordered[i]
                for j in range(i + 1, n):
                    s2 = ordered[j]
                    # If next session starts after this one ends, no further overlaps possible
                    if s2["start"] >= s1["end"]:
                        break
                    # Overlap condition: times intersect
                    if not (s1["end"] <= s2["start"] or s2["end"] <= s1["start"]):
                        # Match Phase 8 rule: only treat overlaps between DIFFERENT courses
                        c1 = (s1.get("course_code") or "").split("-")[0]
                        c2 = (s2.get("course_code") or "").split("-")[0]
                        if c1 == c2:
                            continue
                        key1 = (day, s1["start"], s1["end"])
                        key2 = (day, s2["start"], s2["end"])
                        room_conflicts[room][period][key1].append(s1)
                        room_conflicts[room][period][key1].append(s2)
                        room_conflicts[room][period][key2].append(s1)
                        room_conflicts[room][period][key2].append(s2)

            # Count distinct conflicting time slots for this room
            clash_slots = room_conflicts[room][period]
            room_clash_counts[room] += len(clash_slots)

    return room_conflicts, room_clash_counts


def write_classroom_timetables(
    all_sessions: Iterable[Any],
    output_path: str,
) -> str:
    """
    Write a per-classroom timetable workbook.

    Layout:
      - One sheet per room.
      - For each room: separate PreMid and PostMid tables (Time × Weekday).
      - Cells participating in clashes within that room are highlighted.
      - Per-room summary at the bottom of each sheet.
      - Global SUMMARY sheet aggregating clash info across rooms.
    """
    room_sessions = _collect_room_sessions(all_sessions)
    if not room_sessions:
        # Nothing to write – avoid creating empty files
        return ""

    room_conflicts, room_clash_counts = _build_room_conflicts(room_sessions)

    wb = Workbook()
    # Basic styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    time_font = Font(bold=True, size=10)
    cell_font = Font(size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    conflict_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    days = list(WORKING_DAYS)

    def _write_period_block(
        ws,
        start_row: int,
        room_code: str,
        period_label: str,  # "PRE" or "POST"
        sessions_for_period: List[Dict[str, Any]],
        conflicts_for_period: Dict[Tuple[str, dt_time, dt_time], List[Dict[str, Any]]],
    ) -> int:
        """
        Write one Time × Weekday table for the given period.
        Returns the next free row index after the table.
        """
        # Sub-header
        ws.cell(row=start_row, column=1, value="PreMid" if period_label == "PRE" else "PostMid").font = header_font
        ws.cell(row=start_row, column=1).fill = header_fill
        start_row += 1

        header_row = start_row
        ws.cell(row=header_row, column=1, value="Time").font = time_font
        ws.cell(row=header_row, column=1).border = thin_border

        if not sessions_for_period:
            # No sessions in this period
            msg_row = header_row + 1
            ws.cell(row=msg_row, column=1, value="No sessions in this period").font = cell_font
            return msg_row + 2

        # Create weekday columns
        col_map: Dict[str, int] = {}
        col = 2
        for day in days:
            cell = ws.cell(row=header_row, column=col, value=day)
            cell.font = time_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            col_map[day] = col
            col += 1

        # Collect unique time slots (start,end)
        time_keys: List[Tuple[dt_time, dt_time]] = []
        seen_time_keys = set()
        for s in sessions_for_period:
            key = (s["start"], s["end"])
            if key not in seen_time_keys:
                seen_time_keys.add(key)
                time_keys.append(key)

        time_keys.sort()

        # Build index by (day, start, end)
        index: Dict[Tuple[str, dt_time, dt_time], List[Dict[str, Any]]] = defaultdict(list)
        for s in sessions_for_period:
            key = (str(s["day"]), s["start"], s["end"])
            index[key].append(s)

        current_row = header_row + 1
        for start_time, end_time in time_keys:
            time_label = f"{start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')}"
            time_cell = ws.cell(row=current_row, column=1, value=time_label)
            time_cell.font = time_font
            time_cell.border = thin_border

            for day, col in col_map.items():
                key = (day, start_time, end_time)
                sessions_here = index.get(key, [])
                if not sessions_here:
                    continue

                # Build cell text: combine course, section, and faculty if available
                parts = []
                for s in sessions_here:
                    text = f"{s['course_code']} ({s['section']})"
                    if s.get("faculty"):
                        text = f"{text} [{s['faculty']}]"
                    parts.append(text)
                value = "\n".join(parts)

                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = cell_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

                # Highlight if this time slot in this room is part of a clash
                if key in conflicts_for_period:
                    cell.fill = conflict_fill

            current_row += 1

        return current_row + 1

    # Per-room sheets
    for room_code, sessions in sorted(room_sessions.items(), key=lambda x: x[0]):
        sheet_name = _sanitize_room_sheet_name(room_code)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Title row
        ws["A1"] = f"Classroom timetable: {room_code}"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill

        current_row = 3

        # Split sessions by period
        by_period: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for s in sessions:
            by_period[s["period"]].append(s)

        conflicts_for_room = room_conflicts.get(room_code, {})

        # Always write PRE then POST for determinism
        for period_label in ("PRE", "POST"):
            period_sessions = by_period.get(period_label, [])
            period_conflicts = conflicts_for_room.get(period_label, {})
            current_row = _write_period_block(
                ws,
                current_row,
                room_code,
                period_label,
                period_sessions,
                period_conflicts,
            )
            current_row += 1

        # Per-room summary at the bottom
        total_sessions = len(sessions)
        clash_count = room_clash_counts.get(room_code, 0)
        status = "NO CLASHES" if clash_count == 0 else "CLASHES PRESENT"

        summary_row = current_row + 1
        ws.cell(row=summary_row, column=1, value="Total sessions in this room").font = time_font
        ws.cell(row=summary_row, column=2, value=total_sessions).font = cell_font

        summary_row += 1
        ws.cell(row=summary_row, column=1, value="Total clashes (same room, same period)").font = time_font
        ws.cell(row=summary_row, column=2, value=clash_count).font = cell_font

        summary_row += 1
        ws.cell(row=summary_row, column=1, value="Status").font = time_font
        status_cell = ws.cell(row=summary_row, column=2, value=status)
        status_cell.font = cell_font
        if clash_count == 0:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
        else:
            status_cell.fill = conflict_fill

    # Global SUMMARY sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        summary_ws = wb.active
        summary_ws.title = "SUMMARY"
    else:
        summary_ws = wb.create_sheet(title="SUMMARY")

    header = ["Room", "Total Sessions", "Clash Count", "Status"]
    for col_idx, title in enumerate(header, start=1):
        cell = summary_ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill

    row_idx = 2
    for room_code in sorted(room_sessions.keys()):
        total_sessions = len(room_sessions[room_code])
        clash_count = room_clash_counts.get(room_code, 0)
        status = "NO CLASH" if clash_count == 0 else "CLASH"

        summary_ws.cell(row=row_idx, column=1, value=room_code).font = cell_font
        summary_ws.cell(row=row_idx, column=2, value=total_sessions).font = cell_font
        summary_ws.cell(row=row_idx, column=3, value=clash_count).font = cell_font
        status_cell = summary_ws.cell(row=row_idx, column=4, value=status)
        status_cell.font = cell_font
        if clash_count == 0:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        row_idx += 1

    # Ensure directory exists
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path

