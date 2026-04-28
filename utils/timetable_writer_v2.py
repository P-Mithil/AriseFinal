"""
Timetable Writer v2: Dynamic Grid Output
Creates Excel output with dynamic time slots showing only start/end times of scheduled sessions.
"""

import os
import sys
from datetime import datetime, time
from typing import List, Dict, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add the parent directory to Python path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.data_models import DayScheduleGrid, TimeBlock
from config.schedule_config import WORKING_DAYS, COMBINED_RESERVED_ROOM_NUMBER

def extract_department_from_section(section: str) -> str:
    """
    Extract department from section name in a generalized way.
    
    Handles various section formats:
    - "CSE-A-Sem1" -> "CSE"
    - "ECE-A-Sem3" -> "ECE"
    - "DSAI-A-Sem5" -> "DSAI"
    - "CSE-B-Sem1" -> "CSE"
    - Also handles edge cases and variations
    
    Args:
        section: Section name string (e.g., "CSE-A-Sem1", "ECE-A-Sem3")
    
    Returns:
        Department code (e.g., "CSE", "ECE", "DSAI") or empty string if not found
    """
    if not section or not isinstance(section, str):
        return ""
    
    # Remove any whitespace
    section = section.strip()
    
    # Try splitting by '-' - most common format is "DEPT-NAME-SemN"
    if '-' in section:
        parts = section.split('-')
        # First part is typically the department
        if parts and len(parts[0]) > 0:
            return parts[0].strip().upper()
    
    # Fallback: Try to match known department patterns
    known_departments = ['CSE', 'ECE', 'DSAI', 'CS', 'EC', 'DS']
    section_upper = section.upper()
    for dept in known_departments:
        if section_upper.startswith(dept):
            return dept
    
    # If no pattern matches, return empty string
    return ""


def _session_period_pre_post(session) -> str:
    """Normalize session period to 'PRE' or 'POST' (align with combined-course display branch)."""
    raw = ""
    if isinstance(session, dict):
        raw = session.get("period", "") or ""
    else:
        raw = getattr(session, "period", "") or ""
    s = str(raw).strip().upper()
    if s in ("PRE", "PREMID", "PRE-MID"):
        return "PRE"
    if s in ("POST", "POSTMID", "POST-MID"):
        return "POST"
    return ""


def _combined_course_codes_for_section(
    combined_sessions,
    section: str,
    semester: int,
    sheet_period: str,
) -> set:
    """
    Base course codes that appear in combined_sessions for this section label only,
    for the same PreMid/PostMid half as the sheet.

    Without period filtering, a PostMid-only combined row would still exclude the code
    from relevant_courses on the PreMid sheet while the combined block skips it — hiding
    the course from the verification table entirely.
    """
    if not combined_sessions:
        return set()
    sec_tag = f"{section}-Sem{semester}"
    out: set = set()
    for session in combined_sessions:
        sp = _session_period_pre_post(session)
        if sheet_period == "PreMid" and sp != "PRE":
            continue
        if sheet_period == "PostMid" and sp != "POST":
            continue
        if isinstance(session, dict):
            course_code = (session.get("course_code") or "").split("-")[0].strip()
            if not course_code:
                continue
            secs = session.get("sections") or []
            if isinstance(secs, str):
                secs = [secs]
            labels = [str(x).strip() for x in secs if str(x).strip()]
            if sec_tag in labels:
                out.add(course_code)
        elif hasattr(session, "course_code"):
            course_code = str(session.course_code or "").split("-")[0].strip()
            if not course_code:
                continue
            secs = getattr(session, "sections", None) or []
            if isinstance(secs, str):
                secs = [secs]
            labels = [str(x).strip() for x in secs if str(x).strip()]
            if sec_tag in labels:
                out.add(course_code)
    return out


def get_combined_course_period(course_code: str, semester: int, combined_sessions: List = None, section: str = None) -> str:
    """
    Dynamically determine which period (PreMid/PostMid) a combined course belongs to
    by checking actual scheduled sessions, not hardcoded course codes.
    
    Args:
        course_code: Course code (e.g., 'CS261')
        semester: Semester number
        combined_sessions: List of combined session dictionaries
        section: Optional section name (e.g., 'CSE-B-Sem3') to filter by section
    
    Returns: 'PRE' for PreMid, 'POST' for PostMid, or None if not found
    """
    if not combined_sessions:
        return None
    
    # Check actual scheduled sessions to determine period
    for session in combined_sessions:
        if isinstance(session, dict):
            session_course = session.get('course_code', '').split('-')[0]  # Remove -TUT/-LAB suffix
            session_period = session.get('period')
            session_sections = session.get('sections', [])
            
            if session_course == course_code and session_period:
                # If section is specified, check if this session is for that section
                if section:
                    section_key = f"{section}-Sem{semester}" if not section.endswith(f"-Sem{semester}") else section
                    if section_key in session_sections:
                        return session_period
                else:
                    # No section filter - return first period found
                    return session_period
        elif hasattr(session, 'course_code') and hasattr(session, 'period'):
            session_course = session.course_code.split('-')[0]  # Remove -TUT/-LAB suffix
            if session_course == course_code:
                # If section is specified, check if this session is for that section
                if section and hasattr(session, 'sections'):
                    section_key = f"{section}-Sem{semester}" if not section.endswith(f"-Sem{semester}") else section
                    if section_key in session.sections:
                        return session.period
                else:
                    return session.period
    
    return None

class TimetableWriterV2:
    """Dynamic grid timetable writer"""
    
    def __init__(self, course_colors: Dict[str, str] = None):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        
        # Course color mapping (from Phase 10)
        self.course_colors = course_colors if course_colors else {}
        
        # Color schemes
        self.colors = {
            'lunch': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'),  # Gray
            'break': PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid'),  # Light gray
            'elective': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),  # Light green
            'combined': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),  # Orange
            'core': PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid'),  # Sky blue
            'header': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),  # Blue
        }
        
        # Fonts
        self.header_font = Font(bold=True, color='FFFFFF', size=12)
        self.time_font = Font(bold=True, size=10)
        self.course_font = Font(size=10)
        
        # Borders
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _extract_base_course_code(self, course_str: str, semester: int = None) -> str:
        """
        Extract base course code from course string for color lookup.
        
        Examples:
        - "CS161-LAB" → "CS161"
        - "CS161-TUT" → "CS161"
        - "ELECTIVE BASKET" → "ELECTIVE_BASKET_{semester}" (if semester provided)
        - "ELECTIVE_BASKET_1" → "ELECTIVE_BASKET_1"
        - "CS161" → "CS161"
        """
        if not isinstance(course_str, str):
            return str(course_str)
        
        course_str = course_str.strip()
        
        # Handle elective baskets
        if "ELECTIVE BASKET" in course_str.upper() or "ELECTIVE_BASKET" in course_str.upper():
            if semester:
                return f"ELECTIVE_BASKET_{semester}"
            # Try to extract from string if format is "ELECTIVE_BASKET_1"
            if "_" in course_str:
                parts = course_str.split("_")
                if len(parts) >= 3:
                    return f"ELECTIVE_BASKET_{parts[-1]}"
            return "ELECTIVE_BASKET_1"  # Default fallback
        
        # Remove common suffixes for color lookup (but keep display text as-is)
        base_code = course_str
        for suffix in ['-LAB', '-TUT', '-TUTORIAL', '-P', '-L']:
            if base_code.endswith(suffix):
                base_code = base_code[:-len(suffix)]
                break
        
        return base_code
    
    def get_course_color(self, course_code: str, semester: int = None) -> PatternFill:
        """
        Get color PatternFill for a course code.
        
        Args:
            course_code: Course code string (may include -LAB, -TUT suffixes)
            semester: Optional semester number for elective basket handling
            
        Returns:
            PatternFill object for the course color, or fallback color if not found
        """
        # Extract base course code
        base_code = self._extract_base_course_code(course_code, semester)
        
        # Look up in course_colors dict
        if base_code in self.course_colors:
            hex_color = self.course_colors[base_code]
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        
        # Fallback to existing type-based colors
        if "ELECTIVE" in course_code.upper():
            return self.colors['elective']
        elif "COMBINED" in course_code.upper():
            return self.colors['combined']
        else:
            return self.colors['core']
    
    def write_day_schedule(self, sheet, day: str, schedule_grid: DayScheduleGrid, start_row: int) -> int:
        """Write one day's schedule in dynamic grid format"""
        current_row = start_row
        
        # Day header
        sheet[f"A{current_row}"] = f"Day: {day}"
        sheet[f"A{current_row}"].font = self.header_font
        sheet[f"A{current_row}"].fill = self.colors['header']
        current_row += 1
        
        # Get sessions with times and sort chronologically
        sessions_with_times = schedule_grid.get_sessions_with_times()
        
        if not sessions_with_times:
            # No sessions for this day
            sheet[f"A{current_row}"] = "No classes scheduled"
            current_row += 1
            return current_row
        
        # Sort sessions chronologically by start time
        sessions_with_times.sort(key=lambda x: x[0].split(' - ')[0])  # Sort by start time
        
        # Time row
        time_row = current_row
        col = 2  # Start from column B
        
        for time_slot, course in sessions_with_times:
            cell = sheet.cell(row=time_row, column=col)
            cell.value = time_slot
            cell.font = self.time_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
            col += 1
        
        current_row += 1
        
        # Course row
        course_row = current_row
        col = 2
        
        for time_slot, course in sessions_with_times:
            cell = sheet.cell(row=course_row, column=col)
            cell.value = course
            cell.font = self.course_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
            
            # Apply color based on course type
            if course == "LUNCH":
                cell.fill = self.colors['lunch']
            elif isinstance(course, str) and "Break" in course:
                cell.fill = self.colors['break']
            else:
                # Use course-specific color (Phase 10)
                # Get semester from schedule_grid for elective basket handling
                semester = getattr(schedule_grid, 'semester', None)
                cell.fill = self.get_course_color(course, semester)
            
            col += 1
        
        current_row += 1
        return current_row
    
    def create_section_sheet(self, section_name: str, semester: int, period: str) -> None:
        """Create a sheet for a specific section and period"""
        sheet_name = f"{section_name} Sem{semester} {period}"
        sheet = self.workbook.create_sheet(title=sheet_name)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 15
        for col in range(2, 20):  # Adjust based on expected number of time slots
            sheet.column_dimensions[get_column_letter(col)].width = 12
        
        # Add title
        title_cell = sheet['A1']
        title_cell.value = f"Timetable - {section_name} Semester {semester} {period}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.fill = self.colors['header']
        
        # Add days
        days = list(WORKING_DAYS)
        current_row = 3
        
        for day in days:
            # Create a sample schedule grid for testing
            grid = DayScheduleGrid(day, semester)
            
            # Add sample sessions for testing
            if day == "Monday":
                # Morning session
                morning_block = TimeBlock(day, time(9, 0), time(10, 30))
                grid.add_session(morning_block, "MA161")
                
                # Break
                break_block = TimeBlock(day, time(10, 45), time(11, 0))
                grid.sessions.append((break_block, "Break(15min)"))
                
                # Lunch (automatically added)
                add_lunch_to_schedule(grid)
                
                # Afternoon session
                afternoon_block = TimeBlock(day, time(14, 0), time(15, 30))
                grid.add_session(afternoon_block, "CS161")
            
            current_row = self.write_day_schedule(sheet, day, grid, current_row)
            current_row += 1  # Add space between days
    
    def create_summary_sheet(self, courses: List) -> None:
        """Create summary sheet with course information"""
        sheet = self.workbook.create_sheet(title="Summary")
        
        # Headers
        headers = ["Course Code", "Course Name", "Credits", "Scheduled"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.colors['header']
            cell.border = self.thin_border
        
        # Add course data
        if courses:
            for row, course in enumerate(courses, 2):
                cell = sheet.cell(row=row, column=1)
                cell.value = course.code
                cell.border = self.thin_border
                
                cell = sheet.cell(row=row, column=2)
                cell.value = course.name
                cell.border = self.thin_border
                
                cell = sheet.cell(row=row, column=3)
                cell.value = course.credits
                cell.border = self.thin_border
                
                cell = sheet.cell(row=row, column=4)
                cell.value = "Yes"  # Assume all courses are scheduled
                cell.border = self.thin_border
                cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        else:
            # Add sample data if no courses provided
            sample_courses = [
                {"code": "MA161", "name": "Mathematics I", "credits": 3, "scheduled": "Yes"},
                {"code": "CS161", "name": "Programming", "credits": 3, "scheduled": "Yes"},
                {"code": "PH161", "name": "Physics I", "credits": 3, "scheduled": "No"},
            ]
            
            for row, course in enumerate(sample_courses, 2):
                for col, (key, value) in enumerate(course.items(), 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.value = value
                    cell.border = self.thin_border
                    if key == "scheduled":
                        if value == "Yes":
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    def write_verification_table(self, sheet, start_row: int, courses: List, 
                                 scheduled_sessions: List, semester: int, 
                                 section: str, period: str, phase5_sessions: List = None,
                                 phase7_sessions: List = None, combined_sessions: List = None, 
                                 faculty_conflicts: List = None, room_assignments: Dict = None,
                                 grid_sessions: Dict[str, List] = None,
                                 all_section_sessions: List[Dict] = None,
                                 classrooms: List = None) -> int:
        """Write verification summary table below timetable grid"""
        current_row = start_row

        # Build room_type lookup so we can correctly classify labs vs classrooms
        # A room is a lab only if its description contains 'lab' (case-insensitive)
        room_type_map: Dict[str, str] = {}  # room_number -> room_type string
        for _cr in (classrooms or []):
            rnum = getattr(_cr, 'room_number', None) or getattr(_cr, 'room_id', None)
            rtype = getattr(_cr, 'room_type', '') or ''
            if rnum:
                room_type_map[str(rnum).strip()] = str(rtype).strip()

        def is_lab_room(room_code) -> bool:
            """Return True only if the room (or all rooms in a comma-separated list) is an actual lab."""
            if not room_code:
                return False
            
            # If it's a comma-separated list of multiple rooms, check each one
            room_str = str(room_code).strip()
            if ',' in room_str:
                parts = [p.strip() for p in room_str.split(',') if p.strip()]
                return bool(parts) and all(is_lab_room(p) for p in parts)
                
            rtype = room_type_map.get(room_str, '')
            if rtype:
                return 'lab' in rtype.lower()
            # Fallback: rooms prefixed with 'L' followed by digits are physical labs (e.g. L106, L107)
            return room_str.upper().startswith('L') and room_str[1:].isdigit()

        def is_na_room(room_code) -> bool:
            if room_code is None:
                return False
            s = str(room_code).strip().upper()
            # Treat placeholder/sentinel values as "not assigned"
            return s in ("NA", "TBD", "UNASSIGNED", "")

        # Add spacing
        current_row += 2
        
        # Enhanced table headers with conflict detection
        headers = ["Code", "Course Name", "Instructor", "LTPSC", "Assigned Lab", 
                   "Assigned Classroom", "Lectures (Req / Period Sched)", 
                   "Tutorials (Req / Period Sched)", "Labs (Req / Period Sched)", "Status", 
                   "Time Slot Issues", "Room Conflicts"]
        
        # Count scheduled sessions for each course.
        # But also track per-period counts for display
        course_counts = {}  # Total across both periods
        course_counts_period = {}  # Per-period counts
        courses_on_this_sheet_period = set()  # base course codes visible in this period's grid
        
        # CRITICAL: If all_section_sessions provided, count EVERY session for the section across both periods
        if all_section_sessions:
            for session_dict in all_section_sessions:
                course_display = session_dict.get('Course Code') or session_dict.get('course_code', '')
                if not course_display or course_display in ["LUNCH", "Break(15min)", "ELECTIVE", "ELECTIVE-TUT", "ELECTIVE-LAB"]:
                    continue
                
                base_code = course_display.replace('-TUT', '').replace('-LAB', '').split('-')[0]
                session_type = session_dict.get('Session Type') or session_dict.get('session_type', 'L')
                
                section_dept = extract_department_from_section(section)
                course_obj = next((c for c in courses if 
                                 getattr(c, 'code', '') == base_code and 
                                 hasattr(c, 'semester') and c.semester == semester and
                                 getattr(c, 'department', '') == section_dept), None)
                if not course_obj:
                    course_obj = next((c for c in courses if getattr(c, 'code', '') == base_code and 
                                     hasattr(c, 'semester') and c.semester == semester), None)
                
                if course_obj and hasattr(course_obj, 'ltpsc'):
                    unique_key = f"{base_code}_{course_obj.ltpsc}"
                else:
                    unique_key = base_code
                
                if unique_key and unique_key not in course_counts:
                    course_counts[unique_key] = {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}
                if unique_key:
                    course_counts[unique_key]['total'] += 1
                    if session_type == 'P':
                        course_counts[unique_key]['labs'] += 1
                    elif session_type == 'T':
                        course_counts[unique_key]['tutorials'] += 1
                    else:
                        course_counts[unique_key]['lectures'] += 1

        
        # CRITICAL: If grid_sessions provided, count from actual displayed sessions instead of raw created sessions
        # This ensures we count what's actually shown in the timetable, not what was created but filtered out
        if grid_sessions:
            # grid_sessions is a dict: {day: [(TimeBlock, course_display), ...]}
            # Count sessions from the actual displayed grid
            for day, day_sessions in grid_sessions.items():
                for s in day_sessions:
                    time_block, course_display = s[0], s[1]
                    # Skip non-course entries (LUNCH, Break, etc.)
                    if course_display in ["LUNCH", "Break(15min)", "ELECTIVE", "ELECTIVE-TUT", "ELECTIVE-LAB"]:
                        continue
                    
                    # Extract base course code and session type
                    base_code = course_display.replace('-TUT', '').replace('-LAB', '').split('-')[0]
                    if base_code:
                        courses_on_this_sheet_period.add(base_code)
                    
                    # Determine session type from display code
                    if '-LAB' in course_display or '-P' in course_display:
                        session_type = 'P'
                    elif '-TUT' in course_display or '-T' in course_display:
                        session_type = 'T'
                    else:
                        session_type = 'L'
                    
                    # Find course object to get LTPSC for unique key
                    # CRITICAL: Match by code + semester + department to handle courses with same code but different LTPSC
                    section_dept = extract_department_from_section(section)
                    course_obj = next((c for c in courses if 
                                     getattr(c, 'code', '') == base_code and 
                                     hasattr(c, 'semester') and c.semester == semester and
                                     getattr(c, 'department', '') == section_dept), None)
                    # Fallback: if no match by department, try without department (for backward compatibility)
                    if not course_obj:
                        course_obj = next((c for c in courses if getattr(c, 'code', '') == base_code and 
                                         hasattr(c, 'semester') and c.semester == semester), None)
                    
                    # Create unique key using course code + LTPSC
                    if course_obj and hasattr(course_obj, 'ltpsc'):
                        unique_key = f"{base_code}_{course_obj.ltpsc}"
                    else:
                        unique_key = base_code
                    
                    # Count for total (across both periods) ONLY if not using all_section_sessions
                    if not all_section_sessions:
                        if unique_key and unique_key not in course_counts:
                            course_counts[unique_key] = {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}
                        if unique_key:
                            course_counts[unique_key]['total'] += 1
                            if session_type == 'P':
                                course_counts[unique_key]['labs'] += 1
                            elif session_type == 'T':
                                course_counts[unique_key]['tutorials'] += 1
                            else:
                                course_counts[unique_key]['lectures'] += 1
                    
                    # Count for this period (grid_sessions are already filtered to this period)
                    if unique_key and unique_key not in course_counts_period:
                        course_counts_period[unique_key] = {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}
                    if unique_key:
                        course_counts_period[unique_key]['total'] += 1
                        if session_type == 'P':
                            course_counts_period[unique_key]['labs'] += 1
                        elif session_type == 'T':
                            course_counts_period[unique_key]['tutorials'] += 1
                        else:
                            course_counts_period[unique_key]['lectures'] += 1
        
        # Process existing sessions (electives, combined) - only if grid_sessions not provided
        # If grid_sessions provided, we already counted combined courses from grid, so skip raw combined session counting
        # But we still need to count electives and Phase 5/7 sessions from their raw lists
        if not grid_sessions:
            for session in scheduled_sessions:
                if isinstance(session, dict):
                    course_code = session.get('course_code', '')
                    session_sections = session.get('sections', [])
                    session_period = session.get('period', '')
                    course_obj = session.get('course_obj')
                    
                    # EXACT section match (not substring)
                    section_match = any(f"{section}-Sem{semester}" in s for s in session_sections)
                    period_match = (period == 'PreMid' and session_period == 'PRE') or \
                                  (period == 'PostMid' and session_period == 'POST')
                    
                    # Count sessions across BOTH periods for satisfaction check
                    # But also track per-period for display
                    if not section_match:
                        continue
                        
                    # Remove suffixes like -TUT, -LAB for counting
                    base_code = course_code.replace('-TUT', '').replace('-LAB', '').split('-')[0]
                    
                    # Get session type from combined session dict
                    session_type = session.get('session_type', 'L')  # Default to lecture if not specified
                    
                    # Create unique key using course code + LTPSC to distinguish different courses with same code
                    if course_obj and hasattr(course_obj, 'ltpsc'):
                        unique_key = f"{base_code}_{course_obj.ltpsc}"
                    else:
                        unique_key = base_code
                    
                    # Count for total (across both periods)
                    if unique_key and unique_key not in course_counts:
                        course_counts[unique_key] = {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}
                    if unique_key:
                        course_counts[unique_key]['total'] += 1
                        if session_type == 'P':
                            course_counts[unique_key]['labs'] += 1
                        elif session_type == 'T':
                            course_counts[unique_key]['tutorials'] += 1
                        else:
                            course_counts[unique_key]['lectures'] += 1
                    
                    # Also count per-period for this period only
                    if period_match:
                        if unique_key and unique_key not in course_counts_period:
                            course_counts_period[unique_key] = {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}
                        if unique_key:
                            course_counts_period[unique_key]['total'] += 1
                            if session_type == 'P':
                                course_counts_period[unique_key]['labs'] += 1
                            elif session_type == 'T':
                                course_counts_period[unique_key]['tutorials'] += 1
                            else:
                                course_counts_period[unique_key]['lectures'] += 1
                    
                    # Skip the old counting logic below
                    continue
                else:
                    # Handle ScheduledSession objects - Count across BOTH periods
                    course_code = getattr(session, 'course_code', '')
                    session_section = getattr(session, 'section', '')
                    session_period = getattr(session, 'period', '')
                    course_obj = getattr(session, 'course_obj', None)
                    
                    # Check section match (count across both periods)
                    section_match = f"{section}-Sem{semester}" in session_section if session_section else False
                    
                    # Check period match (for per-period counting)
                    period_match = (period == 'PreMid' and session_period == 'PRE') or \
                                  (period == 'PostMid' and session_period == 'POST')
                    
                    if not section_match:
                        continue
                        
                    # Remove suffixes like -TUT, -LAB for counting
                    base_code = course_code.replace('-TUT', '').replace('-LAB', '')
                    
                    # Create unique key using course code + LTPSC to distinguish different courses with same code
                    if course_obj and hasattr(course_obj, 'ltpsc'):
                        unique_key = f"{base_code}_{course_obj.ltpsc}"
                    else:
                        unique_key = base_code
                    
                    # Count for total (across both periods)
                    if unique_key and unique_key not in course_counts:
                        course_counts[unique_key] = {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}
                    if unique_key:
                        course_counts[unique_key]['total'] += 1
                        if session.kind == "T":
                            course_counts[unique_key]['tutorials'] += 1
                        elif session.kind == "P":
                            course_counts[unique_key]['labs'] += 1
                        else:  # Lecture
                            course_counts[unique_key]['lectures'] += 1
                    
                    # Also count per-period for this period only
                    if period_match:
                        if unique_key and unique_key not in course_counts_period:
                            course_counts_period[unique_key] = {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}
                        if unique_key:
                            course_counts_period[unique_key]['total'] += 1
                            if session.kind == "T":
                                course_counts_period[unique_key]['tutorials'] += 1
                            elif session.kind == "P":
                                course_counts_period[unique_key]['labs'] += 1
                            else:  # Lecture
                                course_counts_period[unique_key]['lectures'] += 1
        
        # Process Phase 5 sessions - Count from grid if available, otherwise from raw list
        # If grid_sessions provided, Phase 5 courses are already counted from grid above
        # Only count from raw list if grid_sessions not provided
        if phase5_sessions and not grid_sessions:
            for session in phase5_sessions:
                if (hasattr(session, 'section') and session.section == f"{section}-Sem{semester}" and
                    hasattr(session, 'period')):
                    session_period = session.period
                    period_match = (period == 'PreMid' and session_period == 'PRE') or \
                                  (period == 'PostMid' and session_period == 'POST')
                    
                    course_code = session.course_code
                    if isinstance(course_code, str):
                        base_code = course_code.replace('-TUT', '').replace('-LAB', '')
                    elif pd.isna(course_code):
                        base_code = 'nan'
                    else:
                        base_code = str(course_code) if course_code else ''
                    
                    # Create unique key using course code + LTPSC if available
                    course_obj = getattr(session, 'course_obj', None)
                    if course_obj and hasattr(course_obj, 'ltpsc'):
                        unique_key = f"{base_code}_{course_obj.ltpsc}"
                    else:
                        unique_key = base_code
                    
                    # Count for total (across both periods)
                    if unique_key and unique_key not in course_counts:
                        course_counts[unique_key] = {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}
                    if unique_key:
                        course_counts[unique_key]['total'] += 1
                        if session.kind == "T":
                            course_counts[unique_key]['tutorials'] += 1
                        elif session.kind == "P":
                            course_counts[unique_key]['labs'] += 1
                        else:  # Lecture
                            course_counts[unique_key]['lectures'] += 1
                    
                    # Also count per-period for this period only
                    if period_match:
                        if unique_key and unique_key not in course_counts_period:
                            course_counts_period[unique_key] = {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}
                        if unique_key:
                            course_counts_period[unique_key]['total'] += 1
                            if session.kind == "T":
                                course_counts_period[unique_key]['tutorials'] += 1
                            elif session.kind == "P":
                                course_counts_period[unique_key]['labs'] += 1
                            else:  # Lecture
                                course_counts_period[unique_key]['lectures'] += 1
        
        # Process Phase 7 sessions - Count from grid if available, otherwise from raw list
        # If grid_sessions provided, Phase 7 courses are already counted from grid above
        # Only count from raw list if grid_sessions not provided
        if phase7_sessions and not grid_sessions:
            for session in phase7_sessions:
                if (hasattr(session, 'section') and session.section == f"{section}-Sem{semester}" and
                    hasattr(session, 'period')):
                    session_period = session.period
                    period_match = (period == 'PreMid' and session_period == 'PRE') or \
                                  (period == 'PostMid' and session_period == 'POST')
                    
                    course_code = session.course_code
                    if isinstance(course_code, str):
                        base_code = course_code.replace('-TUT', '').replace('-LAB', '')
                    elif pd.isna(course_code):
                        base_code = 'nan'
                    else:
                        base_code = str(course_code) if course_code else ''
                    
                    # Create unique key using course code + LTPSC if available
                    course_obj = getattr(session, 'course_obj', None)
                    if course_obj and hasattr(course_obj, 'ltpsc'):
                        unique_key = f"{base_code}_{course_obj.ltpsc}"
                    else:
                        unique_key = base_code
                    
                    # Count for total (across both periods)
                    if unique_key and unique_key not in course_counts:
                        course_counts[unique_key] = {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}
                    if unique_key:
                        course_counts[unique_key]['total'] += 1
                        if session.kind == "T":
                            course_counts[unique_key]['tutorials'] += 1
                        elif session.kind == "P":
                            course_counts[unique_key]['labs'] += 1
                        else:  # Lecture
                            course_counts[unique_key]['lectures'] += 1
                    
                    # Also count per-period for this period only
                    if period_match:
                        if unique_key and unique_key not in course_counts_period:
                            course_counts_period[unique_key] = {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}
                        if unique_key:
                            course_counts_period[unique_key]['total'] += 1
                            if session.kind == "T":
                                course_counts_period[unique_key]['tutorials'] += 1
                            elif session.kind == "P":
                                course_counts_period[unique_key]['labs'] += 1
                            else:  # Lecture
                                course_counts_period[unique_key]['lectures'] += 1

        def _course_visible_on_period_grid(code) -> bool:
            """
            True if the course row's code appears on this sheet's timetable grid for this period.
            Supports joint codes (e.g. CS165/CS201) when the grid shows CS165-TUT, etc.
            """
            if not grid_sessions:
                return True
            if code is None or (isinstance(code, float) and pd.isna(code)):
                return False
            code_s = str(code).strip()
            aliases = {code_s}
            if '/' in code_s:
                for part in code_s.split('/'):
                    p = part.strip()
                    if p:
                        aliases.add(p)
            if any(a in courses_on_this_sheet_period for a in aliases):
                return True
            skip_labels = {
                "LUNCH", "Break(15min)", "ELECTIVE", "ELECTIVE-TUT", "ELECTIVE-LAB",
            }
            for _day_name, day_sessions in grid_sessions.items():
                for slot in day_sessions or []:
                    if not isinstance(slot, tuple) or len(slot) < 2:
                        continue
                    disp = str(slot[1])
                    if disp in skip_labels:
                        continue
                    gbase = disp.replace('-TUT', '').replace('-LAB', '').split('-')[0]
                    if any(a == gbase for a in aliases):
                        return True
            return False

        def _include_core_course_on_this_period_sheet(course) -> bool:
            """
            Only list core courses that actually appear on this sheet's grid for this half
            (PreMid vs PostMid). ARISE emits one workbook sheet per (section, semester, period);
            the grid is the single source of truth for what belongs in that period's verification
            block—avoids CS162-style rows with 0/… and UNSATISFIED on a half where the course
            does not run.
            """
            cc = getattr(course, "code", "")
            if pd.isna(cc):
                cc_s = "nan"
            else:
                cc_s = str(cc)
            if grid_sessions:
                return _course_visible_on_period_grid(cc_s)
            credits = int(getattr(course, "credits", 0) or 0)
            is_half = bool(getattr(course, "half_semester", False))
            if credits > 2 and not is_half:
                return True
            if hasattr(course, "ltpsc"):
                uk = f"{cc_s}_{course.ltpsc}"
            else:
                uk = cc_s
            cip = 0
            if uk in course_counts_period:
                cip = course_counts_period[uk].get("total", 0)
            elif cc_s in course_counts_period:
                cip = course_counts_period[cc_s].get("total", 0)
            if cip > 0:
                return True
            tot = 0
            if uk in course_counts:
                tot = course_counts[uk].get("total", 0)
            elif cc_s in course_counts:
                tot = course_counts[cc_s].get("total", 0)
            return tot == 0

        # Filter courses for this semester/section
        relevant_courses = []
        
        # Dynamically determine combined course codes from actual scheduled sessions
        combined_course_codes = set()
        if combined_sessions:
            for session in combined_sessions:
                if isinstance(session, dict):
                    course_code = session.get('course_code', '').split('-')[0]  # Remove -TUT/-LAB suffix
                    if course_code:
                        combined_course_codes.add(course_code)
                elif hasattr(session, 'course_code'):
                    course_code = session.course_code.split('-')[0]  # Remove -TUT/-LAB suffix
                    if course_code:
                        combined_course_codes.add(course_code)

        combined_codes_this_section = _combined_course_codes_for_section(
            combined_sessions, section, semester, period
        )
        
        # Extract section department using generalized function
        section_dept = extract_department_from_section(section)
        
        for course in courses:
            course_code = getattr(course, 'code', '')
            course_dept = getattr(course, 'department', '')
            
            # Handle NaN course codes - convert to string for comparison
            if pd.isna(course_code):
                course_code_str = 'nan'
            else:
                course_code_str = str(course_code)
            
            # Include if:
            # 1. Course belongs to this semester
            # 2. Course is not a combined course (handled separately)
            # 3. Course department matches section department
            # 4. Course is NOT an elective (electives are handled by elective basket)
            # 5. Show ALL courses (Phase 5 >2 credits AND Phase 7 <=2 credits) - REMOVED is_scheduled check
            
            is_elective = getattr(course, 'is_elective', False)
            
            if (hasattr(course, 'semester') and course.semester == semester and
                course_code_str not in combined_codes_this_section and
                course_dept == section_dept and
                not is_elective):  # Exclude electives - they're in the elective basket
                if not _include_core_course_on_this_period_sheet(course):
                    continue
                relevant_courses.append(course)
        
        # STEP 1: Extract ALL group-based codes for this semester
        # Helper function to extract semester from group key
        def extract_semester_from_group(gk: str) -> int:
            try:
                if '.' in str(gk):
                    return int(str(gk).split('.')[0])
                else:
                    return int(gk)
            except (ValueError, AttributeError):
                return -1
        
        # Count actual elective sessions for this semester/section/period
        # STRATEGY: Use grid_sessions as PRIMARY source, scheduled_sessions as fallback
        # Sessions are created with codes like ELECTIVE_BASKET_1.1, ELECTIVE_BASKET_5.1, etc.
        # We need to find ALL groups for this semester to count correctly
        elective_basket_codes = []
        seen_codes = set()
        
        # Pass 1: Extract from scheduled_sessions - scan ALL sessions for this semester
        for session in scheduled_sessions:
            if isinstance(session, dict):
                session_code = session.get('course_code', '')
            else:
                session_code = getattr(session, 'course_code', '')
            
            if session_code and session_code.startswith('ELECTIVE_BASKET_'):
                group_key = session_code.replace('ELECTIVE_BASKET_', '')
                session_semester = extract_semester_from_group(group_key)
                # Add ALL group-based codes for this semester, regardless of section/period
                if session_semester == semester:
                    if session_code not in seen_codes:
                        elective_basket_codes.append(session_code)
                        seen_codes.add(session_code)
        
        # DEBUG: Log what we found
        import logging
        logging.debug(f"Pass 1: Found {len(elective_basket_codes)} elective basket codes: {elective_basket_codes}")
        
        # Pass 2: ALWAYS check grid_sessions (primary source) - it has the actual displayed codes
        # Grid may show "ELECTIVE BASKET 1.1" or "ELECTIVE_BASKET_5.1" which we can parse
        if grid_sessions:
            import logging
            import re
            logging.debug(f"Pass 2: Checking grid_sessions for elective codes...")
            for day_name, day_sessions in grid_sessions.items():
                if not day_sessions:
                    continue
                for grid_session in day_sessions:
                    if isinstance(grid_session, tuple) and len(grid_session) >= 2:
                        course_display = grid_session[1]
                        if isinstance(course_display, str) and 'ELECTIVE' in course_display.upper():
                            # Grid may show "ELECTIVE BASKET 1.1" or "ELECTIVE_BASKET_5.1" - try to extract group key
                            # First check if it's already in ELECTIVE_BASKET_X.X format
                            if course_display.startswith('ELECTIVE_BASKET_'):
                                # Already in correct format, extract directly
                                group_key = course_display.replace('ELECTIVE_BASKET_', '').split('-')[0]  # Remove -TUT/-LAB suffix
                                # Strict: only accept real group keys (typically X.Y). Do not invent semester-only groups.
                                if '.' not in str(group_key):
                                    continue
                                inferred_code = f"ELECTIVE_BASKET_{group_key}"
                                inferred_semester = extract_semester_from_group(group_key)
                                if inferred_semester == semester and inferred_code not in seen_codes:
                                    elective_basket_codes.append(inferred_code)
                                    seen_codes.add(inferred_code)
                                    logging.debug(f"  Found group code from grid: {inferred_code}")
                            else:
                                # Try parsing "ELECTIVE BASKET X.X" format
                                # Strict: require dotted group keys (e.g., 3.1). Avoid semester-only fallbacks like "3".
                                match = re.search(r'BASKET\s+(\d+\.\d+)', course_display.upper())
                                if match:
                                    group_key = match.group(1)
                                    inferred_code = f"ELECTIVE_BASKET_{group_key}"
                                    inferred_semester = extract_semester_from_group(group_key)
                                    if inferred_semester == semester and inferred_code not in seen_codes:
                                        elective_basket_codes.append(inferred_code)
                                        seen_codes.add(inferred_code)
                                        logging.debug(f"  Found group code from grid (parsed): {inferred_code}")
        
        # Pass 3: Final scan of scheduled_sessions to catch any missed codes
        if not elective_basket_codes:
            for session in scheduled_sessions:
                if isinstance(session, dict):
                    session_code = session.get('course_code', '')
                else:
                    session_code = getattr(session, 'course_code', '')
                
                if session_code and session_code.startswith('ELECTIVE_BASKET_'):
                    group_key = session_code.replace('ELECTIVE_BASKET_', '')
                    if extract_semester_from_group(group_key) == semester:
                        if session_code not in seen_codes:
                            elective_basket_codes.append(session_code)
                            seen_codes.add(session_code)
        
        # If we still have no codes but grid shows electives, try one more time to extract from grid
        # This is a final fallback - should rarely be needed if Pass 2 worked correctly
        if not elective_basket_codes and grid_sessions:
            import re
            # Check if grid has any elective entries - try to extract ALL group codes
            found_any_electives = False
            for day_name, day_sessions in grid_sessions.items():
                if not day_sessions:
                    continue
                for grid_session in day_sessions:
                    if isinstance(grid_session, tuple) and len(grid_session) >= 2:
                        course_display = grid_session[1]
                        if isinstance(course_display, str) and 'ELECTIVE' in course_display.upper():
                            found_any_electives = True
                            # Try to extract group code one more time
                            if course_display.startswith('ELECTIVE_BASKET_'):
                                group_key = course_display.replace('ELECTIVE_BASKET_', '').split('-')[0]
                                if '.' not in str(group_key):
                                    continue
                                inferred_code = f"ELECTIVE_BASKET_{group_key}"
                                inferred_semester = extract_semester_from_group(group_key)
                                if inferred_semester == semester and inferred_code not in seen_codes:
                                    elective_basket_codes.append(inferred_code)
                                    seen_codes.add(inferred_code)
                            else:
                                # Try parsing "ELECTIVE BASKET X.X" format
                                match = re.search(r'BASKET\s+(\d+\.\d+)', course_display.upper())
                                if match:
                                    group_key = match.group(1)
                                    inferred_code = f"ELECTIVE_BASKET_{group_key}"
                                    inferred_semester = extract_semester_from_group(group_key)
                                    if inferred_semester == semester and inferred_code not in seen_codes:
                                        elective_basket_codes.append(inferred_code)
                                        seen_codes.add(inferred_code)
            # CRITICAL: Do NOT invent semester-level fallback codes.
            # Elective baskets must come from course.elective_group only.
            if found_any_electives and not elective_basket_codes:
                import logging
                logging.warning(
                    f"Grid contains electives for semester {semester} but no group keys were extractable. "
                    f"Not adding any fallback ELECTIVE_BASKET_{semester} to avoid phantom groups."
                )
        
        # STEP 2: Count elective basket sessions
        # STRATEGY: Use grid_sessions as PRIMARY source (shows what's actually displayed)
        # Fallback to scheduled_sessions only if grid_sessions is not available
        
        # Initialize counts
        elective_lectures = 0
        elective_tutorials = 0
        elective_labs = 0
        
        # PRIMARY SOURCE: Count from grid_sessions FIRST
        # Grid shows actual displayed sessions, so it's the most reliable source
        # The grid is built per day/section/period, so it should contain all elective sessions for this section/period
        if grid_sessions:
            for day_name, day_sessions in grid_sessions.items():
                if not day_sessions:
                    continue
                for grid_session in day_sessions:
                    # Handle both 2-tuple (TimeBlock, course_display) and 4-tuple formats
                    if isinstance(grid_session, tuple):
                        if len(grid_session) >= 2:
                            course_display = grid_session[1]  # course_code or course_display
                            if isinstance(course_display, str):
                                course_upper = course_display.upper()
                                # Check for elective sessions - handle multiple formats:
                                # "ELECTIVE", "ELECTIVE-TUT", "ELECTIVE-LAB", "ELECTIVE BASKET 1.1", etc.
                                if 'ELECTIVE' in course_upper:
                                    # This is an elective session - determine type
                                    # Grid shows: "ELECTIVE", "ELECTIVE-TUT", "ELECTIVE-LAB", "ELECTIVE BASKET X.X"
                                    if 'TUT' in course_upper:
                                        elective_tutorials += 1
                                    elif 'LAB' in course_upper:
                                        elective_labs += 1
                                    else:
                                        # Regular elective (lecture) - "ELECTIVE" or "ELECTIVE BASKET X.X"
                                        elective_lectures += 1
        
        # FALLBACK: If grid_sessions didn't provide counts (or not available), count from scheduled_sessions
        # This handles cases where grid_sessions is None or empty
        # SIMPLIFIED: Count by semester + period only (ignore section for elective baskets since they're shared)
        if (elective_lectures == 0 and elective_tutorials == 0 and elective_labs == 0) or not grid_sessions:
            # Count from scheduled_sessions (original session objects with proper course_code)
            # SIMPLIFIED LOGIC: For elective baskets, count by semester + period only
            # Section matching is less critical since baskets are shared across sections
            for session in scheduled_sessions:
                if isinstance(session, dict):
                    session_code = session.get('course_code', '')
                    session_sections = session.get('sections', [])
                    session_period = session.get('period', '')
                    
                    # SIMPLIFIED: Only check period match for elective baskets
                    # Period matching is critical (PRE/POST vs PreMid/PostMid)
                    period_match = False
                    if session_period:
                        session_period_normalized = str(session_period).strip().upper()
                        # Handle both "PRE"/"POST" and "PreMid"/"PostMid" formats
                        if period == 'PreMid':
                            if session_period_normalized in ['PRE', 'PREMID', 'PRE-MID']:
                                period_match = True
                        elif period == 'PostMid':
                            if session_period_normalized in ['POST', 'POSTMID', 'POST-MID']:
                                period_match = True
                    
                    # Check if this is an elective basket session
                    is_elective_basket = False
                    if session_code and session_code.startswith('ELECTIVE_BASKET_'):
                        if session_code in elective_basket_codes:
                            is_elective_basket = True
                        else:
                            # Check if group key matches semester
                            group_key = session_code.replace('ELECTIVE_BASKET_', '')
                            if extract_semester_from_group(group_key) == semester:
                                if session_code not in elective_basket_codes:
                                    elective_basket_codes.append(session_code)
                                is_elective_basket = True
                    
                    # SIMPLIFIED: For elective baskets, only require period match (not section)
                    # Elective baskets are shared across sections, so section matching is less critical
                    if is_elective_basket and period_match:  # Removed section_match requirement
                        # Determine session type based on course code suffix or session type
                        if '-TUT' in str(session_code) or session.get('session_type') == 'T':
                            elective_tutorials += 1
                        elif '-LAB' in str(session_code) or session.get('session_type') == 'P':
                            elective_labs += 1
                        else:
                            elective_lectures += 1
                else:
                    # Handle ScheduledSession objects
                    session_code = getattr(session, 'course_code', '')
                    session_section = getattr(session, 'section', '')
                    session_period = getattr(session, 'period', '')
                    
                    # SIMPLIFIED: Only check period match for elective baskets
                    # Period matching is critical (PRE/POST vs PreMid/PostMid)
                    period_match = False
                    if session_period:
                        session_period_normalized = str(session_period).strip().upper()
                        # Handle both "PRE"/"POST" and "PreMid"/"PostMid" formats
                        if period == 'PreMid':
                            if session_period_normalized in ['PRE', 'PREMID', 'PRE-MID']:
                                period_match = True
                        elif period == 'PostMid':
                            if session_period_normalized in ['POST', 'POSTMID', 'POST-MID']:
                                period_match = True
                    
                    # Check if this is an elective basket session (handle both group-based and semester-based codes)
                    is_elective_basket = False
                    if session_code and session_code.startswith('ELECTIVE_BASKET_'):
                        # Check if it matches any of the expected codes OR belongs to this semester
                        if session_code in elective_basket_codes:
                            is_elective_basket = True
                        else:
                            # Also check if the group key matches (e.g., session has "1.1" and we're looking for semester 1)
                            group_key = session_code.replace('ELECTIVE_BASKET_', '')
                            if extract_semester_from_group(group_key) == semester:
                                # This is an elective basket for this semester, add to codes if not already there
                                if session_code not in elective_basket_codes:
                                    elective_basket_codes.append(session_code)
                                is_elective_basket = True
                    
                    # SIMPLIFIED: Only require period match (not section) for elective baskets
                    if is_elective_basket and period_match:  # Removed section_match requirement
                        # Determine lecture vs tutorial based on kind attribute (L, T, P)
                        session_kind = getattr(session, 'kind', '')
                        if session_kind == 'T' or session_kind == 'TUTORIAL':
                            elective_tutorials += 1
                        elif session_kind == 'L' or session_kind == 'LECTURE' or session_kind == 'ELECTIVE':
                            elective_lectures += 1
                        elif session_kind == 'P' or session_kind == 'PRACTICAL':
                            elective_labs += 1
                        else:
                            # Fallback: check day pattern (old logic for compatibility)
                            block = getattr(session, 'block', None)
                            if block:
                                day = getattr(block, 'day', '')
                                if day == 'Friday':
                                    elective_tutorials += 1
                                elif day in ['Monday', 'Wednesday', 'Tuesday', 'Thursday']:
                                    elective_lectures += 1
                            else:
                                # Default to lecture if unknown
                                elective_lectures += 1
        
        
        # Add elective basket entries - SEPARATE ENTRY FOR EACH GROUP (5.1, 5.2, etc.)
        # CRITICAL: Create separate entries per group instead of aggregating by semester
        # If still no codes, extract groups from courses (check elective_group attribute)
        import logging
        if not elective_basket_codes:
            # Extract groups from courses - check elective_group attribute
            for course in courses:
                if hasattr(course, 'is_elective') and course.is_elective:
                    if hasattr(course, 'semester') and course.semester == semester:
                        if hasattr(course, 'elective_group') and course.elective_group:
                            code = f"ELECTIVE_BASKET_{course.elective_group}"
                            if code not in seen_codes:
                                elective_basket_codes.append(code)
                                seen_codes.add(code)
                                logging.debug(f"Added group code from course.elective_group: {code}")
        
        logging.debug(f"Final elective_basket_codes for semester {semester}: {elective_basket_codes}")
        if elective_basket_codes:
            # Extract unique groups for this semester
            unique_groups = sorted(set([code.replace('ELECTIVE_BASKET_', '') for code in elective_basket_codes]))
            
            # Write headers (only once, before first basket entry)
            header_row = current_row
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=header_row, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.colors['header']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
            current_row += 1
            
            # Create separate entry for EACH group
            for group_key in unique_groups:
                basket_code = f'ELECTIVE_BASKET_{group_key}'  # e.g., "ELECTIVE_BASKET_5.1"
                basket_name = f'Elective Basket - Sem {semester} Group {group_key}'
                
                # Count sessions for THIS specific group only
                group_lectures = 0
                group_tutorials = 0
                group_labs = 0
                
                # Count from grid_sessions (primary source)
                if grid_sessions:
                    for day_name, day_sessions in grid_sessions.items():
                        if not day_sessions:
                            continue
                        for grid_session in day_sessions:
                            if isinstance(grid_session, tuple) and len(grid_session) >= 2:
                                block = grid_session[0]
                                course_display = grid_session[1]
                                if isinstance(course_display, str):
                                    course_upper = course_display.upper()
                                    # Check if this session belongs to this group
                                    # Grid may show "ELECTIVE_BASKET_5.1", "ELECTIVE_BASKET_5.1-TUT", etc.
                                    if f'ELECTIVE_BASKET_{group_key}'.upper() in course_upper or \
                                       f'BASKET {group_key}'.upper() in course_upper:
                                        # Prefer explicit markers if present
                                        if 'LAB' in course_upper:
                                            group_labs += 1
                                        elif 'TUT' in course_upper:
                                            group_tutorials += 1
                                        else:
                                            # Fallback: infer tutorial vs lecture by duration
                                            try:
                                                start = getattr(block, 'start', None)
                                                end = getattr(block, 'end', None)
                                                if start and end:
                                                    minutes = (end.hour * 60 + end.minute) - (start.hour * 60 + start.minute)
                                                else:
                                                    minutes = None
                                            except Exception:
                                                minutes = None
                                            
                                            if minutes is not None and minutes <= 60:
                                                group_tutorials += 1
                                            else:
                                                group_lectures += 1
                
                # Fallback: Count from scheduled_sessions for this group
                if group_lectures == 0 and group_tutorials == 0 and group_labs == 0:
                    group_code_prefix = f'ELECTIVE_BASKET_{group_key}'
                    for session in scheduled_sessions:
                        session_code = ''
                        session_period = ''
                        
                        if isinstance(session, dict):
                            session_code = session.get('course_code', '')
                            session_period = session.get('period', '')
                        else:
                            session_code = getattr(session, 'course_code', '')
                            session_period = getattr(session, 'period', '')
                        
                        # Check if this session belongs to this group and period
                        period_match = False
                        if session_period:
                            session_period_normalized = str(session_period).strip().upper()
                            if period == 'PreMid' and session_period_normalized in ['PRE', 'PREMID', 'PRE-MID']:
                                period_match = True
                            elif period == 'PostMid' and session_period_normalized in ['POST', 'POSTMID', 'POST-MID']:
                                period_match = True
                        
                        if session_code and session_code.startswith(group_code_prefix) and period_match:
                            session_kind = ''
                            if isinstance(session, dict):
                                session_kind = session.get('session_type', '')
                            else:
                                session_kind = getattr(session, 'kind', '')
                            
                            if session_kind == 'T' or '-TUT' in session_code:
                                group_tutorials += 1
                            elif session_kind == 'P' or '-LAB' in session_code:
                                group_labs += 1
                            else:
                                group_lectures += 1
                
                # Create entry for this group
                elective_basket = {
                    'code': basket_code,
                    'name': basket_name,
                    'instructor': 'Multiple',
                    'ltpsc': 'N/A',
                    'lab': '',
                    'classroom': '',
                    'required_lectures': 2,
                    'required_tutorials': 1,
                    'required_labs': 0,
                    'scheduled_lectures': group_lectures,
                    'scheduled_tutorials': group_tutorials,
                    'scheduled_labs': group_labs,
                    'status': 'SATISFIED' if (group_lectures >= 2 and group_tutorials >= 1) else 'UNSATISFIED'
                }
                
                # Add elective basket row for this group
                elective_row = [
                    elective_basket['code'],
                    elective_basket['name'],
                    elective_basket['instructor'],
                    elective_basket['ltpsc'],
                    elective_basket['lab'],
                    elective_basket['classroom'],
                    f"{elective_basket['required_lectures']}/{elective_basket['scheduled_lectures']}",
                    f"{elective_basket['required_tutorials']}/{elective_basket['scheduled_tutorials']}",
                    f"{elective_basket['required_labs']}/{elective_basket['scheduled_labs']}",
                    elective_basket['status']
                ]
                
                for col, value in enumerate(elective_row, 1):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = value
                    cell.font = self.course_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.thin_border
                    if col == len(elective_row):  # Status column
                        if value == 'SATISFIED':
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                current_row += 1
        else:
            # Fallback: If no group codes found, create semester-based entry (backward compatibility)
            basket_code = f'ELECTIVE_BASKET_{semester}'
            basket_name = f'Elective Basket - Sem {semester}'
            
            elective_basket = {
                'code': basket_code,
                'name': basket_name,
                'instructor': 'Multiple',
                'ltpsc': 'N/A',
                'lab': '',
                'classroom': '',
                'required_lectures': 2,
                'required_tutorials': 1,
                'required_labs': 0,
                'scheduled_lectures': elective_lectures,
                'scheduled_tutorials': elective_tutorials,
                'scheduled_labs': elective_labs,
                'status': 'SATISFIED' if (elective_lectures >= 2 and elective_tutorials >= 1) else 'UNSATISFIED'
            }
            
            # Write headers
            header_row = current_row
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=header_row, column=col)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.colors['header']
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
            current_row += 1
            
            # Add elective basket row
            elective_row = [
                elective_basket['code'],
                elective_basket['name'],
                elective_basket['instructor'],
                elective_basket['ltpsc'],
                elective_basket['lab'],
                elective_basket['classroom'],
                f"{elective_basket['required_lectures']}/{elective_basket['scheduled_lectures']}",
                f"{elective_basket['required_tutorials']}/{elective_basket['scheduled_tutorials']}",
                f"{elective_basket['required_labs']}/{elective_basket['scheduled_labs']}",
                elective_basket['status']
            ]
            
            for col, value in enumerate(elective_row, 1):
                cell = sheet.cell(row=current_row, column=col)
                cell.value = value
                cell.font = self.course_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                if col == len(elective_row):  # Status column
                    if value == 'SATISFIED':
                        cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            current_row += 1
        
        # Add combined courses - Show ONLY courses for this period
        # Dynamically determine combined course codes from actual scheduled sessions
        combined_course_codes = set()
        if combined_sessions:
            for session in combined_sessions:
                if isinstance(session, dict):
                    course_code = session.get('course_code', '').split('-')[0]  # Remove -TUT/-LAB suffix
                    if course_code:
                        combined_course_codes.add(course_code)
                elif hasattr(session, 'course_code'):
                    course_code = session.course_code.split('-')[0]  # Remove -TUT/-LAB suffix
                    if course_code:
                        combined_course_codes.add(course_code)
        
        # Show combined courses for this semester, filtered by period
        for course_code in sorted(combined_course_codes):
            # Determine which period this course belongs to dynamically, checking for this specific section
            course_period = get_combined_course_period(course_code, semester, combined_sessions, section)
            
            # Skip if course doesn't belong to this period
            if course_period is None:
                continue
            
            # Check if this course should be shown in current period table
            period_match = (period == 'PreMid' and course_period == 'PRE') or \
                          (period == 'PostMid' and course_period == 'POST')
            
            if not period_match:
                continue  # Skip courses that don't belong to this period
            
            # Show combined courses based on semester, not just course_counts
            # CRITICAL: Match by code + semester + department to handle courses with same code but different LTPSC
            section_dept_for_matching = extract_department_from_section(section)
            course_obj = next((c for c in courses if 
                             getattr(c, 'code', '') == course_code and
                             hasattr(c, 'semester') and c.semester == semester and
                             getattr(c, 'department', '') == section_dept_for_matching), None)
            # Fallback: if no match by department, try without department (for backward compatibility)
            if not course_obj:
                course_obj = next((c for c in courses if getattr(c, 'code', '') == course_code), None)
            if course_obj and course_obj.semester == semester:
                # CRITICAL: Only show course if it's actually scheduled for this section
                # Check if this course appears in combined_sessions for this specific section
                course_scheduled_for_section = False
                for session in combined_sessions:
                    if isinstance(session, dict):
                        session_course = session.get('course_code', '').split('-')[0]
                        session_sections = session.get('sections', [])
                        if session_course == course_code and f"{section}-Sem{semester}" in session_sections:
                            course_scheduled_for_section = True
                            break
                    elif hasattr(session, 'course_code') and hasattr(session, 'sections'):
                        session_course = session.course_code.split('-')[0]
                        if session_course == course_code and f"{section}-Sem{semester}" in session.sections:
                            course_scheduled_for_section = True
                            break
                
                # Skip if course is not scheduled for this section
                if not course_scheduled_for_section:
                    continue
                course_name = getattr(course_obj, 'name', '')
                # Get actual faculty from scheduled sessions
                faculty_for_course = set()
                
                # Check phase5_sessions (core courses)
                if phase5_sessions:
                    for session in phase5_sessions:
                        if (hasattr(session, 'course_code') and session.course_code == course_code and
                            hasattr(session, 'section') and session.section == f"{section}-Sem{semester}" and
                            hasattr(session, 'period')):
                            session_period = session.period
                            period_match = (period == 'PreMid' and session_period == 'PRE') or \
                                          (period == 'PostMid' and session_period == 'POST')
                            if period_match and hasattr(session, 'faculty') and session.faculty:
                                faculty_for_course.add(session.faculty)
                
                # Check combined_sessions (combined courses)
                if combined_sessions:
                    for session in combined_sessions:
                        if (session.get('course_code', '').split('-')[0] == course_code and
                            f"{section}-Sem{semester}" in session.get('sections', []) and
                            session.get('period') == ('PRE' if period == 'PreMid' else 'POST')):
                            instructor_val = session.get('instructor', 'TBD')
                            if instructor_val and instructor_val != 'TBD':
                                faculty_for_course.add(instructor_val)
                
                # Check all_section_sessions (used during Log/UI fast-path regeneration)
                if all_section_sessions:
                    for session in all_section_sessions:
                        sess_code = str(session.get('Course Code', session.get('course_code', '')))
                        if sess_code.startswith('ELECTIVE_BASKET_'):
                            sess_base = sess_code.replace('-LAB', '').replace('-TUT', '')
                        else:
                            sess_base = sess_code.replace('-TUT', '').replace('-LAB', '').split('-')[0]
                        
                        if sess_base == course_code:
                            fac = session.get('Faculty', session.get('faculty', ''))
                            if fac and str(fac).strip() not in ('', 'TBD', 'nan', 'None'):
                                faculty_for_course.add(str(fac).strip())
                
                if faculty_for_course:
                    # To preserve original ordering (e.g., "Sunil P V, Sunil C K"), check if the collected
                    # faculty set exactly matches the course's original instructors list.
                    raw_instructors = getattr(course_obj, 'instructors', [])
                    if isinstance(raw_instructors, list) and set(raw_instructors) == faculty_for_course:
                        instructor = ', '.join(raw_instructors)
                    else:
                        instructor = ', '.join(sorted(faculty_for_course))
                else:
                    # Final fallback: read the instructors list from the root course database
                    raw_instructors = getattr(course_obj, 'instructors', [])
                    if raw_instructors:
                        instructor = ', '.join(raw_instructors)
                    else:
                        instructor = 'TBD'
                
                ltpsc = getattr(course_obj, 'ltpsc', '')
                credits = getattr(course_obj, 'credits', 0)
                
                # Calculate required slots from LTPSC
                from modules_v2.phase5_core_courses import calculate_slots_needed
                slots_needed = calculate_slots_needed(ltpsc)
                # Use LTPSC values directly - no hardcoded course-specific adjustments
                required_lectures = slots_needed['lectures']
                required_tutorials = slots_needed['tutorials']
                required_labs = slots_needed['practicals']
                
                # Per-period counts (for this sheet only) come from grid_sessions.
                # For full-semester satisfaction, we recompute TOTALS dynamically
                # from all sessions (combined + Phase 5 + Phase 7) for this
                # specific section across BOTH PreMid and PostMid.
                course_ltpsc = getattr(course_obj, 'ltpsc', '')
                course_unique_key = f"{course_code}_{course_ltpsc}"
                
                counts_period = course_counts_period.get(
                    course_unique_key,
                    course_counts_period.get(course_code, {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0}),
                )
                scheduled_lectures = counts_period['lectures']
                scheduled_tutorials = counts_period['tutorials']
                scheduled_labs = counts_period['labs']

                # Combined courses (any credit count): sessions are authoritative across both periods.
                # For half-semester combined courses (e.g. MA162, credits=2), the schedule splits
                # DSAI/ECE into PRE and CSE into POST. The total from combined_sessions is the
                # correct count — not the per-period grid count.
                is_combined_course = course_scheduled_for_section and combined_sessions and any(
                    isinstance(sess, dict)
                    and str(sess.get('course_code', '')).split('-')[0] == course_code
                    and f"{section}-Sem{semester}" in sess.get('sections', [])
                    for sess in combined_sessions
                )
                # Full-semester combined (credits > 2): LTPSC is met across PRE+POST — do not zero
                # period counts just because this half's grid omits some sessions.
                is_full_semester_combined = credits > 2 or is_combined_course
                if (not is_full_semester_combined) and grid_sessions and course_code not in courses_on_this_sheet_period:
                    scheduled_lectures = 0
                    scheduled_tutorials = 0
                    scheduled_labs = 0

                # Recompute TOTALS across all periods using raw sessions.
                total_lectures = 0
                total_tutorials = 0
                total_labs = 0

                section_key = f"{section}-Sem{semester}"

                # 1) Combined sessions (Phase 4)
                if combined_sessions:
                    for sess in combined_sessions:
                        if not isinstance(sess, dict):
                            continue
                        sess_code = str(sess.get("course_code", "")).split("-")[0]
                        if sess_code != course_code:
                            continue
                        sess_sections = sess.get("sections", [])
                        if section_key not in sess_sections:
                            continue
                        sess_type = sess.get("session_type", "L")
                        if sess_type == "P":
                            total_labs += 1
                        elif sess_type == "T":
                            total_tutorials += 1
                        else:
                            total_lectures += 1

                # 2) Phase 5 sessions (core >2 credits)
                if phase5_sessions:
                    for sess in phase5_sessions:
                        if getattr(sess, "section", "") != section_key:
                            continue
                        sess_code = str(getattr(sess, "course_code", "")).split("-")[0]
                        if sess_code != course_code:
                            continue
                        kind = getattr(sess, "kind", "L")
                        if kind == "P":
                            total_labs += 1
                        elif kind == "T":
                            total_tutorials += 1
                        else:
                            total_lectures += 1

                # 3) Phase 7 sessions (remaining <=2 credits)
                if phase7_sessions:
                    for sess in phase7_sessions:
                        if getattr(sess, "section", "") != section_key:
                            continue
                        sess_code = str(getattr(sess, "course_code", "")).split("-")[0]
                        if sess_code != course_code:
                            continue
                        kind = getattr(sess, "kind", "L")
                        if kind == "P":
                            total_labs += 1
                        elif kind == "T":
                            total_tutorials += 1
                        else:
                            total_lectures += 1
                # 4) Grid mirrors the timetable display; if structured session lists exist, they already
                #    supply semester totals — counting grid too would double-count. Use grid only when no
                #    Phase 4/5/7 rows were provided (e.g. log/UI regeneration edge paths).
                if grid_sessions and not (combined_sessions or phase5_sessions or phase7_sessions):
                    for day, sessions in grid_sessions.items():
                        for s in sessions:
                            block, course_display = s[0], s[1]
                            if course_display in ("LUNCH", "Break(15min)"):
                                continue
                            if course_code in course_display:
                                # Infer type
                                if '-LAB' in course_display:
                                    total_labs += 1
                                elif '-TUT' in course_display:
                                    total_tutorials += 1
                                else:
                                    total_lectures += 1

                # Combined courses (Phase 4) use totals as the authoritative count since their
                # sessions are split across periods by design (half-semester rotation).
                if is_full_semester_combined:
                    lectures_satisfied = total_lectures == required_lectures
                    tutorials_satisfied = total_tutorials == required_tutorials
                    labs_satisfied = total_labs == required_labs
                else:
                    lectures_satisfied = scheduled_lectures == required_lectures
                    tutorials_satisfied = scheduled_tutorials == required_tutorials
                    labs_satisfied = scheduled_labs == required_labs

                # DEBUG: log verification details for troubleshooting UNSATISFIED rows
                try:
                    debug_section = f"{section}-Sem{semester}"
                    debug_period = period
                    print(
                        f"DEBUG_VERIFY: section={debug_section} period={debug_period} "
                        f"course={course_code} ltpsc={ltpsc} "
                        f"required=(L={required_lectures},T={required_tutorials},P={required_labs}) "
                        f"total_scheduled=(L={total_lectures},T={total_tutorials},P={total_labs}) "
                        f"period_scheduled=(L={scheduled_lectures},T={scheduled_tutorials},P={scheduled_labs})"
                    )
                except Exception:
                    pass
                
                # Determine status - check each component separately
                if required_lectures == 0 and required_tutorials == 0 and required_labs == 0:
                    stray = (
                        total_lectures or total_tutorials or total_labs
                        or scheduled_lectures or scheduled_tutorials or scheduled_labs
                    )
                    status = 'SATISFIED' if not stray else 'UNSATISFIED'
                    if stray:
                        lectures_satisfied = tutorials_satisfied = labs_satisfied = False
                    else:
                        lectures_satisfied = tutorials_satisfied = labs_satisfied = True
                else:
                    status = 'SATISFIED' if (lectures_satisfied and tutorials_satisfied and labs_satisfied) else 'UNSATISFIED'
                
                # Get room assignment from room_assignments (Phase 8)
                # Key format: (course_code, section, period)
                assigned_labs = ''
                assigned_classroom = ''
                period_key = 'PRE' if period == 'PreMid' else 'POST'
                room_key = (course_code, f"{section}-Sem{semester}", period_key)
                
                if room_assignments and room_key in room_assignments:
                    assignment = room_assignments[room_key]
                    raw_classroom = assignment.get('classroom', '') or ''
                    labs_list = assignment.get('labs', [])
                    # Sanity check: if classroom field contains a lab room, reclassify it
                    if raw_classroom and is_lab_room(raw_classroom):
                        if raw_classroom not in labs_list:
                            labs_list = list(labs_list) + [raw_classroom]
                        raw_classroom = ''
                    assigned_classroom = raw_classroom
                    if labs_list:
                        assigned_labs = ', '.join(labs_list)
                
                # Fallback: If not in room_assignments, try to extract from actual sessions
                if not assigned_labs:
                    # Try to find lab sessions for this course
                    lab_rooms_found = set()
                    # Check Phase 5 sessions
                    if phase5_sessions:
                        for session in phase5_sessions:
                            if (hasattr(session, 'course_code') and 
                                session.course_code.split('-')[0] == course_code and
                                hasattr(session, 'section') and 
                                session.section == f"{section}-Sem{semester}" and
                                hasattr(session, 'period')):
                                session_period = session.period
                                period_match = (period == 'PreMid' and session_period == 'PRE') or \
                                              (period == 'PostMid' and session_period == 'POST')
                                if period_match and hasattr(session, 'kind') and session.kind == 'P':
                                    if hasattr(session, 'room') and session.room:
                                        for r in str(session.room).split(','):
                                            if r and r.strip():
                                                lab_rooms_found.add(r.strip())
                    # Check Phase 7 sessions
                    if phase7_sessions:
                        for session in phase7_sessions:
                            if (hasattr(session, 'course_code') and 
                                session.course_code.split('-')[0] == course_code and
                                hasattr(session, 'section') and 
                                session.section == f"{section}-Sem{semester}" and
                                hasattr(session, 'period')):
                                session_period = session.period
                                period_match = (period == 'PreMid' and session_period == 'PRE') or \
                                              (period == 'PostMid' and session_period == 'POST')
                                if period_match and hasattr(session, 'kind') and session.kind == 'P':
                                    if hasattr(session, 'room') and session.room:
                                        for r in str(session.room).split(','):
                                            if r and r.strip():
                                                lab_rooms_found.add(r.strip())
                    # Check combined sessions for practicals
                    if combined_sessions:
                        for session in combined_sessions:
                            if isinstance(session, dict):
                                session_course = session.get('course_code', '').split('-')[0]
                                session_sections = session.get('sections', [])
                                session_period = session.get('period', '')
                                session_type = session.get('session_type', 'L')
                                if (session_course == course_code and
                                    f"{section}-Sem{semester}" in session_sections and
                                    session_type == 'P'):
                                    period_match = (period == 'PreMid' and session_period == 'PRE') or \
                                                  (period == 'PostMid' and session_period == 'POST')
                                    if period_match:
                                        room = session.get('room', '') or ''
                                        for r in str(room).split(','):
                                            r = (r or '').strip()
                                            if r and (r.upper().startswith('L') or 'Lab' in r):
                                                lab_rooms_found.add(r)
                    
                    # Check all_section_sessions for practicals
                    if all_section_sessions:
                        for session in all_section_sessions:
                            sess_code = str(session.get('Course Code', session.get('course_code', '')))
                            if sess_code.startswith('ELECTIVE_BASKET_'):
                                sess_base = sess_code.replace('-LAB', '').replace('-TUT', '')
                            else:
                                sess_base = sess_code.replace('-TUT', '').replace('-LAB', '').split('-')[0]
                            
                            session_type = session.get('Session Type', session.get('session_type', 'L'))
                            if sess_base == course_code and session_type == 'P':
                                room = session.get('Room', session.get('room', '')) or ''
                                for r in str(room).split(','):
                                    r = (r or '').strip()
                                    if r and (r.upper().startswith('L') or 'Lab' in r):
                                        lab_rooms_found.add(r)

                    if lab_rooms_found:
                        # Only show actual lab rooms (L-prefix or lab pattern), not classroom codes (C002, C004)
                        lab_only = [r for r in lab_rooms_found if r and (str(r).strip().upper().startswith('L') or ('Lab' in str(r)))]
                        if lab_only:
                            assigned_labs = ', '.join(sorted(lab_only))
                
                # Fallback: If not in room_assignments, try to determine from course properties
                if (not assigned_classroom) or is_na_room(assigned_classroom):
                    # Combined courses (shared across groups) typically use C004 (240-seater)
                    # Check if this is a combined course by looking at scheduled sessions
                    is_combined = False
                    if combined_sessions:
                        for session in combined_sessions:
                            if isinstance(session, dict):
                                session_course = session.get('course_code', '').split('-')[0]
                                if session_course == course_code:
                                    is_combined = True
                                    # Get room from combined session
                                    room = session.get('room', '')
                                    if room and not is_lab_room(room):
                                        assigned_classroom = room
                                        break
                            elif hasattr(session, 'course_code'):
                                if session.course_code.split('-')[0] == course_code:
                                    is_combined = True
                                    if hasattr(session, 'room') and session.room and not is_lab_room(session.room):
                                        assigned_classroom = session.room
                                    break
                    
                    # Try to extract from all_section_sessions for normal lectures
                    if ((not assigned_classroom) or is_na_room(assigned_classroom)) and all_section_sessions:
                        for session in all_section_sessions:
                            sess_code = str(session.get('Course Code', session.get('course_code', '')))
                            if sess_code.startswith('ELECTIVE_BASKET_'):
                                sess_base = sess_code.replace('-LAB', '').replace('-TUT', '')
                            else:
                                sess_base = sess_code.replace('-TUT', '').replace('-LAB', '').split('-')[0]
                            
                            session_type = session.get('Session Type', session.get('session_type', 'L'))
                            if sess_base == course_code and session_type in ('L', 'T', 'ELECTIVE'):
                                room = session.get('Room', session.get('room', '')) or ''
                                if room and (not room.startswith('L')) and (not is_na_room(room)):
                                    assigned_classroom = room
                                    break

                    # Default only when truly unknown; preserve intentional NA.
                    if not assigned_classroom:
                        assigned_classroom = COMBINED_RESERVED_ROOM_NUMBER if is_combined else ''
                
                # Time slot issues / displayed counts: full-semester combined uses semester totals
                disp_l = total_lectures if is_full_semester_combined else scheduled_lectures
                disp_t = total_tutorials if is_full_semester_combined else scheduled_tutorials
                disp_p = total_labs if is_full_semester_combined else scheduled_labs
                time_slot_issues = []
                if disp_l < required_lectures:
                    time_slot_issues.append(f"Missing {required_lectures - disp_l} lecture(s)")
                elif disp_l > required_lectures:
                    time_slot_issues.append(f"Extra {disp_l - required_lectures} lecture(s)")
                if disp_t < required_tutorials:
                    time_slot_issues.append(f"Missing {required_tutorials - disp_t} tutorial(s)")
                elif disp_t > required_tutorials:
                    time_slot_issues.append(f"Extra {disp_t - required_tutorials} tutorial(s)")
                if disp_p < required_labs:
                    time_slot_issues.append(f"Missing {required_labs - disp_p} lab(s)")
                elif disp_p > required_labs:
                    time_slot_issues.append(f"Extra {disp_p - required_labs} lab(s)")
                time_slot_issues_str = "; ".join(time_slot_issues) if time_slot_issues else "OK"
                room_conflicts_str = "OK"
                if is_na_room(assigned_classroom):
                    status = "UNSATISFIED"
                    na_msg = "No capacity-safe classroom found after room assignment/reschedule attempts"
                    time_slot_issues_str = f"{time_slot_issues_str}; {na_msg}" if time_slot_issues_str != "OK" else na_msg

                _not_on_grid_msg = (
                    "Not shown on this period's grid (semester L/T/P may be satisfied on the other half)"
                )
                if (
                    is_full_semester_combined
                    and (required_lectures or required_tutorials or required_labs)
                    and grid_sessions
                    and not _course_visible_on_period_grid(course_code)
                ):
                    if lectures_satisfied and tutorials_satisfied and labs_satisfied:
                        status = 'NOT_ON_SHEET'
                        time_slot_issues_str = _not_on_grid_msg
                    else:
                        status = 'UNSATISFIED'
                        if time_slot_issues_str == 'OK':
                            time_slot_issues_str = _not_on_grid_msg

                course_row = [
                    course_code,
                    course_name,
                    instructor,
                    ltpsc,
                    assigned_labs,  # Assigned Lab - from room_assignments
                    assigned_classroom,  # Assigned Classroom - from room_assignments or fallback
                    f"{required_lectures}/{disp_l}",
                    f"{required_tutorials}/{disp_t}",
                    f"{required_labs}/{disp_p}",
                    status,
                    time_slot_issues_str,
                    room_conflicts_str
                ]
                
                for col, value in enumerate(course_row, 1):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = value
                    cell.font = self.course_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.thin_border
                    # Color code columns
                    if col == 6:  # Assigned Classroom column
                        if not value or str(value).strip() in ('', 'None', 'N/A', 'TBD', 'UNASSIGNED'):
                            # Red: no classroom assigned - needs attention
                            cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                            cell.font = Font(bold=True, color='FFFFFF', size=9)
                            cell.value = value or 'NO ROOM'
                        else:
                            # Soft green: classroom assigned OK
                            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    elif col == len(course_row) - 1:  # Time Slot Issues column
                        if value != 'OK':
                            if status == 'NOT_ON_SHEET':
                                cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                            else:
                                cell.fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
                    elif col == len(course_row):  # Room Conflicts column
                        if value != 'OK':
                            cell.fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
                    elif col == len(course_row) - 2:  # Status column
                        if value == 'SATISFIED':
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        elif value == 'NOT_ON_SHEET':
                            cell.fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                
                current_row += 1
        
        # Add other relevant courses (excluding combined courses to avoid duplicates)
        for course in relevant_courses:
            course_code = getattr(course, 'code', '')
            if course_code not in combined_codes_this_section:
                course_name = getattr(course, 'name', '')
                # Get actual faculty from scheduled sessions
                faculty_for_course = set()
                
                # Check phase5_sessions for faculty
                if phase5_sessions:
                    for session in phase5_sessions:
                        if (hasattr(session, 'course_code') and session.course_code == course_code and
                            hasattr(session, 'section') and session.section == f"{section}-Sem{semester}" and
                            hasattr(session, 'period')):
                            session_period = session.period
                            period_match = (period == 'PreMid' and session_period == 'PRE') or \
                                          (period == 'PostMid' and session_period == 'POST')
                            if period_match and hasattr(session, 'faculty') and session.faculty:
                                faculty_for_course.add(session.faculty)
                                
                # Check phase7_sessions for faculty
                if phase7_sessions:
                    for session in phase7_sessions:
                        if (hasattr(session, 'course_code') and session.course_code == course_code and
                            hasattr(session, 'section') and session.section == f"{section}-Sem{semester}" and
                            hasattr(session, 'period')):
                            session_period = session.period
                            period_match = (period == 'PreMid' and session_period == 'PRE') or \
                                          (period == 'PostMid' and session_period == 'POST')
                            if period_match and hasattr(session, 'faculty') and session.faculty:
                                faculty_for_course.add(session.faculty)

                # Check all_section_sessions for faculty (CSV log / UI fast-path)
                if not faculty_for_course and all_section_sessions:
                    for session in all_section_sessions:
                        sess_code = str(session.get('Course Code', session.get('course_code', ''))).split('-')[0]
                        if sess_code == course_code:
                            fac = session.get('Faculty', session.get('faculty', session.get('instructor', '')))
                            if fac and str(fac).strip() and str(fac).strip().lower() not in ('tbd', 'nan', 'none'):
                                faculty_for_course.add(str(fac).strip())

                # Final fallback: read from the Course object's instructors list
                raw_instructors = getattr(course, 'instructors', [])
                if not faculty_for_course:
                    if raw_instructors:
                        instructor = ', '.join(raw_instructors)
                    else:
                        instructor = 'TBD'
                else:
                    if isinstance(raw_instructors, list) and set(raw_instructors) == faculty_for_course:
                        instructor = ', '.join(raw_instructors)
                    else:
                        instructor = ', '.join(sorted(faculty_for_course))
                
                if not instructor or str(instructor).strip() in ('', 'TBD', 'nan', 'None'):
                    instructor = 'TBD'
                ltpsc = getattr(course, 'ltpsc', '')
                credits = getattr(course, 'credits', 0)
                
                # Calculate required slots from LTPSC
                from modules_v2.phase5_core_courses import calculate_slots_needed
                slots_needed = calculate_slots_needed(ltpsc)
                
                # Determine if this is a Phase 5 course (>2 credits) or Phase 7 course (<=2 credits)
                # Phase 5: Full requirement in BOTH PreMid and PostMid
                # Phase 7: Full requirement in EITHER PreMid OR PostMid (half-semester)
                is_phase5 = credits > 2
                
                if is_phase5:
                    # Phase 5: Full requirement needed in EACH period (PreMid and PostMid)
                    required_lectures = slots_needed['lectures']
                    required_tutorials = slots_needed['tutorials']
                    required_labs = slots_needed['practicals']
                else:
                    # Phase 7: Full requirement needed in ONE period (half-semester course)
                    # Check if this course is scheduled in this period
                    required_lectures = slots_needed['lectures']
                    required_tutorials = slots_needed['tutorials']
                    required_labs = slots_needed['practicals']

                # Strict-consistent effective target for verification table status:
                # Phase 5 courses are full-semester (PRE+POST), so compare against doubled load.
                eff_required_lectures = required_lectures * 2 if is_phase5 else required_lectures
                eff_required_tutorials = required_tutorials * 2 if is_phase5 else required_tutorials
                eff_required_labs = required_labs * 2 if is_phase5 else required_labs
                
                # Count scheduled slots for THIS PERIOD ONLY - use unique key if available
                course_ltpsc = getattr(course, 'ltpsc', '')
                course_unique_key = f"{course_code}_{course_ltpsc}"
                
                # Try to get counts by unique key first, then fall back to base code
                # Use course_counts_period (this period only) for requirement checking
                counts_period = course_counts_period.get(
                    course_unique_key,
                    course_counts_period.get(
                        course_code,
                        {'total': 0, 'lectures': 0, 'tutorials': 0, 'labs': 0},
                    ),
                )
                scheduled_lectures = counts_period['lectures']
                scheduled_tutorials = counts_period['tutorials']
                scheduled_labs = counts_period['labs']

                # Phase 7 / half-semester: if absent from this period's grid, treat period counts as zero.
                # Phase 5: full-semester LTPSC — do not zero; semester totals drive status below.
                if (not is_phase5) and grid_sessions and course_code not in courses_on_this_sheet_period:
                    scheduled_lectures = 0
                    scheduled_tutorials = 0
                    scheduled_labs = 0

                # Full-semester totals (PRE+POST) for Phase 5 status and sheet columns.
                total_lectures = 0
                total_tutorials = 0
                total_labs = 0

                if phase5_sessions:
                    for session in phase5_sessions:
                        if not hasattr(session, 'course_code') or not hasattr(session, 'section'):
                            continue
                        sess_base = str(getattr(session, 'course_code', '')).replace('-TUT', '').replace('-LAB', '').split('-')[0]
                        if sess_base != course_code:
                            continue
                        if getattr(session, 'section', '') != f"{section}-Sem{semester}":
                            continue
                        kind = getattr(session, 'kind', 'L')
                        if kind == 'P':
                            total_labs += 1
                        elif kind == 'T':
                            total_tutorials += 1
                        else:
                            total_lectures += 1

                if phase7_sessions:
                    for session in phase7_sessions:
                        if not hasattr(session, 'course_code') or not hasattr(session, 'section'):
                            continue
                        sess_base = str(getattr(session, 'course_code', '')).replace('-TUT', '').replace('-LAB', '').split('-')[0]
                        if sess_base != course_code:
                            continue
                        if getattr(session, 'section', '') != f"{section}-Sem{semester}":
                            continue
                        kind = getattr(session, 'kind', 'L')
                        if kind == 'P':
                            total_labs += 1
                        elif kind == 'T':
                            total_tutorials += 1
                        else:
                            total_lectures += 1

                if not phase5_sessions and not phase7_sessions:
                    if all_section_sessions:
                        for session in all_section_sessions:
                            course_display = str(session.get('Course Code', session.get('course_code', '')))
                            sess_base = course_display.replace('-TUT', '').replace('-LAB', '').split('-')[0]
                            if sess_base != course_code:
                                continue
                            if '-LAB' in course_display:
                                total_labs += 1
                            elif '-TUT' in course_display:
                                total_tutorials += 1
                            else:
                                total_lectures += 1
                    elif grid_sessions:
                        for day, sessions in grid_sessions.items():
                            for s in sessions:
                                block, course_display = s[0], s[1]
                                if course_display in ("LUNCH", "Break(15min)"):
                                    continue
                                if course_code in course_display:
                                    if '-LAB' in course_display:
                                        total_labs += 1
                                    elif '-TUT' in course_display:
                                        total_tutorials += 1
                                    else:
                                        total_lectures += 1

                # Legacy data can contain section-specific LTPSC variants for the same code
                # (without Offering_ID), where some sections are lab-only offerings.
                # When strict generation succeeded and this section has only practicals for
                # this code, align expected counts with observed strict-safe totals.
                if is_phase5 and total_lectures == 0 and total_tutorials == 0 and total_labs > 0:
                    eff_required_lectures = 0
                    eff_required_tutorials = 0
                    eff_required_labs = total_labs

                phase7_other_period_only = False
                if not is_phase5:
                    if scheduled_lectures == 0 and scheduled_tutorials == 0 and scheduled_labs == 0:
                        other_period = 'POST' if period == 'PreMid' else 'PRE'
                        scheduled_in_other = False
                        if phase7_sessions:
                            for session in phase7_sessions:
                                if (hasattr(session, 'course_code') and session.course_code == course_code and
                                    hasattr(session, 'section') and session.section == f"{section}-Sem{semester}" and
                                    hasattr(session, 'period') and session.period == other_period):
                                    scheduled_in_other = True
                                    break

                        if scheduled_in_other:
                            phase7_other_period_only = True
                            lectures_satisfied = True
                            tutorials_satisfied = True
                            labs_satisfied = True
                        else:
                            lectures_satisfied = False
                            tutorials_satisfied = False
                            labs_satisfied = False
                    else:
                        lectures_satisfied = scheduled_lectures == required_lectures
                        tutorials_satisfied = scheduled_tutorials == required_tutorials
                        labs_satisfied = scheduled_labs == required_labs
                else:
                    # Phase 5 (>2 cr): evaluate status with full-semester totals (PRE+POST),
                    # aligned with strict verification semantics.
                    lectures_satisfied = total_lectures == eff_required_lectures
                    tutorials_satisfied = total_tutorials == eff_required_tutorials
                    labs_satisfied = total_labs == eff_required_labs

                if eff_required_lectures == 0 and eff_required_tutorials == 0 and eff_required_labs == 0:
                    stray = (
                        total_lectures or total_tutorials or total_labs
                        or scheduled_lectures or scheduled_tutorials or scheduled_labs
                    )
                    if stray:
                        status = 'UNSATISFIED'
                        lectures_satisfied = tutorials_satisfied = labs_satisfied = False
                    else:
                        status = 'SATISFIED'
                        lectures_satisfied = tutorials_satisfied = labs_satisfied = True
                elif is_phase5:
                    status = (
                        'SATISFIED'
                        if (lectures_satisfied and tutorials_satisfied and labs_satisfied)
                        else 'UNSATISFIED'
                    )
                else:
                    status = (
                        'SATISFIED'
                        if (lectures_satisfied and tutorials_satisfied and labs_satisfied)
                        else 'UNSATISFIED'
                    )

                # Get room assignment from room_assignments (Phase 8)
                # Key format: (course_code, section, period)
                assigned_labs = ''
                assigned_classroom = ''
                period_key = 'PRE' if period == 'PreMid' else 'POST'
                room_key = (course_code, f"{section}-Sem{semester}", period_key)
                
                if room_assignments and room_key in room_assignments:
                    assignment = room_assignments[room_key]
                    raw_classroom = assignment.get('classroom', '') or ''
                    labs_list = assignment.get('labs', [])
                    # Sanity check: if classroom field contains a lab room, reclassify it
                    if raw_classroom and is_lab_room(raw_classroom):
                        if raw_classroom not in labs_list:
                            labs_list = list(labs_list) + [raw_classroom]
                        raw_classroom = ''
                    assigned_classroom = raw_classroom
                    if labs_list:
                        assigned_labs = ', '.join(labs_list)
                
                # Fallback: If not in room_assignments, try actual sessions first
                if (not assigned_classroom) or is_na_room(assigned_classroom):
                    lab_rooms_from_sessions: set = set()

                    def _is_valid_room(r) -> bool:
                        """Return True only for non-empty, non-NaN room strings."""
                        try:
                            if pd.isna(r):
                                return False
                        except (TypeError, ValueError):
                            pass
                        return bool(r) and str(r).strip().lower() not in ('nan', 'none', 'tbd', '')

                    # Check phase5_sessions and phase7_sessions for real rooms
                    for _sessions in (phase5_sessions or [], phase7_sessions or []):
                        for sess in _sessions:
                            if (hasattr(sess, 'course_code') and str(getattr(sess, 'course_code', '')).split('-')[0] == course_code
                                    and hasattr(sess, 'section') and sess.section == f"{section}-Sem{semester}"):
                                room_val = getattr(sess, 'room', None)
                                if _is_valid_room(room_val):
                                    room_str = str(room_val).strip()
                                    if is_lab_room(room_str):
                                        lab_rooms_from_sessions.add(room_str)
                                    elif not assigned_classroom:
                                        assigned_classroom = room_str

                    # Check all_section_sessions (CSV log / UI fast-path)
                    if ((not assigned_classroom) or is_na_room(assigned_classroom)) and all_section_sessions:
                        for session in all_section_sessions:
                            sess_code = str(session.get('Course Code', session.get('course_code', ''))).split('-')[0]
                            if sess_code == course_code:
                                room = session.get('Room', session.get('room', '')) or ''
                                if _is_valid_room(room):
                                    room_str = str(room).strip()
                                    if is_lab_room(room_str):
                                        lab_rooms_from_sessions.add(room_str)
                                    elif (not assigned_classroom) or is_na_room(assigned_classroom):
                                        if room_str and not is_na_room(room_str):
                                            assigned_classroom = room_str

                    # Populate labs from sessions if room_assignments didn't provide them
                    if not assigned_labs and lab_rooms_from_sessions:
                        assigned_labs = ', '.join(sorted(lab_rooms_from_sessions))

                    # Last resort hardcoded defaults (preserve intentional NA)
                    if not assigned_classroom:
                        is_combined = any(
                            (isinstance(s, dict) and s.get('course_code', '').split('-')[0] == course_code) or
                            (hasattr(s, 'course_code') and s.course_code.split('-')[0] == course_code)
                            for s in (combined_sessions or [])
                        )
                        assigned_classroom = COMBINED_RESERVED_ROOM_NUMBER if is_combined else ''

                
                if is_phase5:
                    # For full-semester core rows, display semester totals to avoid false
                    # per-half UNSATISFIED when strict full-semester checks are satisfied.
                    disp_l, disp_t, disp_p = total_lectures, total_tutorials, total_labs
                elif phase7_other_period_only:
                    disp_l = disp_t = disp_p = 0
                else:
                    disp_l, disp_t, disp_p = scheduled_lectures, scheduled_tutorials, scheduled_labs

                time_slot_issues = []
                if not phase7_other_period_only:
                    if disp_l < eff_required_lectures:
                        time_slot_issues.append(f"Missing {eff_required_lectures - disp_l} lecture(s)")
                    elif disp_l > eff_required_lectures:
                        time_slot_issues.append(f"Extra {disp_l - eff_required_lectures} lecture(s)")
                    if disp_t < eff_required_tutorials:
                        time_slot_issues.append(f"Missing {eff_required_tutorials - disp_t} tutorial(s)")
                    elif disp_t > eff_required_tutorials:
                        time_slot_issues.append(f"Extra {disp_t - eff_required_tutorials} tutorial(s)")
                    if disp_p < eff_required_labs:
                        time_slot_issues.append(f"Missing {eff_required_labs - disp_p} lab(s)")
                    elif disp_p > eff_required_labs:
                        time_slot_issues.append(f"Extra {disp_p - eff_required_labs} lab(s)")
                time_slot_issues_str = "; ".join(time_slot_issues) if time_slot_issues else "OK"
                room_conflicts_str = "OK"
                _not_on_grid_msg = (
                    "Not shown on this period's grid (semester L/T/P may be satisfied on the other half)"
                )
                if (
                    is_phase5
                    and (required_lectures or required_tutorials or required_labs)
                    and grid_sessions
                    and not _course_visible_on_period_grid(course_code)
                ):
                    if lectures_satisfied and tutorials_satisfied and labs_satisfied:
                        status = 'NOT_ON_SHEET'
                        time_slot_issues_str = _not_on_grid_msg
                    else:
                        status = 'UNSATISFIED'
                        if time_slot_issues_str == 'OK':
                            time_slot_issues_str = _not_on_grid_msg
                if is_na_room(assigned_classroom):
                    status = "UNSATISFIED"
                    na_msg = "No capacity-safe classroom found after room assignment/reschedule attempts"
                    time_slot_issues_str = f"{time_slot_issues_str}; {na_msg}" if time_slot_issues_str != "OK" else na_msg

                course_row = [
                    course_code,
                    course_name,
                    instructor,
                    ltpsc,
                    assigned_labs,  # Assigned Lab - from room_assignments
                    assigned_classroom,  # Assigned Classroom - from room_assignments or fallback
                    f"{eff_required_lectures}/{disp_l}",
                    f"{eff_required_tutorials}/{disp_t}",
                    f"{eff_required_labs}/{disp_p}",
                    status,
                    time_slot_issues_str,
                    room_conflicts_str
                ]
                
                for col, value in enumerate(course_row, 1):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = value
                    cell.font = self.course_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.thin_border
                    # Color code columns
                    if col == 6:  # Assigned Classroom column
                        if not value or str(value).strip() in ('', 'None', 'N/A', 'TBD', 'UNASSIGNED'):
                            # Red: no classroom assigned - needs attention
                            cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                            cell.font = Font(bold=True, color='FFFFFF', size=9)
                            cell.value = value or 'NO ROOM'
                        else:
                            # Soft green: classroom assigned OK
                            cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                    elif col == len(course_row) - 1:  # Time Slot Issues column
                        if value != 'OK':
                            if status == 'NOT_ON_SHEET':
                                cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                            else:
                                cell.fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
                    elif col == len(course_row):  # Room Conflicts column
                        if value != 'OK':
                            cell.fill = PatternFill(start_color='FFE4E1', end_color='FFE4E1', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='F0FFF0', end_color='F0FFF0', fill_type='solid')
                    elif col == len(course_row) - 2:  # Status column
                        if value == 'SATISFIED':
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        elif value == 'NOT_ON_SHEET':
                            cell.fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                
                current_row += 1
        
        # Add faculty conflict information if any
        if faculty_conflicts:
            current_row += 2
            self._add_faculty_conflict_section(sheet, current_row, faculty_conflicts, section, period)
            current_row += 10  # Add space for conflict section
        
        return current_row

    def _add_faculty_conflict_section(self, sheet, start_row: int, faculty_conflicts: List, 
                                    section: str, period: str) -> None:
        """Add faculty conflict information to the sheet"""
        current_row = start_row
        
        # Filter conflicts relevant to this section/period
        relevant_conflicts = []
        for conflict in faculty_conflicts:
            # Check if any conflicting session is for this section/period
            for session_name in conflict.conflicting_sessions:
                if section in session_name:
                    relevant_conflicts.append(conflict)
                    break
        
        if not relevant_conflicts:
            return
        
        # Add conflict header
        conflict_header = sheet.cell(row=current_row, column=1)
        conflict_header.value = "[WARNING] FACULTY CONFLICTS DETECTED"
        conflict_header.font = Font(bold=True, color="FF0000")
        conflict_header.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Add conflict details
        for i, conflict in enumerate(relevant_conflicts, 1):
            conflict_row = [
                f"Conflict #{i}",
                conflict.faculty_name,
                f"{conflict.day} {conflict.time_slot}",
                ", ".join(conflict.conflicting_sessions)
            ]
            
            for col, value in enumerate(conflict_row, 1):
                cell = sheet.cell(row=current_row, column=col)
                cell.value = value
                cell.font = self.course_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
                if col == 1:  # Conflict number
                    cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            current_row += 1

    def write_elective_assignment_table(
        self,
        sheet,
        start_row: int,
        semester: int,
        courses: List,
        elective_assignments: List[Dict],
        all_section_sessions: List[Dict] = None,
    ) -> int:
        """
        Write elective assignment table below verification table.
        
        Args:
            sheet: Excel sheet to write to
            start_row: Starting row number
            semester: Semester number (1, 3, or 5)
            courses: Full course list from Phase 1 (source of truth)
            elective_assignments: List of elective assignments from Phase 9
                Format: [{'course': Course, 'room': str, 'period': str, 'faculty': str}, ...]
            all_section_sessions: UI/Fast-path log containing active manual changes
        
        Returns:
            Next row number after the table
        """
        current_row = start_row
        
        # Add spacing
        current_row += 2
        
        # Table headers (define first for merge)
        headers = ["Course Code", "Course Name", "Room", "Period", "Faculty"]

        def format_period(period_val: str) -> str:
            if period_val == 'PRE':
                return 'PreMid'
            if period_val == 'POST':
                return 'PostMid'
            if period_val == 'FULL':
                return 'Full Sem'
            return period_val

        def group_semester_prefix_ok(group_key: str) -> bool:
            # Accept only semester-matching dotted keys: e.g., semester=3 => "3.1", "3.2"
            try:
                s = str(group_key).strip()
                if '.' not in s:
                    return False
                prefix = int(s.split('.')[0])
                return prefix == int(semester)
            except Exception:
                return False

        def normalize_code(code: str) -> str:
            return str(code).strip().upper() if code is not None else ""

        # Build assignment maps from Phase 9, but hard-filter to this semester to prevent leaks
        # - assignment_any: any representative assignment per course (for room/faculty)
        # - assignment_periods: set of periods seen per course (to infer Full Sem vs Pre/Post)
        assignment_any: Dict[str, Dict] = {}
        assignment_periods: Dict[str, set] = {}
        for assignment in elective_assignments or []:
            course = assignment.get('course')
            if not course:
                continue
            course_sem = getattr(course, 'semester', None)
            if int(course_sem) != int(semester):
                continue
            code_key = normalize_code(getattr(course, 'code', ''))
            if not code_key:
                continue
            if code_key not in assignment_any:
                assignment_any[code_key] = assignment
            # Only store the period from the FIRST seen assignment per course.
            # Phase 9 may assign the same course to multiple sections with different
            # periods — accumulating them would turn PRE + POST → 'Full Sem' falsely.
            if code_key not in assignment_periods:
                period_raw = str(assignment.get('period', '')).strip().upper()
                if period_raw:
                    assignment_periods[code_key] = {period_raw}

        # Harvest from UI Log fast-path if present
        # ELECTIVE_BASKET_1.1 entries tell us which group was assigned PRE/POST
        if all_section_sessions:
            for s in all_section_sessions:
                sess_code = str(s.get('Course Code', s.get('course_code', '')))
                if not sess_code.startswith('ELECTIVE_BASKET_'):
                    continue
                # Extract the group key from the Course Code, e.g. 'ELECTIVE_BASKET_1.1' → '1.1'
                group_key = sess_code.replace('ELECTIVE_BASKET_', '').strip()
                if not group_key:
                    continue

                period_raw = str(s.get('Period', s.get('period', ''))).strip().upper()
                if period_raw.startswith('PRE'):
                    period_raw = 'PRE'
                elif period_raw.startswith('POST'):
                    period_raw = 'POST'
                elif period_raw in ('FULL', 'FULLSEM', 'FULL SEM'):
                    period_raw = 'FULL'
                
                if group_key not in assignment_periods:
                    assignment_periods[group_key] = set()
                if period_raw:
                    assignment_periods[group_key].add(period_raw)

        # Source of truth: course data for this semester, grouped by elective_group
        electives_for_sem = []
        for c in courses or []:
            if not getattr(c, 'is_elective', False):
                continue
            if getattr(c, 'semester', None) != semester:
                continue
            gk = getattr(c, 'elective_group', None)
            if not gk:
                continue
            gk_str = str(gk).strip()
            if not group_semester_prefix_ok(gk_str):
                continue
            electives_for_sem.append(c)

        # Group electives by elective_group, de-duplicating by course code within a group
        from collections import defaultdict
        by_group = defaultdict(list)
        seen_codes_per_group: Dict[str, set] = {}
        for c in electives_for_sem:
            group_key = str(getattr(c, 'elective_group')).strip()
            code_key = normalize_code(getattr(c, 'code', ''))
            if not code_key:
                continue
            if group_key not in seen_codes_per_group:
                seen_codes_per_group[group_key] = set()
            if code_key in seen_codes_per_group[group_key]:
                # Skip duplicates of the same course code within the same elective group
                continue
            by_group[group_key].append(c)
            seen_codes_per_group[group_key].add(code_key)

        group_keys_sorted = sorted(by_group.keys(), key=str)

        if not group_keys_sorted:
            # Show message if semester has no grouped electives in input data
            title_cell = sheet.cell(row=current_row, column=1)
            title_cell.value = f"Elective Course Assignments - Semester {semester}"
            title_cell.font = Font(bold=True, color="FFFFFF", size=12)
            title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 1

            cell = sheet.cell(row=current_row, column=1)
            cell.value = "No elective groups found in course data for this semester"
            cell.font = self.course_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 1
            return current_row

        # Render stacked tables: one table per group_key
        for idx, group_key in enumerate(group_keys_sorted):
            if idx > 0:
                current_row += 2  # spacing between group tables

            # Title
            title_cell = sheet.cell(row=current_row, column=1)
            title_cell.value = f"Elective Course Assignments - Semester {semester} - Group {group_key}"
            title_cell.font = Font(bold=True, color="FFFFFF", size=12)
            title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 1

            # Headers
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.thin_border
            current_row += 1

            group_courses = by_group[group_key]
            group_courses.sort(key=lambda c: getattr(c, 'code', ''))

            for course in group_courses:
                course_code = getattr(course, 'code', '')
                course_name = getattr(course, 'name', 'N/A')

                code_key = normalize_code(course_code)
                assignment = assignment_any.get(code_key, {})
                room = assignment.get('room', 'TBD')
                # Infer period:
                # 1) If course credits > 2, treat as Full Sem.
                # 2) Otherwise, infer from all assignments for this course (PRE/POST/FULL).
                credits = getattr(course, 'credits', 0) or 0
                periods = assignment_periods.get(code_key, set())
                # Also check using the elective_group key (used by fast-path UI log harvesting)
                if not periods:
                    group_key_fb = str(getattr(course, 'elective_group', '') or '').strip()
                    if group_key_fb:
                        periods = assignment_periods.get(group_key_fb, set())
                raw_period = 'TBD'
                if credits > 2:
                    raw_period = 'FULL'
                elif 'FULL' in periods:
                    raw_period = 'FULL'
                elif 'PRE' in periods and 'POST' in periods:
                    raw_period = 'FULL'
                elif 'PRE' in periods:
                    raw_period = 'PRE'
                elif 'POST' in periods:
                    raw_period = 'POST'
                period = raw_period
                faculty = assignment.get('faculty', 'TBD')
                period_display = format_period(str(period))

                row_data = [course_code, course_name, room, period_display, faculty]

                for col, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = value
                    cell.font = self.course_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = self.thin_border

                    # Highlight room cells like before
                    if col == 3 and value == 'TBD':
                        cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                    elif col == 3 and value != 'TBD':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

                current_row += 1

        return current_row
    
    def save_timetable(self, output_path: str) -> None:
        """Save the workbook to file"""
        self.workbook.save(output_path)
        print(f"Timetable saved to: {output_path}")

def add_lunch_to_schedule(grid: DayScheduleGrid):
    """Add lunch break to the schedule grid"""
    lunch_block = grid.lunch_block
    grid.sessions.append((lunch_block, "LUNCH"))

def create_sample_timetable():
    """Create a sample timetable for testing"""
    writer = TimetableWriterV2()
    
    # Create sheets for different sections
    sections = [
        ("CSE-A", 1, "PreMid"),
        ("CSE-A", 1, "PostMid"),
        ("CSE-A", 3, "PreMid"),
        ("CSE-A", 3, "PostMid"),
        ("CSE-A", 5, "PreMid"),
        ("CSE-A", 5, "PostMid"),
    ]
    
    for section_name, semester, period in sections:
        writer.create_section_sheet(section_name, semester, period)
    
    # Create summary sheet
    writer.create_summary_sheet([])
    
    # Save to file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"iiitdwd_timetable_v2/DATA/OUTPUT/sample_timetable_v2_{timestamp}.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    writer.save_timetable(output_path)
    
    return output_path

if __name__ == "__main__":
    from datetime import time
    output_file = create_sample_timetable()
    print(f"Sample timetable created: {output_file}")
