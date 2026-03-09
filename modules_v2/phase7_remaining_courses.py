"""
Phase 7: Remaining ≤2 Credit Courses Scheduling
Handles courses that are not scheduled in Phase 4 (combined) but still need
half-semester (PreMid/PostMid) placement.

This version makes period (PreMid/PostMid) selection user-driven on a
per-course-per-section basis, persisted to an Excel file, similar to Phase 4.
"""

import os
import sys
import math
from datetime import time, datetime, timedelta
from typing import List, Dict, Tuple
from collections import defaultdict

import openpyxl

# Add the current directory to Python path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from utils.data_models import Course, Section, TimeBlock, ClassRoom, ScheduledSession, section_has_time_conflict
from utils.time_slot_logger import get_logger
from utils.session_rules_validator import SessionRulesValidator
from utils.time_validator import validate_time_range
from config.schedule_config import WORKING_DAYS, LUNCH_WINDOWS, DAY_START_TIME, DAY_END_TIME

def calculate_slots_from_ltpsc(ltpsc: str) -> Dict[str, int]:
    """Calculate slots using LTPSC with ceiling for lectures"""
    try:
        parts = str(ltpsc).split('-')
        if len(parts) != 5:
            return {'lectures': 2, 'tutorials': 1, 'practicals': 0, 'total': 3}
        
        L = int(parts[0]) if parts[0] else 0
        T = int(parts[1]) if parts[1] else 0
        P = int(parts[2]) if parts[2] else 0
        
        lectures = math.ceil(L / 1.5)
        tutorials = int(T / 1)
        practicals = int(P / 2)
        
        return {
            'lectures': lectures,
            'tutorials': tutorials,
            'practicals': practicals,
            'total': lectures + tutorials + practicals
        }
    except:
        return {'lectures': 2, 'tutorials': 1, 'practicals': 0, 'total': 3}


def get_phase7_period_assignments_path() -> str:
    """
    Return absolute path for Phase 7 period assignment Excel file.

    File: DATA/INPUT/phase7_period_assignments.xlsx
    """
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_dir, "DATA", "INPUT", "phase7_period_assignments.xlsx")


def load_phase7_period_assignments(path: str) -> Dict[Tuple[str, int, str], bool]:
    """
    Load per-course, per-section period assignments from Excel.

    Returns:
        {(course_code, semester, section_label): True for PreMid, False for PostMid}
    """
    assignments: Dict[Tuple[str, int, str], bool] = {}

    if not os.path.exists(path):
        return assignments

    try:
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
    except Exception:
        # Corrupt or unreadable file – ignore and start fresh
        return assignments

    # Invalidate old assignments if the section grouping changed
    try:
        from config.structure_config import get_grouping_signature
        expected_sig = get_grouping_signature()
        meta = wb["META"] if "META" in wb.sheetnames else None
        if meta is None:
            return {}  # legacy -> re-ask
        stored_sig = None
        for row in meta.iter_rows(min_row=1, max_row=100):
            k = row[0].value
            v = row[1].value if len(row) > 1 else None
            if str(k).strip().upper() == "GROUPING_SIGNATURE":
                stored_sig = str(v) if v is not None else ""
                break
        if not stored_sig or stored_sig != expected_sig:
            return {}
    except Exception:
        return {}

    # Expect header: Course_Code | Semester | Section | Pre | Post
    header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if not header_row:
        return assignments

    header = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in header_row
    ]
    name_to_idx = {name.upper(): idx for idx, name in enumerate(header)}

    def _get_col(name: str) -> int:
        return name_to_idx.get(name.upper(), -1)

    col_code = _get_col("COURSE_CODE")
    col_sem = _get_col("SEMESTER")
    col_section = _get_col("SECTION")
    col_pre = _get_col("PRE")
    col_post = _get_col("POST")

    if min(col_code, col_sem, col_section, col_pre, col_post) < 0:
        return assignments

    for row in sheet.iter_rows(min_row=2):
        code_cell = row[col_code].value
        sem_cell = row[col_sem].value
        section_cell = row[col_section].value
        pre_cell = row[col_pre].value
        post_cell = row[col_post].value

        if not code_cell or sem_cell is None or not section_cell:
            continue

        try:
            code = str(code_cell).strip().upper()
            semester = int(sem_cell)
            section_label = str(section_cell).strip()
        except (TypeError, ValueError):
            continue

        pre_val = int(pre_cell) if pre_cell is not None else 0
        post_val = int(post_cell) if post_cell is not None else 0

        if pre_val == 1 and post_val == 0:
            assignments[(code, semester, section_label)] = True
        elif pre_val == 0 and post_val == 1:
            assignments[(code, semester, section_label)] = False
        else:
            # Invalid or ambiguous row – ignore; it will be re-asked
            continue

    return assignments


def save_phase7_period_assignments(
    path: str,
    assignments: Dict[Tuple[str, int, str], bool],
) -> None:
    """
    Persist Phase 7 period assignments to Excel.

    Each row: Course_Code | Semester | Section | Pre | Post
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Phase7Periods"

    # META sheet to track grouping signature (invalidate on group change)
    try:
        from config.structure_config import get_grouping_signature
        meta = wb.create_sheet(title="META")
        meta.append(["Key", "Value"])
        meta.append(["GROUPING_SIGNATURE", get_grouping_signature()])
    except Exception:
        pass

    sheet.append(["Course_Code", "Semester", "Section", "Pre", "Post"])

    for (code, semester, section_label) in sorted(
        assignments.keys(), key=lambda x: (x[1], x[0], x[2])
    ):
        is_premid = assignments[(code, semester, section_label)]
        pre_val = 1 if is_premid else 0
        post_val = 0 if is_premid else 1
        sheet.append([code, semester, section_label, pre_val, post_val])

    dir_name = os.path.dirname(path)
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name, exist_ok=True)

    wb.save(path)


def _normalize_period(raw: object) -> str:
    """Normalize period values to 'PRE' or 'POST'."""
    if raw is None:
        return "PRE"
    val = str(raw).strip().upper()
    if val in ("PRE", "PREMID"):
        return "PRE"
    if val in ("POST", "POSTMID"):
        return "POST"
    return val or "PRE"


def add_session_to_occupied_slots(
    session: object,
    occupied_slots: Dict[str, List[Tuple[TimeBlock, str]]],
) -> None:
    """
    Add a session (dict or ScheduledSession) into occupied_slots with
    normalized section and period keys.

    Key format: f"{section_label}_{period}", where period is 'PRE' or 'POST'.
    """
    if session is None:
        return

    # Dict format – used by Phase 4 combined sessions and some helpers
    if isinstance(session, dict):
        course_code = session.get("course_code", "")
        block = session.get("time_block") or session.get("block")
        if not isinstance(block, TimeBlock):
            return

        period = _normalize_period(session.get("period", "PRE"))

        sections = session.get("sections")
        if not sections and "section" in session:
            sections = [session.get("section")]
        if not sections:
            return

        for sec in sections:
            if not sec:
                continue
            section_key = f"{sec}_{period}"
            occupied_slots.setdefault(section_key, []).append((block, str(course_code)))
        return

    # ScheduledSession-like object
    if hasattr(session, "block") and hasattr(session, "section"):
        block = getattr(session, "block")
        if not isinstance(block, TimeBlock):
            return
        period_raw = getattr(session, "period", "PRE")
        period = _normalize_period(period_raw)
        
        # Try to use .sections list if present (added by Phase 4)
        sections = getattr(session, "sections", None)
        if not sections:
            section_label = str(getattr(session, "section", "")).strip()
            if not section_label:
                return
            # Split comma-separated labels if present
            sections = [s.strip() for s in section_label.split(',')]
            
        course_code = getattr(session, "course_code", "")
        
        for sec in sections:
            if not sec:
                continue
            section_key = f"{sec}_{period}"
            occupied_slots.setdefault(section_key, []).append((block, str(course_code)))


def prompt_user_for_phase7_period(
    course: Course,
    section_label: str,
) -> bool:
    """
    Ask the user whether this course+section should be in PreMid or PostMid.

    Returns:
        True  -> PreMid
        False -> PostMid
    """
    prompt = (
        f"\nPhase 7: Choose period for {course.code} - {course.name} "
        f"(Sem {course.semester}, section {section_label}).\n"
        f"  [1] PreMid\n"
        f"  [2] PostMid\n"
        f"Enter 1 or 2 (or 'pre'/'post'): "
    )

    while True:
        try:
            choice = input(prompt).strip().lower()
        except EOFError:
            # Non-interactive env: default to PreMid
            return True

        if choice in ("1", "pre", "premid", "pre-mid", "pre_mid"):
            return True
        if choice in ("2", "post", "postmid", "post-mid", "post_mid"):
            return False

        print("Invalid input. Please enter 1 for PreMid or 2 for PostMid.")


def get_phase7_period_assignments(
    within_group: List[List[Course]],
    non_combined: List[Course],
    sections: List[Section],
    assignments_path: str,
) -> Tuple[Dict[Tuple[str, int, str], bool], bool]:
    """
    Build / update Phase 7 period assignments for all remaining ≤2 credit courses.

    Returns:
        (assignments, any_prompted)
        - assignments: {(course_code, semester, section_label): True for PreMid, False for PostMid}
        - any_prompted: True if at least one prompt was shown this run
    """
    existing = load_phase7_period_assignments(assignments_path)
    assignments: Dict[Tuple[str, int, str], bool] = dict(existing)
    any_prompted = False

    # Helper to ensure a mapping for given course & section
    def ensure_assignment(course: Course, section_label: str) -> None:
        nonlocal any_prompted
        key = (course.code.upper(), course.semester, section_label)
        if key in assignments:
            return
        is_premid = prompt_user_for_phase7_period(course, section_label)
        assignments[key] = is_premid
        any_prompted = True

    # Within-group: course is common within a department's multiple sections
    seen_within: Dict[Tuple[str, int, str], Course] = {}
    for group in within_group:
        if not group:
            continue
        course = group[0]
        key = (course.code.upper(), course.semester, course.department)
        if key in seen_within:
            continue
        seen_within[key] = course

        dept_sections = [
            s for s in sections
            if s.program == course.department and s.semester == course.semester
        ]
        for sec in dept_sections:
            section_label = f"{sec.program}-{sec.name}-Sem{sec.semester}"
            ensure_assignment(course, section_label)

    # Non-combined: section-specific or only one section in that department
    seen_non: Dict[Tuple[str, int], Course] = {}
    for course in non_combined:
        key = (course.code.upper(), course.semester)
        if key in seen_non:
            continue
        seen_non[key] = course

        # Phase 7 currently assumes A-section for departments with only one section
        section_label = f"{course.department}-A-Sem{course.semester}"
        ensure_assignment(course, section_label)

    return assignments, any_prompted
def get_phase4_course_codes(combined_sessions: List = None) -> List[str]:
    """
    Get list of Phase 4 combined course codes dynamically from combined_sessions.
    Phase 7 must exclude these to avoid duplicate scheduling.
    """
    phase4_codes = set()
    if not combined_sessions:
        return list(phase4_codes)

    for session in combined_sessions:
        if hasattr(session, 'course_code'):
            base_code = session.course_code.split('-')[0]  # Remove -TUT, -LAB suffix
            phase4_codes.add(base_code)
        elif isinstance(session, dict) and 'course_code' in session:
            course_code = session['course_code']
            base_code = course_code.split('-')[0]
            phase4_codes.add(base_code)
    return list(phase4_codes)

def identify_phase7_courses(courses: List[Course], sections: List[Section], combined_sessions: List = None) -> Tuple[List[List[Course]], List[Course]]:
    """
    Identify ≤2 credit courses not in Phase 4
    Returns: (within_group_combined, non_combined)
    """
    phase4_codes = get_phase4_course_codes(combined_sessions)
    
    # Filter ≤2 credit core courses not in Phase 4
    remaining_courses = [c for c in courses 
                        if c.credits <= 2 
                        and not c.is_elective 
                        and c.code not in phase4_codes]
    
    print(f"Found {len(remaining_courses)} remaining <=2 credit courses not in Phase 4:")
    for course in remaining_courses:
        print(f"  {course.code} - {course.name} ({course.department}, Sem{course.semester}) - LTPSC: {course.ltpsc}")
    
    # Group by (department, semester, course_code) to find within-group combined
    dept_course_map = defaultdict(list)
    for course in remaining_courses:
        # Check if this department has multiple sections
        dept_sections = [s for s in sections if s.program == course.department and s.semester == course.semester]
        if len(dept_sections) > 1:
            # This department has multiple sections (e.g., CSE-A, CSE-B)
            # So this course can be within-group combined
            key = (course.department, course.semester, course.code)
            dept_course_map[key].append(course)
    
    # Separate within-group combined vs non-combined
    within_group_combined = []
    non_combined = []
    
    for course in remaining_courses:
        key = (course.department, course.semester, course.code)
        if key in dept_course_map:
            # This course can be within-group combined
            # Only add once (not for each section)
            if course not in [c for group in within_group_combined for c in group]:
                within_group_combined.append([course])
        else:
            # This course is non-combined (section-specific or only one section)
            non_combined.append(course)
    
    return within_group_combined, non_combined

def get_lunch_blocks() -> Dict[int, TimeBlock]:
    """Get lunch break time blocks for each semester"""
    blocks: Dict[int, TimeBlock] = {}
    for sem, (start_t, end_t) in LUNCH_WINDOWS.items():
        # Day here is a placeholder; callers should project it onto the slot's day.
        blocks[sem] = TimeBlock("Monday", start_t, end_t)
    return blocks

def is_slot_available(slot: TimeBlock, section: str, period: str, occupied_slots: Dict, lunch_block: TimeBlock) -> bool:
    """Check if a time slot is available"""
    # Check lunch overlap
    if lunch_block:
        lunch_for_day = TimeBlock(slot.day, lunch_block.start, lunch_block.end)
        if slot.overlaps(lunch_for_day):
            return False
    
    # Check occupied slots
    section_key = f"{section}_{period}"
    if section_key in occupied_slots:
        for occupied_block, _ in occupied_slots[section_key]:
            if slot.overlaps(occupied_block):
                return False

    # Extra safety: handle any legacy occupied-slot formats
    if section_has_time_conflict(occupied_slots, section, period, slot):
        return False
    
    return True


def check_elective_conflict_phase7(day: str, start: time, end: time, semester: int) -> bool:
    """
    Check if a time slot conflicts with any elective basket slot for this semester.

    NOTE: Phase 3 uses group keys like '3.1', '3.2' (not integer semester keys).
    """
    try:
        from modules_v2.phase3_elective_baskets_v2 import ELECTIVE_BASKET_SLOTS

        def extract_semester_from_group(gk: str) -> int:
            try:
                s = str(gk)
                if '.' in s:
                    return int(s.split('.')[0])
                return int(s)
            except Exception:
                return -1

        test_block = TimeBlock(day, start, end)
        for group_key, slots in (ELECTIVE_BASKET_SLOTS or {}).items():
            if extract_semester_from_group(group_key) != semester:
                continue
            for slot_type in ("lecture_1", "lecture_2", "tutorial"):
                elective_block = slots.get(slot_type) if isinstance(slots, dict) else None
                if elective_block and elective_block.day == day and test_block.overlaps(elective_block):
                    return True
    except Exception:
        return False
    return False

def find_available_slots_phase7(semester: int, section: str, period: str,
                                occupied_slots: Dict,
                                num_slots: int,
                                slot_durations: List[int]) -> List[TimeBlock]:
    """
    Find available time slots avoiding:
    - Electives (Phase 3)
    - Combined courses (Phase 4)
    - Core courses (Phase 5)
    - Lunch breaks
    
    EXPANDED: Use full day (9:00-18:00) with 15-minute intervals for maximum flexibility
    """
    days = list(WORKING_DAYS)
    available_slots = []
    
    lunch_blocks = get_lunch_blocks()
    lunch_block_base = lunch_blocks.get(semester)
    
    used_days = set()
    max_iterations = 5000  # Increased limit for full day search
    iteration_count = 0
    
    # Generate time slots dynamically across full day with 15-minute intervals.
    # Start/end are driven by schedule_config.DAY_START_TIME / DAY_END_TIME.
    start_hour, end_hour = DAY_START_TIME.hour, DAY_END_TIME.hour
    
    for duration in slot_durations:
        if len(available_slots) >= num_slots:
            break
            
        for day in days:
            if len(available_slots) >= num_slots:
                break
                
            if day in used_days:
                continue
            
            # Try all 15-minute intervals from 9:00 to 18:00 (excluding lunch)
            current_time = time(start_hour, 0)
            
            while current_time.hour < end_hour:
                iteration_count += 1
                if iteration_count > max_iterations:
                    print(f"  WARNING: Max iterations reached in find_available_slots_phase7")
                    return available_slots
                
                # Calculate end time
                start_minutes = current_time.hour * 60 + current_time.minute
                end_minutes = start_minutes + duration
                
                if end_minutes > end_hour * 60:
                    break  # Slot would extend beyond working hours
                
                end_time = time(end_minutes // 60, end_minutes % 60)
                slot = TimeBlock(day, current_time, end_time)
                
                # Check if slot overlaps with lunch
                if lunch_block_base:
                    lunch_block = TimeBlock(day, lunch_block_base.start, lunch_block_base.end)
                else:
                    lunch_block = None

                if lunch_block and slot.overlaps(lunch_block):
                    # Skip to after lunch
                    if current_time < lunch_block.start:
                        current_time = lunch_block.end
                    else:
                        # Already past lunch, move to next interval
                        current_dt = datetime.combine(datetime.min, current_time)
                        current_dt += timedelta(minutes=15)
                        current_time = current_dt.time()
                    continue
                
                # CRITICAL: Check elective basket conflict
                if check_elective_conflict_phase7(day, current_time, end_time, semester):
                    # Skip to after elective time
                    current_dt = datetime.combine(datetime.min, current_time)
                    current_dt += timedelta(minutes=15)
                    current_time = current_dt.time()
                    continue
                
                # Check if slot is available
                if is_slot_available(slot, section, period, occupied_slots, lunch_block):
                    available_slots.append(slot)
                    used_days.add(day)
                    # Move to next 15-minute interval for next slot
                    current_dt = datetime.combine(datetime.min, current_time)
                    current_dt += timedelta(minutes=15)
                    current_time = current_dt.time()
                    break  # Found slot for this duration, move to next duration
                
                # Move to next 15-minute interval
                current_dt = datetime.combine(datetime.min, current_time)
                current_dt += timedelta(minutes=15)
                current_time = current_dt.time()
            
            if len(available_slots) >= num_slots:
                return available_slots
    
    return available_slots

def assign_classroom(capacity_needed: int, classrooms: List[ClassRoom], slot: TimeBlock) -> ClassRoom:
    """Assign a classroom based on capacity"""
    # Filter classrooms that can accommodate the capacity
    suitable_rooms = [r for r in classrooms if r.capacity >= capacity_needed and r.room_type == 'classroom']
    
    if not suitable_rooms:
        return classrooms[0]  # Fallback
    
    # Sort by capacity (prefer closest match)
    suitable_rooms.sort(key=lambda r: r.capacity)
    return suitable_rooms[0]

def schedule_within_group_combined(course: Course, sections: List[Section],
                                   occupied_slots: Dict,
                                   classrooms: List[ClassRoom],
                                   room_occupancy: Dict,
                                   period_assignments: Dict[Tuple[str, int, str], bool]) -> List[ScheduledSession]:
    """
    Schedule within-group combined courses
    Conditionally combines if sections map to the same period and have a single instructor.
    """
    import time
    sessions = []
    
    # Get sections for this department and semester
    dept_sections = [
        s for s in sections
        if s.program == course.department and s.semester == course.semester
    ]
    dept_sections.sort(key=lambda s: s.name)  # Sort A, B, etc.
    
    # Calculate slots needed from LTPSC
    slots_info = calculate_slots_from_ltpsc(course.ltpsc)
    
    # Prepare slot durations
    slot_durations = []
    for _ in range(slots_info['lectures']):
        slot_durations.append(90)  # 1.5 hours
    for _ in range(slots_info['tutorials']):
        slot_durations.append(60)  # 1 hour
    for _ in range(slots_info['practicals']):
        slot_durations.append(120)  # 2 hours
        
    def schedule_labels(label_list: List[str], target_period: str, total_students: int):
        # Find available time slots for the first section
        base_slots = find_available_slots_phase7(
            course.semester, label_list[0], target_period, occupied_slots, 
            slots_info['total'] * 3, slot_durations
        )
        
        # Filter for all labels
        available_slots = []
        for slot in base_slots:
            valid = True
            for lbl in label_list[1:]:
                sec_key = f"{lbl}_{target_period}"
                for existing in occupied_slots.get(sec_key, []):
                    eb = existing[0] if isinstance(existing, tuple) else existing
                    if eb and slot.overlaps(eb):
                        valid = False
                        break
                if not valid:
                    break
            if valid:
                available_slots.append(slot)
                
        # Fill strictly if we strictly need more slots (simplified fallback)
        if len(available_slots) < slots_info['total']:
            additional_needed = slots_info['total'] - len(available_slots)
            additional_slots = find_available_slots_phase7(
                course.semester, label_list[0], target_period, occupied_slots,
                additional_needed * 3, slot_durations
            )
            for slot in additional_slots:
                valid = True
                for lbl in label_list[1:]:
                    sec_key = f"{lbl}_{target_period}"
                    for existing in occupied_slots.get(sec_key, []):
                        eb = existing[0] if isinstance(existing, tuple) else existing
                        if eb and slot.overlaps(eb):
                            valid = False; break
                    if not valid: break
                if valid and slot not in available_slots:
                    available_slots.append(slot)
                    
        used_days_by_course = SessionRulesValidator.get_used_days_tracker()
        used_days_lectures = set()
        used_days_tutorials = set()
        
        # Schedule lectures
        slot_idx = 0
        lecture_count = 0
        max_retries = 3
        
        for i in range(slots_info['lectures']):
            if lecture_count >= slots_info['lectures']: break
            found_slot = False
            retry_count = 0
            while not found_slot and retry_count < max_retries:
                if slot_idx >= len(available_slots):
                    break
                while slot_idx < len(available_slots):
                    slot = available_slots[slot_idx]
                    if slot.day in used_days_lectures or not SessionRulesValidator.can_schedule_session_type(course.code, slot.day, "L", used_days_by_course):
                        slot_idx += 1; continue
                    room = assign_classroom(total_students, classrooms, slot)
                    for lbl in label_list:
                        session = ScheduledSession(
                            course_code=course.code, section=lbl, kind="L", block=slot,
                            room=room.room_number, period=target_period,
                            faculty=course.instructors[0] if course.instructors else 'TBD'
                        )
                        sessions.append(session)
                        sec_key = f"{lbl}_{target_period}"
                        if sec_key not in occupied_slots: occupied_slots[sec_key] = []
                        occupied_slots[sec_key].append((slot, course.code))
                    used_days_lectures.add(slot.day)
                    SessionRulesValidator.mark_day_used(course.code, slot.day, "L", used_days_by_course)
                    slot_idx += 1
                    lecture_count += 1
                    found_slot = True
                    break
                if not found_slot: retry_count += 1
            
            if not found_slot:
                # Emergency slots
                emergency_slots = find_available_slots_phase7(course.semester, label_list[0], target_period, occupied_slots, 1, [90])
                for slot in emergency_slots:
                    if not validate_time_range(slot.start, slot.end): continue
                    if slot.day not in used_days_lectures:
                        room = assign_classroom(total_students, classrooms, slot)
                        for lbl in label_list:
                            session = ScheduledSession(
                                course_code=course.code, section=lbl, kind="L", block=slot,
                                room=room.room_number, period=target_period,
                                faculty=course.instructors[0] if course.instructors else 'TBD'
                            )
                            sessions.append(session)
                            sec_key = f"{lbl}_{target_period}"
                            if sec_key not in occupied_slots: occupied_slots[sec_key] = []
                            occupied_slots[sec_key].append((slot, course.code))
                        used_days_lectures.add(slot.day)
                        SessionRulesValidator.mark_day_used(course.code, slot.day, "L", used_days_by_course)
                        lecture_count += 1
                        found_slot = True
                        break

        # Schedule tutorials
        tutorial_count = 0
        for i in range(slots_info['tutorials']):
            if tutorial_count >= slots_info['tutorials']: break
            found_slot = False
            retry_count = 0
            while not found_slot and retry_count < max_retries:
                if slot_idx >= len(available_slots):
                    break
                while slot_idx < len(available_slots):
                    slot = available_slots[slot_idx]
                    if slot.day in used_days_tutorials or slot.day in used_days_lectures or not SessionRulesValidator.can_schedule_session_type(course.code, slot.day, "T", used_days_by_course):
                        slot_idx += 1; continue
                    if not validate_time_range(slot.start, slot.end):
                        slot_idx += 1; continue
                    room = assign_classroom(total_students, classrooms, slot)
                    for lbl in label_list:
                        session = ScheduledSession(
                            course_code=course.code, section=lbl, kind="T", block=slot,
                            room=room.room_number, period=target_period,
                            faculty=course.instructors[0] if course.instructors else 'TBD'
                        )
                        sessions.append(session)
                        sec_key = f"{lbl}_{target_period}"
                        if sec_key not in occupied_slots: occupied_slots[sec_key] = []
                        occupied_slots[sec_key].append((slot, course.code))
                    used_days_tutorials.add(slot.day)
                    SessionRulesValidator.mark_day_used(course.code, slot.day, "T", used_days_by_course)
                    tutorial_count += 1
                    slot_idx += 1
                    found_slot = True
                    break
                if not found_slot: retry_count += 1
            if not found_slot:
                emergency_slots = find_available_slots_phase7(course.semester, label_list[0], target_period, occupied_slots, 1, [60])
                for slot in emergency_slots:
                    if slot.day not in used_days_lectures and slot.day not in used_days_tutorials:
                        room = assign_classroom(total_students, classrooms, slot)
                        for lbl in label_list:
                            session = ScheduledSession(
                                course_code=course.code, section=lbl, kind="T", block=slot,
                                room=room.room_number, period=target_period,
                                faculty=course.instructors[0] if course.instructors else 'TBD'
                            )
                            sessions.append(session)
                            sec_key = f"{lbl}_{target_period}"
                            if sec_key not in occupied_slots: occupied_slots[sec_key] = []
                            occupied_slots[sec_key].append((slot, course.code))
                        used_days_tutorials.add(slot.day)
                        SessionRulesValidator.mark_day_used(course.code, slot.day, "T", used_days_by_course)
                        tutorial_count += 1
                        found_slot = True
                        break

        # Schedule practicals
        for i in range(slots_info['practicals']):
            if slot_idx >= len(available_slots): break
            slot = available_slots[slot_idx]
            if not validate_time_range(slot.start, slot.end):
                slot_idx += 1; continue
            slot_duration = (slot.end.hour * 60 + slot.end.minute) - (slot.start.hour * 60 + slot.start.minute)
            if slot_duration >= 120 and (slot.start.hour > 16 or (slot.start.hour == 16 and slot.start.minute > 0)):
                slot_idx += 1; continue
            lab_rooms = [r for r in classrooms if hasattr(r, 'room_type') and r.room_type.lower() == 'lab']
            if not lab_rooms:
                lab_rooms = classrooms
            assigned_room = None
            for lab in lab_rooms:
                room_num = lab.room_number if hasattr(lab, 'room_number') else str(lab)
                if room_occupancy and room_num in room_occupancy:
                    has_conflict = False
                    for occupied_block in room_occupancy[room_num]:
                        if slot.overlaps(occupied_block): has_conflict = True; break
                    if has_conflict: continue
                assigned_room = room_num
                break
            if not assigned_room and lab_rooms:
                assigned_room = lab_rooms[0].room_number if hasattr(lab_rooms[0], 'room_number') else str(lab_rooms[0])
            elif not assigned_room:
                assigned_room = classrooms[0].room_number if classrooms and hasattr(classrooms[0], 'room_number') else 'LAB1'
            
            for lbl in label_list:
                session = ScheduledSession(
                    course_code=course.code, section=lbl, kind="P", block=slot,
                    room=assigned_room, period=target_period, faculty=None
                )
                sessions.append(session)
                sec_key = f"{lbl}_{target_period}"
                if sec_key not in occupied_slots: occupied_slots[sec_key] = []
                occupied_slots[sec_key].append((slot, course.code))
                
            if room_occupancy is not None:
                if assigned_room not in room_occupancy: room_occupancy[assigned_room] = []
                room_occupancy[assigned_room].append(slot)
            slot_idx += 1
            
    # Group sections by period
    period_groups = {'PRE': [], 'POST': []}
    for section in dept_sections:
        section_label = f"{section.program}-{section.name}-Sem{section.semester}"
        key = (course.code.upper(), course.semester, section_label)
        is_premid = period_assignments.get(key, True)
        period = 'PRE' if is_premid else 'POST'
        period_groups[period].append((section, section_label))

    for period, section_list in period_groups.items():
        if not section_list: continue
        
        # Check if we should combine: Multiple sections mapping to same period AND single instructor
        if len(section_list) > 1 and course.num_faculty == 1:
            labels = [s[1] for s in section_list]
            total_students = sum(s[0].students for s in section_list)
            schedule_labels(labels, period, total_students)
        else:
            for section, section_label in section_list:
                schedule_labels([section_label], period, section.students)
                
    return sessions

def schedule_non_combined(course: Course, sections: List[Section],
                         occupied_slots: Dict,
                         classrooms: List[ClassRoom],
                         room_occupancy: Dict,
                         period_assignments: Dict[Tuple[str, int, str], bool]) -> List[ScheduledSession]:
    """
    Schedule non-combined section-specific courses
    Schedule in EITHER PreMid OR PostMid (not both) - half-semester courses
    """
    import time
    start_time = time.time()
    timeout = 15  # 15 second timeout per course (increased for better scheduling)
    
    sessions = []
    
    # Calculate slots from LTPSC
    slots_info = calculate_slots_from_ltpsc(course.ltpsc)
    
    # Prepare slot durations
    slot_durations = []
    for _ in range(slots_info['lectures']):
        slot_durations.append(90)  # 1.5 hours
    for _ in range(slots_info['tutorials']):
        slot_durations.append(60)  # 1 hour
    for _ in range(slots_info['practicals']):
        slot_durations.append(120)  # 2 hours
    
    # Find the section for this course (Phase 7 uses A-section for single-section depts)
    section_label = f"{course.department}-A-Sem{course.semester}"
    
    # User-driven period selection (default to PRE if missing for robustness)
    key = (course.code.upper(), course.semester, section_label)
    is_premid = period_assignments.get(key, True)
    preferred_period = 'PRE' if is_premid else 'POST'
    # Try preferred period first, then the other as fallback
    periods_to_try = [preferred_period, 'POST' if preferred_period == 'PRE' else 'PRE']
    
    available_slots = []
    selected_period = None
    
    # Try both periods
    for period in periods_to_try:
        available_slots = find_available_slots_phase7(
            course.semester, section_label, period, occupied_slots,
            slots_info['total'], slot_durations
        )
        
        if len(available_slots) >= slots_info['total']:
            selected_period = period
            break
    
    if len(available_slots) < slots_info['total']:
        # Try harder - search more thoroughly
        # Try to find slots one by one if initial search didn't find enough
        for period_retry in periods_to_try:
            additional_needed = slots_info['total'] - len(available_slots)
            if additional_needed > 0:
                additional_slots = find_available_slots_phase7(
                    course.semester, section_label, period_retry, occupied_slots,
                    additional_needed, slot_durations
                )
                if additional_slots:
                    available_slots.extend(additional_slots)
                    selected_period = period_retry
                    break
        
        # If still not enough, return what we have (better than nothing)
        if len(available_slots) < slots_info['total']:
            return sessions
    
    period = selected_period
    
    # Check timeout
    if time.time() - start_time > timeout:
        return sessions
    
    # Track used days for one-session-per-day rule
    used_days_by_course = SessionRulesValidator.get_used_days_tracker()
    used_days_lectures = set()
    used_days_tutorials = set()
    
    # Schedule lectures
    slot_idx = 0
    lecture_count = 0
    max_retries = 3  # Try up to 3 times to find more slots
    
    for i in range(slots_info['lectures']):
        # Check timeout
        if time.time() - start_time > timeout:
            break
            
        if lecture_count >= slots_info['lectures']:
            break
        
        # Try to find a valid slot
        found_slot = False
        retry_count = 0
        
        while not found_slot and retry_count < max_retries:
            # If we've exhausted available slots, try to find more
            if slot_idx >= len(available_slots):
                if retry_count < max_retries - 1:
                    # Try to find more slots
                    additional_slots = find_available_slots_phase7(
                        course.semester, section_label, period, occupied_slots,
                        slots_info['lectures'] - lecture_count, [90]  # Just need lecture slots
                    )
                    if additional_slots:
                        available_slots.extend(additional_slots)
                        slot_idx = len(available_slots) - len(additional_slots)
                    retry_count += 1
                else:
                    break
            
            # Find next available slot that doesn't violate one-session-per-day rule
            max_slot_checks = len(available_slots) * 2  # Safety limit
            slot_check_count = 0
            
            while slot_idx < len(available_slots) and slot_check_count < max_slot_checks:
                slot_check_count += 1
                
                # Check timeout inside loop
                if time.time() - start_time > timeout:
                    break
                    
                slot = available_slots[slot_idx]
                
                # Check one-session-per-day rule
                if slot.day in used_days_lectures:
                    slot_idx += 1
                    continue
                
                if not SessionRulesValidator.can_schedule_session_type(
                    course.code, slot.day, "L", used_days_by_course
                ):
                    slot_idx += 1
                    continue
                
                # Found valid slot
                room = assign_classroom(85, classrooms, slot)
                session = ScheduledSession(
                    course_code=course.code,
                    section=section_label,
                    kind="L",
                    block=slot,
                    room=room.room_number,
                    period=period,
                    faculty=course.instructors[0] if course.instructors else 'TBD'
                )
                sessions.append(session)
                
                # Update occupied slots
                section_key = f"{section_label}_{period}"
                if section_key not in occupied_slots:
                    occupied_slots[section_key] = []
                occupied_slots[section_key].append((slot, course.code))
                
                # Mark day as used
                used_days_lectures.add(slot.day)
                SessionRulesValidator.mark_day_used(course.code, slot.day, "L", used_days_by_course)
                
                slot_idx += 1
                lecture_count += 1
                found_slot = True
                break
            
            if not found_slot:
                retry_count += 1
        
        if not found_slot:
            # Last resort: try to find ANY available slot
            emergency_slots = find_available_slots_phase7(
                course.semester, section_label, period, occupied_slots,
                slots_info['lectures'] - lecture_count, [90]
            )
            for slot in emergency_slots:
                if slot.day not in used_days_lectures:
                    room = assign_classroom(85, classrooms, slot)
                    session = ScheduledSession(
                        course_code=course.code,
                        section=section_label,
                        kind="L",
                        block=slot,
                        room=room.room_number,
                        period=period,
                        faculty=course.instructors[0] if course.instructors else 'TBD'
                    )
                    sessions.append(session)
                    section_key = f"{section_label}_{period}"
                    if section_key not in occupied_slots:
                        occupied_slots[section_key] = []
                    occupied_slots[section_key].append((slot, course.code))
                    used_days_lectures.add(slot.day)
                    SessionRulesValidator.mark_day_used(course.code, slot.day, "L", used_days_by_course)
                    lecture_count += 1
                    # Continue to schedule all remaining lectures
                    if lecture_count >= slots_info['lectures']:
                        break
    
    # Schedule tutorials
    tutorial_count = 0
    max_retries = 3  # Try up to 3 times to find more slots
    
    for i in range(slots_info['tutorials']):
        if tutorial_count >= slots_info['tutorials']:
            break
            
        # Check timeout
        if time.time() - start_time > timeout:
            break
        
        # Try to find a valid slot
        found_slot = False
        retry_count = 0
        
        while not found_slot and retry_count < max_retries:
            # If we've exhausted available slots, try to find more
            if slot_idx >= len(available_slots):
                if retry_count < max_retries - 1:
                    # Try to find more slots
                    additional_slots = find_available_slots_phase7(
                        course.semester, section_label, period, occupied_slots,
                        slots_info['tutorials'] - tutorial_count, [60]  # Just need tutorial slots
                    )
                    if additional_slots:
                        available_slots.extend(additional_slots)
                        slot_idx = len(available_slots) - len(additional_slots)
                    retry_count += 1
                else:
                    break
            
            # Find next available slot that doesn't violate one-session-per-day rule
            max_slot_checks = len(available_slots) * 2  # Safety limit
            slot_check_count = 0
            
            while slot_idx < len(available_slots) and slot_check_count < max_slot_checks:
                slot_check_count += 1
                
                # Check timeout inside loop
                if time.time() - start_time > timeout:
                    break
                    
                slot = available_slots[slot_idx]
                
                # Check one-session-per-day rule - cannot be on same day as lecture
                if slot.day in used_days_tutorials or slot.day in used_days_lectures:
                    slot_idx += 1
                    continue
                
                if not SessionRulesValidator.can_schedule_session_type(
                    course.code, slot.day, "T", used_days_by_course
                ):
                    slot_idx += 1
                    continue
                
                # Found valid slot
                room = assign_classroom(85, classrooms, slot)
                session = ScheduledSession(
                    course_code=course.code,
                    section=section_label,
                    kind="T",
                    block=slot,
                    room=room.room_number,
                    period=period,
                    faculty=course.instructors[0] if course.instructors else 'TBD'
                )
                sessions.append(session)
                
                # Update occupied slots
                section_key = f"{section_label}_{period}"
                if section_key not in occupied_slots:
                    occupied_slots[section_key] = []
                occupied_slots[section_key].append((slot, course.code))
                
                # Mark day as used
                used_days_tutorials.add(slot.day)
                SessionRulesValidator.mark_day_used(course.code, slot.day, "T", used_days_by_course)
                tutorial_count += 1
                slot_idx += 1
                found_slot = True
                break
            
            if not found_slot:
                retry_count += 1
        
        if not found_slot:
            # Last resort: try to find ANY available slot on a different day
            emergency_slots = find_available_slots_phase7(
                course.semester, section_label, period, occupied_slots,
                slots_info['tutorials'] - tutorial_count, [60]
            )
            for slot in emergency_slots:
                if slot.day not in used_days_lectures and slot.day not in used_days_tutorials:
                    room = assign_classroom(85, classrooms, slot)
                    session = ScheduledSession(
                        course_code=course.code,
                        section=section_label,
                        kind="T",
                        block=slot,
                        room=room.room_number,
                        period=period,
                        faculty=course.instructors[0] if course.instructors else 'TBD'
                    )
                    sessions.append(session)
                    section_key = f"{section_label}_{period}"
                    if section_key not in occupied_slots:
                        occupied_slots[section_key] = []
                    occupied_slots[section_key].append((slot, course.code))
                    used_days_tutorials.add(slot.day)
                    SessionRulesValidator.mark_day_used(course.code, slot.day, "T", used_days_by_course)
                    tutorial_count += 1
                    # Continue to schedule all remaining tutorials
                    if tutorial_count >= slots_info['tutorials']:
                        break
    
    # Check timeout before practicals
    if time.time() - start_time > timeout:
        return sessions
    
    # Schedule practicals (practicals can be on same day as lectures/tutorials)
    for i in range(slots_info['practicals']):
        if time.time() - start_time > timeout:
            break
            
        # If we've exhausted available slots, try to find more
        if slot_idx >= len(available_slots):
            additional_slots = find_available_slots_phase7(
                course.semester, section_label, period, occupied_slots,
                slots_info['practicals'] - i, [120]  # Just need practical slots
            )
            if additional_slots:
                available_slots.extend(additional_slots)
                slot_idx = len(available_slots) - len(additional_slots)
            else:
                break
            
        slot = available_slots[slot_idx]
        
        # CRITICAL: Validate time range (9:00-18:00)
        # For 2-hour practicals, ensure start time allows completion before 18:00
        # Practical starting at 17:00 would end at 19:00, which is invalid
        if not validate_time_range(slot.start, slot.end):
            slot_idx += 1
            continue  # Skip this slot - extends beyond 18:00 or starts before 9:00
        
        # Additional check: For 2-hour practicals, start must be <= 16:00
        # Check if this is a practical slot (2 hours duration)
        slot_duration = (slot.end.hour * 60 + slot.end.minute) - (slot.start.hour * 60 + slot.start.minute)
        if slot_duration >= 120 and (slot.start.hour > 16 or (slot.start.hour == 16 and slot.start.minute > 0)):
            slot_idx += 1
            continue  # Skip - practical would extend beyond 18:00
        
        # For practicals, use labs - find available lab room
        lab_rooms = [r for r in classrooms if hasattr(r, 'room_type') and r.room_type.lower() == 'lab']
        if not lab_rooms:
            # Fallback to any classroom if no labs found
            lab_rooms = classrooms
        
        # Find first available lab room (check room_occupancy if available)
        assigned_room = None
        for lab in lab_rooms:
            room_num = lab.room_number if hasattr(lab, 'room_number') else str(lab)
            # Check if room is available (if room_occupancy tracking exists)
            if room_occupancy and room_num in room_occupancy:
                # Check for conflicts
                has_conflict = False
                for occupied_block in room_occupancy[room_num]:
                    if slot.overlaps(occupied_block):
                        has_conflict = True
                        break
                if has_conflict:
                    continue
            assigned_room = room_num
            break
        
        # Fallback to first lab if no available one found
        if not assigned_room and lab_rooms:
            assigned_room = lab_rooms[0].room_number if hasattr(lab_rooms[0], 'room_number') else str(lab_rooms[0])
        elif not assigned_room:
            assigned_room = classrooms[0].room_number if classrooms and hasattr(classrooms[0], 'room_number') else 'LAB1'
        
        session = ScheduledSession(
            course_code=course.code,
            section=section_label,
            kind="P",
            block=slot,
            room=assigned_room,
            period=period,
            faculty=None  # No faculty for labs
        )
        sessions.append(session)
        
        # Update occupied slots
        section_key = f"{section_label}_{period}"
        if section_key not in occupied_slots:
            occupied_slots[section_key] = []
        occupied_slots[section_key].append((slot, course.code))
        
        # Update room occupancy if tracking exists
        if room_occupancy is not None:
            if assigned_room not in room_occupancy:
                room_occupancy[assigned_room] = []
            room_occupancy[assigned_room].append(slot)
        
        slot_idx += 1
    
    return sessions

def run_phase7(courses: List[Course], sections: List[Section],
              classrooms: List[ClassRoom],
              occupied_slots: Dict,
              room_occupancy: Dict,
              combined_sessions: List = None,
              timeout_seconds: int = 60) -> List[ScheduledSession]:
    """Main Phase 7 execution with timeout protection"""
    
    import time
    start_time = time.time()
    
    print("=== PHASE 7: REMAINING <=2 CREDIT COURSES ===")
    print()
    
    # Identify courses
    within_group, non_combined = identify_phase7_courses(courses, sections, combined_sessions)
    
    # Check timeout before starting
    if time.time() - start_time > timeout_seconds:
        print(f"WARNING: Phase 7 timeout ({timeout_seconds}s) - skipping")
        return []
    
    print(f"\nWithin-group combined courses: {len(within_group)}")
    for group in within_group:
        course = group[0]
        slots_info = calculate_slots_from_ltpsc(course.ltpsc)
        print(f"  {course.code} - {course.name} ({course.department}, Sem{course.semester})")
        print(f"    LTPSC: {course.ltpsc} -> {slots_info['lectures']}L + {slots_info['tutorials']}T + {slots_info['practicals']}P")
    
    print(f"\nNon-combined courses: {len(non_combined)}")
    for course in non_combined:
        slots_info = calculate_slots_from_ltpsc(course.ltpsc)
        print(f"  {course.code} - {course.name} ({course.department}, Sem{course.semester})")
        print(f"    LTPSC: {course.ltpsc} -> {slots_info['lectures']}L + {slots_info['tutorials']}T + {slots_info['practicals']}P")
    
    # Build or load user-driven period assignments for Phase 7
    assignments_path = get_phase7_period_assignments_path()
    period_assignments, any_prompted = get_phase7_period_assignments(
        within_group,
        non_combined,
        sections,
        assignments_path,
    )

    all_sessions = []
    
    # Schedule within-group combined
    print("\nScheduling within-group combined courses...")
    total_within = len(within_group)
    for idx, course_group in enumerate(within_group, 1):
        # Check timeout
        if time.time() - start_time > timeout_seconds:
            print(f"\nWARNING: Phase 7 timeout ({timeout_seconds}s) after {idx-1}/{total_within} within-group courses")
            break
            
        course = course_group[0]
        print(f"  [{idx}/{total_within}] Scheduling {course.code}...", end=" ", flush=True)
        sessions = schedule_within_group_combined(
            course,
            sections,
            occupied_slots,
            classrooms,
            room_occupancy,
            period_assignments,
        )
        all_sessions.extend(sessions)
        print(f"[OK] {len(sessions)} sessions scheduled")
    
    # Schedule non-combined
    print("\nScheduling non-combined courses...")
    total_non = len(non_combined)
    for idx, course in enumerate(non_combined, 1):
        # Check timeout
        if time.time() - start_time > timeout_seconds:
            print(f"\nWARNING: Phase 7 timeout ({timeout_seconds}s) after {idx-1}/{total_non} non-combined courses")
            break
            
        print(f"  [{idx}/{total_non}] Scheduling {course.code}...", end=" ", flush=True)
        sessions = schedule_non_combined(
            course,
            sections,
            occupied_slots,
            classrooms,
            room_occupancy,
            period_assignments,
        )
        all_sessions.extend(sessions)
        print(f"[OK] {len(sessions)} sessions scheduled")
    
    print(f"\nPhase 7 completed: {len(all_sessions)} sessions scheduled")
    
    # Persist any newly entered period assignments
    if any_prompted:
        save_phase7_period_assignments(assignments_path, period_assignments)

    # Log time slots
    logger = get_logger()
    for session in all_sessions:
        logger.log_session("Phase 7", session)
    
    # Print time slot logging summary
    print("\nTime slot logging summary...")
    phase7_entries = logger.get_entries_by_phase("Phase 7")
    if phase7_entries:
        phase7_summary = logger.get_phase_summary("Phase 7")
        print(f"Phase 7 logged {phase7_summary['total_slots']} time slots")
        print(f"  - Unique courses: {phase7_summary['unique_courses']}")
        print(f"  - Unique sections: {phase7_summary['unique_sections']}")
        print(f"  - By day: {phase7_summary['by_day']}")
        print(f"  - By session type: {phase7_summary['by_session_type']}")
    
    return all_sessions

if __name__ == "__main__":
    # Test the implementation
    from modules_v2.phase1_data_validation_v2 import run_phase1
    from modules_v2.phase3_elective_baskets_v2 import run_phase3
    from modules_v2.phase4_combined_classes_v2_corrected import run_phase4_corrected as run_phase4
    from modules_v2.phase5_core_courses import run_phase5
    
    print("Testing Phase 7 implementation...")
    courses, classrooms, statistics = run_phase1()
    
    # Extract unique semesters from course data
    unique_semesters = sorted(set(course.semester for course in courses 
                                 if course.department in ['CSE', 'DSAI', 'ECE']))
    
    # Create sections
    sections = []
    from config.structure_config import DEPARTMENTS, SECTIONS_BY_DEPT, STUDENTS_PER_SECTION, get_group_for_section
    for dept in DEPARTMENTS:
        for sem in unique_semesters:
            for sec_label in SECTIONS_BY_DEPT.get(dept, []):
                group = get_group_for_section(dept, sec_label)
                sections.append(Section(dept, group, sec_label, sem, STUDENTS_PER_SECTION))
    
    # Run previous phases to build occupied_slots
    elective_baskets, elective_sessions = run_phase3(courses, sections)
    phase4_result = run_phase4(courses, sections)
    
    # Create occupied_slots from elective sessions
    occupied_slots = {}
    for session in elective_sessions:
        # elective_sessions are ScheduledSession objects
        section_key = f"{session.section}_{session.period}"
        if section_key not in occupied_slots:
            occupied_slots[section_key] = []
        occupied_slots[section_key].append((session.block, session.course_code))
    
    # Run Phase 7
    phase7_sessions = run_phase7(courses, sections, classrooms, occupied_slots, {})
    
    print("\n=== PHASE 7 TEST COMPLETE ===")
    print(f"Total Phase 7 sessions: {len(phase7_sessions)}")

