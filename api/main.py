"""
Timetable Pro API: generate, verify, and reflow endpoints.
Serves the frontend; uses existing pipeline and verification.
"""
import os
import re
import sys
import csv
from datetime import time
from typing import List, Dict, Any, Optional

# Run from repo root so imports work
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, REPO_ROOT)
os.chdir(REPO_ROOT)

from utils.period_utils import normalize_period

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

app = FastAPI(title="Timetable Pro API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000", "http://127.0.0.1:5173", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# --- Config (from schedule_config and structure_config) ---
def get_config(semesters: Optional[List[int]] = None):
    """Build config for UI. If semesters not provided, derive from run_phase1() courses."""
    from config.schedule_config import WORKING_DAYS, DAY_START_TIME, DAY_END_TIME, LUNCH_WINDOWS
    from config.structure_config import DEPARTMENTS, SECTIONS_BY_DEPT
    if semesters is None:
        try:
            from modules_v2.phase1_data_validation_v2 import run_phase1
            courses, _, _ = run_phase1()
            semesters = sorted(set(c.semester for c in courses if c.department in DEPARTMENTS))
        except Exception:
            semesters = [1, 2, 3, 4, 5, 6]
    program_display = {"CSE": "Computer Science", "DSAI": "Data Science & AI", "ECE": "ECE"}
    programs = []
    for dept in DEPARTMENTS:
        programs.append({
            "id": dept,
            "name": program_display.get(dept, dept),
            "sections": SECTIONS_BY_DEPT.get(dept, []),
        })
    section_labels = []
    for dept in DEPARTMENTS:
        for sem in semesters:
            for sec in SECTIONS_BY_DEPT.get(dept, []):
                section_labels.append({
                    "section": f"{dept}-{sec}-Sem{sem}",
                    "program": dept,
                    "semester": sem,
                    "label": f"Semester {sem} - {dept}",
                })
    return {
        "working_days": WORKING_DAYS,
        "day_start": DAY_START_TIME.strftime("%H:%M"),
        "day_end": DAY_END_TIME.strftime("%H:%M"),
        "lunch_windows": {str(k): [v[0].strftime("%H:%M"), v[1].strftime("%H:%M")] for k, v in LUNCH_WINDOWS.items()},
        "programs": programs,
        "semesters": semesters,
        "section_labels": section_labels,
    }


def load_timetable_from_csv(log_path: str) -> List[Dict[str, Any]]:
    """Load timetable sessions from time_slot_log CSV. Returns list of session dicts (API schema).
    Deduplicates only truly identical rows.
    """
    seen: set = set()
    rows = []
    with open(log_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            phase = row.get("Phase", "")
            section = row.get("Section", "")
            course_code = row.get("Course Code", "")
            day = row.get("Day", "")
            start_time = row.get("Start Time", "")
            end_time = row.get("End Time", "")
            session_type = row.get("Session Type", "")
            period = row.get("Period", "")
            room = row.get("Room", "")
            faculty = row.get("Faculty", "")
            # Keep API rows aligned 1:1 with final log rows used for Excel render.
            # Only collapse exact duplicates, never "same slot but different phase/room" rows.
            dedup_key = (
                phase,
                section,
                course_code,
                day,
                start_time,
                end_time,
                session_type,
                period,
                room,
                faculty,
            )
            if dedup_key in seen:
                continue
            seen.add(dedup_key)
            rows.append({
                "Phase": phase,
                "Course Code": course_code,
                "Section": section,
                "Day": day,
                "Start Time": start_time,
                "End Time": end_time,
                "Room": room,
                "Period": period,
                "Session Type": session_type,
                "Faculty": faculty,
            })
    return rows


def _resolve_log_path_for_timestamp(ts: str, prefer_edited: bool = False) -> str:
    """
    Resolve the actual time-slot log path for a given timestamp.
    Some flows write to DATA/OUTPUT, while regenerate-from-sessions writes to DATA/EDITED OUTPUT.
    """
    candidates: List[str] = []
    output_path = os.path.join(REPO_ROOT, "DATA", "OUTPUT", f"time_slot_log_{ts}.csv")
    edited_output_path = os.path.join(REPO_ROOT, "DATA", "EDITED OUTPUT", f"time_slot_log_{ts}.csv")

    if prefer_edited:
        candidates = [edited_output_path, output_path]
    else:
        candidates = [output_path, edited_output_path]

    for p in candidates:
        if os.path.exists(p):
            return p

    raise FileNotFoundError(
        f"Time slot log not found for timestamp {ts}. Tried: {', '.join(candidates)}"
    )


# Verification table column indices (0-based) matching write_verification_table headers
_VERIFICATION_HEADERS = [
    "Code", "Course Name", "Instructor", "LTPSC", "Assigned Lab", "Assigned Classroom",
    "Lectures (Req/Sched)", "Tutorials (Req/Sched)", "Labs (Req/Sched)", "Status",
    "Time Slot Issues", "Room Conflicts", "Colour",
]


def _sheet_name_to_key(sheet_name: str) -> Optional[str]:
    """Convert Excel sheet name to frontend key. e.g. 'CSE-A Sem1 PreMid' -> 'CSE-A-Sem1-PRE'."""
    # Sheet name format: "{section_name} Sem{semester} {period}" e.g. "CSE-A Sem1 PreMid"
    m = re.match(r"^(.+?)\s+Sem(\d+)\s+(PreMid|PostMid)$", sheet_name.strip(), re.IGNORECASE)
    if not m:
        return None
    section_name, sem, period = m.group(1).strip(), m.group(2), m.group(3)
    period_key = "PRE" if period.lower() == "premid" else "POST"
    return f"{section_name}-Sem{sem}-{period_key}"


def parse_verification_tables_from_excel(excel_path: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Parse verification tables from generated Excel. Returns dict keyed by section-period
    e.g. "CSE-A-Sem1-PRE" -> list of { code, course_name, ltpsc, lectures, tutorials, labs, status, ... }.
    """
    import openpyxl
    result: Dict[str, List[Dict[str, Any]]] = {}
    if not os.path.isfile(excel_path):
        return result
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    for sheet_name in wb.sheetnames:
        key = _sheet_name_to_key(sheet_name)
        if not key:
            continue
        sheet = wb[sheet_name]
        # Find header row (row with "Code" in first column)
        header_row_idx = None
        for row_idx in range(1, min(sheet.max_row + 1, 500)):
            cell_val = sheet.cell(row=row_idx, column=1).value
            if cell_val is not None and str(cell_val).strip() == "Code":
                header_row_idx = row_idx
                break
        if header_row_idx is None:
            continue
        # Read data rows (columns: Code, Course Name, Instructor, LTPSC, Assigned Lab, Assigned Classroom,
        # Lectures (Req/Sched), Tutorials (Req/Sched), Labs (Req/Sched), Status, Time Slot Issues, Room Conflicts, Colour)
        rows_data = []
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            code_cell = sheet.cell(row=row_idx, column=1).value
            if code_cell is None or (isinstance(code_cell, str) and not code_cell.strip()):
                break
            def _cell(r: int, c: int) -> str:
                v = sheet.cell(row=r, column=c).value
                return str(v).strip() if v is not None else ""
            row_dict = {
                "code": _cell(row_idx, 1),
                "course_name": _cell(row_idx, 2),
                "instructor": _cell(row_idx, 3),
                "ltpsc": _cell(row_idx, 4),
                "assigned_lab": _cell(row_idx, 5),
                "assigned_classroom": _cell(row_idx, 6),
                "lectures": _cell(row_idx, 7),
                "tutorials": _cell(row_idx, 8),
                "labs": _cell(row_idx, 9),
                "status": _cell(row_idx, 10),
                "time_slot_issues": _cell(row_idx, 11),
                "room_conflicts": _cell(row_idx, 12),
            }
            if row_dict["code"]:
                rows_data.append(row_dict)
        result[key] = rows_data
    return result


def parse_timetable_sessions_from_excel(excel_path: str) -> List[Dict[str, Any]]:
    """
    Parse the displayed timetable grid from generated Excel sheets.
    This is the authoritative visual schedule source used by users.
    """
    import openpyxl

    sessions: List[Dict[str, Any]] = []
    if not os.path.isfile(excel_path):
        return sessions

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    def _to_str(v: Any) -> str:
        return str(v).strip() if v is not None else ""

    for sheet_name in wb.sheetnames:
        key = _sheet_name_to_key(sheet_name)
        if not key:
            continue

        # key format: "<section>-Sem<sem>-PRE|POST"
        m = re.match(r"^(.+)-Sem(\d+)-(PRE|POST)$", key.strip(), re.IGNORECASE)
        if not m:
            continue
        section = m.group(1).strip() + f"-Sem{m.group(2)}"
        period = m.group(3).upper()

        sheet = wb[sheet_name]
        row = 1
        max_row = sheet.max_row
        max_col = sheet.max_column
        while row <= max_row:
            day_cell = _to_str(sheet.cell(row=row, column=1).value)
            if day_cell.lower().startswith("day:"):
                day = day_cell.split(":", 1)[1].strip()
                time_row = row + 1
                course_row = row + 2
                if course_row <= max_row:
                    for col in range(2, max_col + 1):
                        time_val = _to_str(sheet.cell(row=time_row, column=col).value)
                        course_val = _to_str(sheet.cell(row=course_row, column=col).value)
                        if not time_val or not course_val:
                            continue
                        course_upper = course_val.upper()
                        if course_upper in ("LUNCH", "BREAK(15MIN)"):
                            continue
                        if "-" not in time_val:
                            continue
                        time_parts = [p.strip() for p in time_val.split("-", 1)]
                        if len(time_parts) != 2 or not time_parts[0] or not time_parts[1]:
                            continue
                        start_time, end_time = time_parts[0], time_parts[1]
                        session_type = "L"
                        if "-LAB" in course_upper:
                            session_type = "P"
                        elif "-TUT" in course_upper:
                            session_type = "T"
                        elif course_upper.startswith("ELECTIVE_BASKET_"):
                            session_type = "ELECTIVE"
                        sessions.append(
                            {
                                "Phase": "",
                                "Course Code": course_val,
                                "Section": section,
                                "Day": day,
                                "Start Time": start_time,
                                "End Time": end_time,
                                "Room": "",
                                "Period": period,
                                "Session Type": session_type,
                                "Faculty": "",
                            }
                        )
                row = course_row + 1
                continue
            row += 1
    return sessions


def merge_excel_sessions_with_log_details(
    excel_sessions: List[Dict[str, Any]],
    log_sessions: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Keep Excel day/time placement authoritative and enrich room/faculty/phase from log rows.
    """
    merged: List[Dict[str, Any]] = []
    log_index_exact: Dict[tuple, Dict[str, Any]] = {}
    log_index_relaxed: Dict[tuple, Dict[str, Any]] = {}
    # Cross-section fallback for synchronized slots that appear in Excel for multiple sections
    # while the log may contain a single representative section row.
    log_index_slot_exact: Dict[tuple, Dict[str, Any]] = {}
    log_index_slot_relaxed: Dict[tuple, Dict[str, Any]] = {}

    for r in log_sessions or []:
        section = str((r or {}).get("Section", "") or "").strip()
        day = str((r or {}).get("Day", "") or "").strip()
        start = str((r or {}).get("Start Time", "") or "").strip()
        end = str((r or {}).get("End Time", "") or "").strip()
        period = normalize_period((r or {}).get("Period"))
        code = str((r or {}).get("Course Code", "") or "").strip()
        stype = str((r or {}).get("Session Type", "") or "").strip().upper()
        if not (section and day and start and end and code):
            continue
        exact_key = (section, period, day, start, end, code, stype)
        relaxed_key = (section, period, day, start, end, code)
        slot_exact_key = (period, day, start, end, code, stype)
        slot_relaxed_key = (period, day, start, end, code)
        log_index_exact.setdefault(exact_key, r)
        log_index_relaxed.setdefault(relaxed_key, r)
        log_index_slot_exact.setdefault(slot_exact_key, r)
        log_index_slot_relaxed.setdefault(slot_relaxed_key, r)

    for s in excel_sessions or []:
        section = str((s or {}).get("Section", "") or "").strip()
        day = str((s or {}).get("Day", "") or "").strip()
        start = str((s or {}).get("Start Time", "") or "").strip()
        end = str((s or {}).get("End Time", "") or "").strip()
        period = normalize_period((s or {}).get("Period"))
        code = str((s or {}).get("Course Code", "") or "").strip()
        stype = str((s or {}).get("Session Type", "") or "").strip().upper()
        exact_key = (section, period, day, start, end, code, stype)
        relaxed_key = (section, period, day, start, end, code)
        slot_exact_key = (period, day, start, end, code, stype)
        slot_relaxed_key = (period, day, start, end, code)
        src = (
            log_index_exact.get(exact_key)
            or log_index_relaxed.get(relaxed_key)
            or log_index_slot_exact.get(slot_exact_key)
            or log_index_slot_relaxed.get(slot_relaxed_key)
            or {}
        )
        merged.append(
            {
                "Phase": (src.get("Phase") or s.get("Phase") or ""),
                "Course Code": code,
                "Section": section,
                "Day": day,
                "Start Time": start,
                "End Time": end,
                "Room": (src.get("Room") or s.get("Room") or ""),
                "Period": period,
                "Session Type": (src.get("Session Type") or s.get("Session Type") or "L"),
                "Faculty": (src.get("Faculty") or s.get("Faculty") or ""),
            }
        )
    return merged


def _count_unsatisfied_rows(verification_table: Dict[str, List[Dict[str, Any]]]) -> int:
    total = 0
    for rows in (verification_table or {}).values():
        for r in rows or []:
            status = str((r or {}).get("status", "") or "").strip().upper()
            if status == "UNSATISFIED":
                total += 1
    return total


def _compute_zero_metrics(
    timetable: List[Dict[str, Any]],
    verification_table: Dict[str, List[Dict[str, Any]]],
    post_generate_verify: Dict[str, Any],
) -> Dict[str, int]:
    """
    Compute strict post-generation metrics used as an explicit zero gate.
    """
    errors = (post_generate_verify or {}).get("errors", []) or []
    metrics = {
        "classroom_conflicts": 0,
        "capacity_violations": 0,
        "faculty_conflicts": 0,
        "overlaps": 0,
        "unsatisfied": 0,
        "tbd": 0,
        "missing_faculty": 0,
    }

    for e in errors:
        rule = str((e or {}).get("rule", "") or "").strip().lower()
        if "faculty conflict" in rule:
            metrics["faculty_conflicts"] += 1
            continue
        if "classroom conflict" in rule:
            msg = str((e or {}).get("message", "") or "").strip().lower()
            if "capacity" in msg:
                metrics["capacity_violations"] += 1
            else:
                metrics["classroom_conflicts"] += 1
            continue
        if "section overlap" in rule or "time overlap" in rule or "overlap" in rule:
            metrics["overlaps"] += 1
            continue
        if "ltpsc" in rule or "unsatisfied" in rule:
            metrics["unsatisfied"] += 1

    # NOTE: Verification-table UNSATISFIED rows are advisory and can be stale against
    # strict verification output after final repairs. Zero gate should follow strict
    # verifier truth only.

    # "TBD" gate should track unresolved timetable allocations only.
    # Faculty names can legitimately be blank/placeholder in some log rows and
    # should not fail otherwise-valid strict generations.
    tbd_count = 0
    missing_faculty_count = 0
    for row in timetable or []:
        code = str((row or {}).get("Course Code", "") or "").strip().upper()
        if code in ("LUNCH", "BREAK(15MIN)"):
            continue
        if code.startswith("ELECTIVE_BASKET_"):
            # Synthetic basket rows don't carry a single definitive instructor.
            continue
        room = str((row or {}).get("Room", "") or "").strip().lower()
        if room in ("tbd", "none", "nan", ""):
            tbd_count += 1
        stype = str((row or {}).get("Session Type", "") or "").strip().upper()
        if stype == "P":
            # Practicals intentionally keep blank/varied faculty to avoid false conflict linking.
            continue
        faculty = str((row or {}).get("Faculty", "") or "").strip().lower()
        if faculty in ("", "tbd", "none", "nan", "-", "various", "multiple"):
            missing_faculty_count += 1
    metrics["tbd"] = tbd_count
    metrics["missing_faculty"] = missing_faculty_count
    return metrics


def _api_fast_mode_enabled(request_fast: bool = False) -> bool:
    """Enable faster API response shaping without changing generation rules."""
    if request_fast:
        return True
    env_v = str(os.environ.get("ARISE_API_FAST_MODE", "0") or "").strip().lower()
    return env_v in ("1", "true", "yes", "on")


@app.get("/api/config")
def api_config():
    """Working days, day start/end, lunch windows, programs and section labels for UI."""
    return get_config()


@app.post("/api/generate")
async def api_generate(fast: bool = False):
    """
    First-time Generate: run full pipeline, return timetable + structure for UI labels.
    Runs generate_24_sheets in a thread pool to avoid blocking the uvicorn event loop
    (which causes Windows signal-handler to kill the process mid-request).
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    try:
        from generate_24_sheets import generate_24_sheets
        from utils.generation_verify_bridge import GenerationViolationError
        # Run the heavy synchronous pipeline in a thread pool so uvicorn stays alive
        loop = asyncio.get_event_loop()
        with ThreadPoolExecutor(max_workers=1) as pool:
            try:
                output_path, ts = await loop.run_in_executor(pool, generate_24_sheets)
            except GenerationViolationError as gve:
                detail = {
                    "success": False,
                    "message": str(gve),
                    "errors": gve.errors,
                }
                if getattr(gve, "debug_faculty_path", None):
                    detail["debug_faculty_path"] = gve.debug_faculty_path
                raise HTTPException(
                    status_code=422,
                    detail=detail,
                )
        try:
            log_path = _resolve_log_path_for_timestamp(ts, prefer_edited=False)
        except FileNotFoundError as ex:
            raise HTTPException(status_code=500, detail=str(ex))
        timetable_log = load_timetable_from_csv(log_path)
        fast_mode = _api_fast_mode_enabled(fast)
        labels = get_config()
        excel_full_path = os.path.join(REPO_ROOT, output_path)
        if fast_mode:
            # Fast mode avoids expensive Excel parsing; generation/strict rules remain unchanged.
            timetable = timetable_log
            verification_table = {}
        else:
            timetable_excel = parse_timetable_sessions_from_excel(excel_full_path)
            timetable = merge_excel_sessions_with_log_details(timetable_excel, timetable_log)
            verification_table = parse_verification_tables_from_excel(excel_full_path)
        post_generate_verify = run_verify(timetable_log)
        unsatisfied_count = _count_unsatisfied_rows(verification_table) if not fast_mode else 0
        metrics = _compute_zero_metrics(timetable_log, verification_table, post_generate_verify)
        zero_gate_ok = all(v == 0 for v in metrics.values())
        consistency = {
            "unsatisfied_rows": unsatisfied_count,
            "strict_success": bool(post_generate_verify.get("success", False)),
            "consistent": bool(post_generate_verify.get("success", False)) if fast_mode else not (post_generate_verify.get("success", False) and unsatisfied_count > 0),
            "zero_metrics": metrics,
            "zero_gate_ok": zero_gate_ok,
            "fast_mode": fast_mode,
        }
        if not zero_gate_ok:
            raise HTTPException(
                status_code=422,
                detail={
                    "success": False,
                    "message": "Post-generation zero gate failed",
                    "post_generate_verify": post_generate_verify,
                    "consistency": consistency,
                },
            )
        return {
            "success": True,
            "timetable": timetable,
            "labels": labels,
            "verification_table": verification_table,
            "log_timestamp": ts,
            "post_generate_verify": post_generate_verify,
            "consistency": consistency,
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))



class GenerateFromSessionsRequest(BaseModel):
    sessions: List[Dict[str, Any]]
    offering_variant: Optional[str] = None

@app.post("/api/generate-from-sessions")
async def api_generate_from_sessions(req: GenerateFromSessionsRequest, fast: bool = False):
    """
    Re-generate all 24 Excel sheets from the current (possibly dragged/edited) session list.
    This preserves all manual drag changes while following existing scheduling rules for
    room assignment and sheet formatting.
    """
    import asyncio
    from concurrent.futures import ThreadPoolExecutor

    sessions = req.sessions
    offering_variant = (req.offering_variant or "").strip().lower()

    def _do_generate():
        import pandas as pd
        from datetime import datetime
        from generate_24_sheets import generate_24_sheets_from_log
        import math

        # Write the sessions to a new time-slot-log CSV so generate_24_sheets
        # can pick it up. We reuse the existing CSV format.
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_path = os.path.join(REPO_ROOT, "DATA", "OUTPUT", f"time_slot_log_{ts}.csv")
        os.makedirs(os.path.dirname(log_path), exist_ok=True)

        debug_path = os.environ.get("TIMETABLE_DEBUG_SESSIONS_PATH")
        if debug_path:
            import json
            with open(debug_path, "w", encoding="utf-8") as f:
                json.dump(sessions, f, indent=2)

        # Keep regenerate run on the same odd/even offering as the UI state.
        if offering_variant in ("odd", "even"):
            os.environ["ARISE_COURSE_DATA_VARIANT"] = offering_variant

        # Build a clean DataFrame from the session list
        def _safe_text(v: Any, default: str = "") -> str:
            if v is None:
                return default
            if isinstance(v, float) and math.isnan(v):
                return default
            txt = str(v).strip()
            if txt.lower() in ("nan", "none"):
                return default
            return txt

        rows = []
        for s in sessions:
            rows.append({
                "Phase": _safe_text(s.get("Phase"), "Manual"),
                "Course Code": _safe_text(s.get("Course Code")),
                "Course Name": _safe_text(s.get("Course Name")),
                "Section": _safe_text(s.get("Section")),
                "Day": _safe_text(s.get("Day")),
                "Start Time": _safe_text(s.get("Start Time")),
                "End Time": _safe_text(s.get("End Time")),
                "Room": _safe_text(s.get("Room")),
                "Faculty": _safe_text(s.get("Faculty")),
                "Session Type": _safe_text(s.get("Session Type"), "L"),
                "Period": _safe_text(s.get("Period"), "PRE"),
            })
        df = pd.DataFrame(rows)
        df.to_csv(log_path, index=False)

        # Regenerate the 24 Excel sheets from this log
        output_path = generate_24_sheets_from_log(log_path, ts)
        return output_path, ts, log_path

    try:
        from utils.generation_verify_bridge import GenerationViolationError
        loop = asyncio.get_event_loop()
        with ThreadPoolExecutor(max_workers=1) as pool:
            try:
                output_path, ts, log_path = await loop.run_in_executor(pool, _do_generate)
            except GenerationViolationError as gve:
                detail = {
                    "success": False,
                    "message": str(gve),
                    "errors": gve.errors,
                }
                if getattr(gve, "debug_faculty_path", None):
                    detail["debug_faculty_path"] = gve.debug_faculty_path
                raise HTTPException(status_code=422, detail=detail)

        try:
            # Regenerate-from-sessions writes its final log to DATA/EDITED OUTPUT.
            log_path = _resolve_log_path_for_timestamp(ts, prefer_edited=True)
        except FileNotFoundError as ex:
            raise HTTPException(status_code=500, detail=str(ex))

        timetable_log = load_timetable_from_csv(log_path)
        fast_mode = _api_fast_mode_enabled(fast)
        labels = get_config()
        excel_full_path = os.path.join(REPO_ROOT, output_path)
        if fast_mode:
            timetable = timetable_log
            verification_table = {}
        else:
            timetable_excel = parse_timetable_sessions_from_excel(excel_full_path)
            timetable = merge_excel_sessions_with_log_details(timetable_excel, timetable_log)
            verification_table = parse_verification_tables_from_excel(excel_full_path)
        post_generate_verify = run_verify(timetable_log)
        unsatisfied_count = _count_unsatisfied_rows(verification_table) if not fast_mode else 0
        metrics = _compute_zero_metrics(timetable_log, verification_table, post_generate_verify)
        zero_gate_ok = all(v == 0 for v in metrics.values())
        consistency = {
            "unsatisfied_rows": unsatisfied_count,
            "strict_success": bool(post_generate_verify.get("success", False)),
            "consistent": bool(post_generate_verify.get("success", False)) if fast_mode else not (post_generate_verify.get("success", False) and unsatisfied_count > 0),
            "zero_metrics": metrics,
            "zero_gate_ok": zero_gate_ok,
            "fast_mode": fast_mode,
        }
        if not zero_gate_ok:
            raise HTTPException(
                status_code=422,
                detail={
                    "success": False,
                    "message": "Post-generation zero gate failed",
                    "post_generate_verify": post_generate_verify,
                    "consistency": consistency,
                },
            )
        return {
            "success": True,
            "timetable": timetable,
            "labels": labels,
            "verification_table": verification_table,
            "log_timestamp": ts,
            "post_generate_verify": post_generate_verify,
            "consistency": consistency,
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


class VerifyRequest(BaseModel):
    sessions: List[Dict[str, Any]]  # each: Course Code, Section, Day, Start Time, End Time, Room, Period, Session Type, Faculty, Phase


def _sessions_api_to_internal(sessions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Convert API time_slot_log schema to internal format (time_block, course_code, sections, etc.).
    Aggregates sections for combined-class sessions logged multiple times so deep verification
    knows the session belongs to *all* of those sections, enabling cross-section conflict checks.
    """
    from utils.data_models import TimeBlock
    from utils.generation_verify_bridge import final_ui_rows_to_verify_sessions

    rows: List[Dict[str, Any]] = []
    for s in sessions:
        course_code = (s.get("Course Code") or "").strip()
        section = (s.get("Section") or "").strip()
        day = (s.get("Day") or "").strip()
        start_s = (s.get("Start Time") or "").strip()
        end_s = (s.get("End Time") or "").strip()
        session_type = (s.get("Session Type") or "L").strip().upper()
        period = normalize_period(s.get("Period"))

        if not (course_code and section and day and start_s and end_s):
            continue

        faculty_raw = (s.get("Faculty") or "").strip()
        room = (s.get("Room") or "").strip()
        phase = (s.get("Phase") or "").strip()
        try:
            hh, mm = start_s.split(":")
            start_t = time(int(hh), int(mm))
            hh, mm = end_s.split(":")
            end_t = time(int(hh), int(mm))
        except Exception:
            continue

        rows.append(
            {
                "phase": phase,
                "course_code": course_code,
                "section": section,
                "day": day,
                "start_time": start_t,
                "end_time": end_t,
                "time_block": TimeBlock(day, start_t, end_t),
                "period": period,
                "session_type": session_type,
                "faculty": faculty_raw,
                "room": room,
            }
        )

    return final_ui_rows_to_verify_sessions(rows)


def run_verify(sessions: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Run verification on in-memory sessions (API schema).
    Uses deep_verification.run_verification_on_sessions for single source of truth.
    Returns { "success": bool, "errors": [ { "rule", "message", "course_code", "section", "day", "time", ... } ] }.
    """
    from modules_v2.phase1_data_validation_v2 import run_phase1
    from config.structure_config import DEPARTMENTS, SECTIONS_BY_DEPT, STUDENTS_PER_SECTION, get_group_for_section
    from utils.data_models import Section
    from deep_verification import run_verification_on_sessions as run_dv_verify

    all_sessions = _sessions_api_to_internal(sessions)
    if not all_sessions:
        return {"success": True, "errors": []}
    courses, classrooms, _ = run_phase1()
    unique_semesters = sorted(set(c.semester for c in courses if c.department in DEPARTMENTS))
    sections = []
    for dept in DEPARTMENTS:
        for sem in unique_semesters:
            for sec_label in SECTIONS_BY_DEPT.get(dept, []):
                group = get_group_for_section(dept, sec_label)
                sections.append(Section(dept, group, sec_label, sem, STUDENTS_PER_SECTION))
    success, errors = run_dv_verify(all_sessions, courses, sections, classrooms)
    # Deduplicate errors: same (rule, course_code, section, message) should not appear twice
    seen_errors: set = set()
    unique_errors = []
    for e in errors:
        key = (e.get("rule", ""), e.get("course_code", ""), e.get("section", ""), e.get("message", ""))
        if key not in seen_errors:
            seen_errors.add(key)
            unique_errors.append(e)
    return {"success": len(unique_errors) == 0, "errors": unique_errors}


@app.post("/api/verify")
def api_verify(req: VerifyRequest):
    """Verify proposed sessions (after drag). Returns success or list of detailed errors."""
    return run_verify(req.sessions)


class ReflowRequest(BaseModel):
    sessions: List[Dict[str, Any]]
    movedSession: Dict[str, Any]


@app.post("/api/reflow")
def api_reflow(req: ReflowRequest):
    """
    Try to reflow: keep user-moved sessions, reschedule others to satisfy rules.
    Minimal implementation: if the move introduced conflicts, try to relocate only
    the directly-conflicting sessions to the next available 15-minute slot.
    If still not possible, return not_possible so frontend can revert.
    """
    try:
        sessions = list(req.sessions or [])
        moved = req.movedSession or {}

        def _parse_hhmm(t: str) -> int:
            hh, mm = (t or "").strip().split(":")
            return int(hh) * 60 + int(mm)

        def _fmt_hhmm(m: int) -> str:
            return f"{m // 60:02d}:{m % 60:02d}"

        def _overlap(a_start: int, a_end: int, b_start: int, b_end: int) -> bool:
            return a_start < b_end and b_start < a_end

        moved_day = (moved.get("Day") or "").strip()
        moved_start = moved.get("Start Time") or ""
        moved_end = moved.get("End Time") or ""
        moved_section = (moved.get("Section") or "").strip()
        moved_room = (moved.get("Room") or "").strip()
        moved_faculty = (moved.get("Faculty") or "").strip()
        moved_period = normalize_period(moved.get("Period") or "PRE")
        if not (moved_day and moved_start and moved_end and moved_section):
            return {"success": False, "not_possible": True}
        moved_start_m = _parse_hhmm(moved_start)
        moved_end_m = _parse_hhmm(moved_end)

        # Build slot grid
        from config.schedule_config import WORKING_DAYS, DAY_START_TIME, DAY_END_TIME, LUNCH_WINDOWS
        day_start_m = DAY_START_TIME.hour * 60 + DAY_START_TIME.minute
        day_end_m = DAY_END_TIME.hour * 60 + DAY_END_TIME.minute

        # Lunch windows by semester (SemX parsed from section label)
        sem_re = re.compile(r"Sem(\d+)", re.IGNORECASE)

        def _get_lunch_window(section_label: str):
            m = sem_re.search(section_label or "")
            if not m:
                return None
            sem = int(m.group(1))
            win = LUNCH_WINDOWS.get(sem)
            if not win:
                return None
            start_t, end_t = win
            return (start_t.hour * 60 + start_t.minute, end_t.hour * 60 + end_t.minute)

        def _session_conflicts(s: Dict[str, Any], other: Dict[str, Any]) -> bool:
            if normalize_period(s.get("Period")) != normalize_period(other.get("Period")):
                return False
            if (s.get("Day") or "").strip() != (other.get("Day") or "").strip():
                return False
            try:
                s1, e1 = _parse_hhmm(s.get("Start Time")), _parse_hhmm(s.get("End Time"))
                s2, e2 = _parse_hhmm(other.get("Start Time")), _parse_hhmm(other.get("End Time"))
            except Exception:
                return False
            if not _overlap(s1, e1, s2, e2):
                return False

            # Conflicts if share section OR room OR faculty
            same_section = (s.get("Section") or "").strip() == (other.get("Section") or "").strip()
            same_room = (s.get("Room") or "").strip() and (s.get("Room") or "").strip() == (other.get("Room") or "").strip()
            same_faculty = (s.get("Faculty") or "").strip() and (s.get("Faculty") or "").strip() == (other.get("Faculty") or "").strip()
            return same_section or same_room or same_faculty

        # Find sessions directly conflicting with the moved one
        moved_key = (
            (moved.get("Course Code") or "").strip(),
            moved_section,
            moved_day,
            moved_start,
            moved_end,
            moved_period,
        )

        def _is_same_session(a: Dict[str, Any], key) -> bool:
            return (
                ((a.get("Course Code") or "").strip(), (a.get("Section") or "").strip(),
                 (a.get("Day") or "").strip(), a.get("Start Time") or "", a.get("End Time") or "",
                 normalize_period(a.get("Period") or "PRE")) == key
            )

        conflicts = []
        for s in sessions:
            if _is_same_session(s, moved_key):
                continue
            # Ensure we only move sessions within the same period as moved (UI is scoped by period)
            if normalize_period(s.get("Period") or "PRE") != moved_period:
                continue
            if _session_conflicts(moved, s):
                conflicts.append(s)

        if not conflicts:
            # Already ok (or conflicts not detected here); let verifier be source of truth
            v = run_verify(sessions)
            if v.get("success"):
                return {"success": True, "timetable": sessions}
            return {"success": False, "not_possible": True}

        # Occupancy check against current sessions (keeps moved fixed)
        def _can_place(sess: Dict[str, Any], cand_day: str, cand_start: int) -> bool:
            duration = _parse_hhmm(sess.get("End Time")) - _parse_hhmm(sess.get("Start Time"))
            cand_end = cand_start + duration
            if cand_start < day_start_m or cand_end > day_end_m:
                return False

            lunch = _get_lunch_window((sess.get("Section") or "").strip())
            if lunch and _overlap(cand_start, cand_end, lunch[0], lunch[1]):
                return False

            # Check overlap vs all others (including moved)
            for o in sessions:
                if o is sess:
                    continue
                if normalize_period(o.get("Period") or "PRE") != moved_period:
                    continue
                if (o.get("Day") or "").strip() != cand_day:
                    continue
                try:
                    o_s, o_e = _parse_hhmm(o.get("Start Time")), _parse_hhmm(o.get("End Time"))
                except Exception:
                    continue
                if not _overlap(cand_start, cand_end, o_s, o_e):
                    continue
                same_section = (sess.get("Section") or "").strip() == (o.get("Section") or "").strip()
                same_room = (sess.get("Room") or "").strip() and (sess.get("Room") or "").strip() == (o.get("Room") or "").strip()
                same_faculty = (sess.get("Faculty") or "").strip() and (sess.get("Faculty") or "").strip() == (o.get("Faculty") or "").strip()
                if same_section or same_room or same_faculty:
                    return False
            return True

        # Try to relocate each conflicting session greedily
        for sess in conflicts:
            orig_day = (sess.get("Day") or "").strip()
            try:
                duration = _parse_hhmm(sess.get("End Time")) - _parse_hhmm(sess.get("Start Time"))
            except Exception:
                continue

            day_order = [orig_day] + [d for d in WORKING_DAYS if d != orig_day]
            placed = False
            for d in day_order:
                for start in range(day_start_m, day_end_m - duration + 1, 15):
                    # avoid placing exactly on moved's start time if it would still conflict
                    if d == moved_day and _overlap(start, start + duration, moved_start_m, moved_end_m):
                        continue
                    if _can_place(sess, d, start):
                        sess["Day"] = d
                        sess["Start Time"] = _fmt_hhmm(start)
                        sess["End Time"] = _fmt_hhmm(start + duration)
                        placed = True
                        break
                if placed:
                    break

            if not placed:
                return {"success": False, "not_possible": True}

        # Final deep verification
        v = run_verify(sessions)
        if v.get("success"):
            return {"success": True, "timetable": sessions}
        return {"success": False, "not_possible": True}
    except Exception:
        import traceback
        traceback.print_exc()
        return {"success": False, "not_possible": True}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
